import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import json
import os
import io
import re
import webbrowser
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MONTHS = ["", "January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


def _hex_tint(hex_color, factor):
    """Blend hex_color toward white by factor (0.0 = original, 1.0 = white)."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"#{r:02x}{g:02x}{b:02x}"


def _today_value():
    """Return today's date as a float date value (CE)."""
    import datetime
    t = datetime.date.today()
    return t.year + (t.month - 1) / 12 + (t.day - 1) / 365


def _date_value(num_str, unit, month=0, day=0):
    """Return a float for sorting. More negative = further in the past."""
    if unit == "Present":
        return _today_value()
    try:
        num = float(num_str)
    except (ValueError, TypeError):
        return None
    m = int(month) if month else 0
    d = int(day) if day else 0
    frac = ((m - 1) / 12 + (d - 1) / 365) if m > 0 else 0.0
    if unit == "CE":
        return num + frac
    elif unit == "BCE":
        return -(num + frac)
    elif unit == "MYA":
        return -(num * 1_000_000)
    elif unit == "BYA":
        return -(num * 1_000_000_000)
    return num


def _date_display(num_str, unit, month=0, day=0):
    """Return a human-readable date string."""
    if unit == "Present":
        return "Present"
    try:
        num = float(num_str)
    except (ValueError, TypeError):
        return ""
    if not str(num_str).strip():
        return ""
    if unit in ("MYA", "BYA"):
        return f"{num:g} {unit}"
    m = int(month) if month else 0
    d = int(day) if day else 0
    year_int = int(num)
    if m > 0 and d > 0:
        suffix = f" {unit}" if unit == "BCE" else ""
        return f"{m:02}/{d:02}/{year_int}{suffix}"
    if m > 0:
        return f"{m:02}/{year_int} {unit}"
    return f"{year_int} {unit}"


def _parse_date_display(text):
    """Parse an exported date display string back into (num_str, unit, month, day).
    Returns (None, None, None, None) if empty or unparseable."""
    if not text or not str(text).strip():
        return None, None, None, None
    text = str(text).strip()
    if text.lower() == "present":
        return None, "Present", None, None

    for unit in ("MYA", "BYA"):
        if text.endswith(unit):
            num_str = text[:-len(unit)].strip()
            try:
                float(num_str)
                return num_str, unit, None, None
            except ValueError:
                pass

    # mm/dd/yyyy or mm/dd/yyyy BCE
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d+)\s*(BCE)?$', text)
    if m:
        unit = "BCE" if m.group(4) else "CE"
        return m.group(3), unit, int(m.group(1)), int(m.group(2))

    # mm/yyyy CE|BCE
    m = re.match(r'^(\d{1,2})/(\d+)\s+(CE|BCE)$', text)
    if m:
        return m.group(2), m.group(3), int(m.group(1)), None

    # yyyy CE|BCE
    m = re.match(r'^(\d+)\s+(CE|BCE)$', text)
    if m:
        return m.group(1), m.group(2), None, None

    return None, None, None, None


def _px_per_year_to_display(px_per_year):
    """Convert px_per_year to a human-readable (value, unit) pair."""
    if px_per_year >= 1:
        return px_per_year, "px/yr"
    yr_per_px = 1.0 / px_per_year
    if yr_per_px < 1_000:
        return yr_per_px, "yr/px"
    elif yr_per_px < 1_000_000:
        return yr_per_px / 1_000, "kyr/px"
    elif yr_per_px < 1_000_000_000:
        return yr_per_px / 1_000_000, "Myr/px"
    else:
        return yr_per_px / 1_000_000_000, "Byr/px"


def _display_to_px_per_year(value, unit):
    """Convert a display value + unit back to px_per_year."""
    if unit == "px/yr":
        return value
    elif unit == "yr/px":
        return 1.0 / value
    elif unit == "kyr/px":
        return 1.0 / (value * 1_000)
    elif unit == "Myr/px":
        return 1.0 / (value * 1_000_000)
    else:  # Byr/px
        return 1.0 / (value * 1_000_000_000)


def _date_val_to_display(val):
    """Convert a float date value to a human-readable string (e.g. '1066 CE', '65 MYA', '4.5 BYA')."""
    if val is None:
        return ""
    if abs(val) >= 1_000_000_000:
        return f"{-val / 1e9:g} BYA"
    if abs(val) >= 1_000_000:
        return f"{-val / 1e6:g} MYA"
    year = int(val)
    if year < 0:
        return f"{-year} BCE"
    return f"{year} CE"


def _date_val_to_num_unit(val):
    """Split a float date value into (num_str, unit) for use in a num+unit widget pair."""
    if val is None:
        return "", "CE"
    if abs(val) >= 1_000_000_000:
        return f"{-val / 1e9:g}", "BYA"
    if abs(val) >= 1_000_000:
        return f"{-val / 1e6:g}", "MYA"
    year = int(abs(val))
    if val < 0:
        return str(year), "BCE"
    return str(year), "CE"


def _date_val_to_components(val):
    """Split a float date value into (num_str, unit, month_name, day_str).

    Month and day are recovered from the fractional part of the float encoding
    (year + (month-1)/12 + (day-1)/365).  Returns empty strings when not encoded.
    """
    if val is None:
        return "", "CE", "", ""
    if abs(val) >= 1_000_000_000:
        return f"{-val / 1e9:g}", "BYA", "", ""
    if abs(val) >= 1_000_000:
        return f"{-val / 1e6:g}", "MYA", "", ""
    abs_val = abs(val)
    unit    = "BCE" if val < 0 else "CE"
    year    = int(abs_val)
    frac    = abs_val - year
    if frac < 0.0005:          # no month/day encoded
        return str(year), unit, "", ""
    month_idx = round(frac * 12)   # 0-based
    if not (0 <= month_idx <= 11):
        return str(year), unit, "", ""
    month_name = MONTHS[month_idx + 1]
    day_frac   = frac - month_idx / 12.0
    day        = round(day_frac * 365) + 1
    day_str    = str(day) if 1 <= day <= 31 else ""
    return str(year), unit, month_name, day_str


class TimelineDB:

    def __init__(self, db_file="timeline.db"):
        self.db_file  = db_file
        self.config_file = db_file.replace(".db", "_config.json")
        self.events = []
        self.categories = []
        self.active_timeline_id = None
        self._init_db()

    def load_config(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as f:
                return json.load(f)
        return {}

    def save_config(self, data):
        cfg = self.load_config()
        cfg.update(data)
        with open(self.config_file, "w") as f:
            json.dump(cfg, f, indent=2)

    def last_timeline_id(self):
        return self.load_config().get("last_timeline_id")

    # ── timeline management ───────────────────────────────────────────────────

    def load_timelines(self):
        """Return list of (TimelineID, Title) tuples."""
        with sqlite3.connect(self.db_file) as conn:
            return conn.execute(
                "SELECT TimelineID, Title FROM Timeline ORDER BY Title"
            ).fetchall()

    def save_timeline_view_state(self, timeline_id, px_per_year, freeze_scale):
        """Save view state. Pass px_per_year=None to clear the saved scale."""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET px_per_year=?, freeze_scale=? WHERE TimelineID=?",
                (px_per_year, 1 if freeze_scale else 0, timeline_id)
            )

    def load_timeline_view_state(self, timeline_id):
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT px_per_year, freeze_scale FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        if row and row[0] is not None:
            return {"px_per_year": row[0], "freeze_scale": bool(row[1])}
        return None

    def load_timeline_ruler(self, timeline_id):
        """Return (ruler_min, ruler_max, ruler_max_is_present).

        ruler_min / ruler_max are floats or None.
        ruler_max_is_present=True means the end is always today's date.
        """
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT ruler_min, ruler_max, ruler_max_is_present FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        if row:
            return row[0], row[1], bool(row[2])
        return None, None, False

    def load_timeline_icons(self, timeline_id):
        """Return (icon_short, icon_long) for the timeline.
        Defaults: icon_short='Diamond', icon_long='Box' when not set."""
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT default_icon_short, default_icon_long FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        if row:
            return row[0] or "Diamond", row[1] or "Box"
        return "Diamond", "Box"

    def save_timeline_icons(self, timeline_id, icon_short, icon_long):
        """Save default icon styles. Pass None/empty to revert to defaults."""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET default_icon_short=?, default_icon_long=? WHERE TimelineID=?",
                (icon_short or None, icon_long or None, timeline_id)
            )

    def load_timeline_cat_header_style(self, timeline_id):
        """Return cat_header_style for the timeline. Default: 'Left'."""
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT cat_header_style FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        return (row[0] or "Left") if row else "Left"

    def save_timeline_cat_header_style(self, timeline_id, style):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET cat_header_style=? WHERE TimelineID=?",
                (style, timeline_id)
            )

    def load_timeline_cat_header_title_pos(self, timeline_id):
        """Return cat_header_title_pos for the timeline. Default: 'Center (View)'."""
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT cat_header_title_pos FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        return (row[0] or "Center (View)") if row else "Center (View)"

    def save_timeline_cat_header_title_pos(self, timeline_id, pos):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET cat_header_title_pos=? WHERE TimelineID=?",
                (pos, timeline_id)
            )

    def load_timeline_canvas_bg(self, timeline_id):
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT canvas_bg_color FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        return (row[0] or "") if row else ""

    def save_timeline_canvas_bg(self, timeline_id, color):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET canvas_bg_color=? WHERE TimelineID=?",
                (color or None, timeline_id)
            )

    def load_timeline_bg_image(self, timeline_id):
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT bg_image, bg_image_name, bg_image_pos FROM Timeline WHERE TimelineID=?",
                (timeline_id,)
            ).fetchone()
        if row:
            return {"image": row[0], "name": row[1] or "", "pos": row[2] or "Top"}
        return {"image": None, "name": "", "pos": "Top"}

    def save_timeline_bg_image(self, timeline_id, image_blob, image_name, image_pos):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET bg_image=?, bg_image_name=?, bg_image_pos=? "
                "WHERE TimelineID=?",
                (image_blob, image_name or None, image_pos or "Top", timeline_id)
            )

    def save_timeline_ruler(self, timeline_id, ruler_min, ruler_max,
                            ruler_max_is_present=False):
        """Save ruler date range. Pass None for a bound to clear it (auto-calculate)."""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE Timeline SET ruler_min=?, ruler_max=?, ruler_max_is_present=? "
                "WHERE TimelineID=?",
                (ruler_min, ruler_max, 1 if ruler_max_is_present else 0, timeline_id)
            )

    def load_timeline_breaks(self, timeline_id):
        """Return list of {id, start, end} dicts sorted by start."""
        with sqlite3.connect(self.db_file) as conn:
            rows = conn.execute(
                "SELECT id, break_start, break_end FROM TimelineBreak "
                "WHERE timeline_id=? ORDER BY break_start",
                (timeline_id,)
            ).fetchall()
        return [{"id": r[0], "start": r[1], "end": r[2]} for r in rows]

    def add_timeline_break(self, timeline_id, start, end):
        """Insert a break and return its id."""
        with sqlite3.connect(self.db_file) as conn:
            cur = conn.execute(
                "INSERT INTO TimelineBreak (timeline_id, break_start, break_end) VALUES (?,?,?)",
                (timeline_id, start, end)
            )
            return cur.lastrowid

    def delete_timeline_break(self, break_id):
        """Remove a break by id."""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("DELETE FROM TimelineBreak WHERE id=?", (break_id,))

    def update_timeline_break(self, break_id, start, end):
        """Update the date range of an existing break."""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute(
                "UPDATE TimelineBreak SET break_start=?, break_end=? WHERE id=?",
                (start, end, break_id)
            )

    def add_timeline(self, title):
        """Insert a new timeline and return its ID."""
        with sqlite3.connect(self.db_file) as conn:
            cur = conn.execute("INSERT INTO Timeline (Title) VALUES (?)", (title,))
            return cur.lastrowid

    def rename_timeline(self, timeline_id, new_title):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("UPDATE Timeline SET Title=? WHERE TimelineID=?",
                         (new_title, timeline_id))

    def count_events(self, timeline_id):
        with sqlite3.connect(self.db_file) as conn:
            return conn.execute(
                "SELECT COUNT(*) FROM events WHERE timelineid=?", (timeline_id,)
            ).fetchone()[0]

    def delete_timeline(self, timeline_id):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("DELETE FROM events   WHERE timelineid=?", (timeline_id,))
            conn.execute("DELETE FROM Category WHERE timelineid=?", (timeline_id,))
            conn.execute("DELETE FROM Timeline WHERE TimelineID=?", (timeline_id,))

    # ── categories ────────────────────────────────────────────────────────────

    def load_categories(self):
        """Load all categories as a flat list of dicts and build self.cat_tree.

        self.categories     — flat list of title strings (depth-first, for combos)
        self.cat_nodes      — list of dicts: {id, title, parent_id, sort_order, depth, path}
        self.cat_by_id      — dict: CategoryID -> node dict
        self.cat_tree       — nested: list of root nodes, each with 'children' key
        """
        with sqlite3.connect(self.db_file) as conn:
            rows = conn.execute(
                "SELECT CategoryID, Title, parent_id, sort_order, hidden, color, row_bg_color, show_row_guide, "
                "cat_image, cat_image_name, cat_image_pos, cat_pad_top, cat_pad_bottom "
                "FROM Category WHERE timelineid=? ORDER BY sort_order, CategoryID",
                (self.active_timeline_id,)
            ).fetchall()

        # Build lookup
        by_id = {r[0]: {"id": r[0], "title": r[1], "parent_id": r[2],
                         "sort_order": r[3], "hidden": bool(r[4]),
                         "color": r[5], "row_bg_color": r[6],
                         "show_row_guide": (r[7] != 0),
                         "cat_image": r[8], "cat_image_name": r[9],
                         "cat_image_pos": r[10] or "Row",
                         "cat_pad_top":    int(r[11] or 0),
                         "cat_pad_bottom": int(r[12] or 0),
                         "children": []} for r in rows}

        roots = []
        for node in by_id.values():
            pid = node["parent_id"]
            if pid and pid in by_id:
                by_id[pid]["children"].append(node)
            else:
                roots.append(node)

        # Sort children by sort_order
        def _sort(nodes):
            nodes.sort(key=lambda n: (n["sort_order"] or 0, n["id"]))
            for n in nodes:
                _sort(n["children"])

        _sort(roots)

        # Flatten depth-first for self.categories / self.cat_nodes
        flat_nodes = []

        def _flatten(nodes, depth=0, path=""):
            for n in nodes:
                n["depth"] = depth
                n["path"]  = (path + " > " + n["title"]) if path else n["title"]
                flat_nodes.append(n)
                _flatten(n["children"], depth + 1, n["path"])

        _flatten(roots)

        self.cat_tree   = roots
        self.cat_nodes  = flat_nodes
        self.cat_by_id  = by_id
        self.categories = [n["title"] for n in flat_nodes]   # kept for compat

    def _category_id(self, conn, category_name):
        row = conn.execute(
            "SELECT CategoryID FROM Category WHERE LOWER(Title)=LOWER(?) AND timelineid=?",
            (category_name, self.active_timeline_id)
        ).fetchone()
        return row[0] if row else None

    def _ensure_category(self, conn, category_name):
        """Return category ID, creating it if missing (always at root level)."""
        cat_id = self._category_id(conn, category_name)
        if cat_id is None:
            cur = conn.execute(
                "INSERT INTO Category (Title, timelineid) VALUES (?, ?)",
                (category_name, self.active_timeline_id)
            )
            cat_id = cur.lastrowid
        return cat_id

    def add_category(self, name, parent_id=None):
        """Add a new category under parent_id (None = root) at end of siblings."""
        with sqlite3.connect(self.db_file) as conn:
            if parent_id:
                max_order = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) FROM Category "
                    "WHERE timelineid=? AND parent_id=?",
                    (self.active_timeline_id, parent_id)
                ).fetchone()[0]
            else:
                max_order = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) FROM Category "
                    "WHERE timelineid=? AND parent_id IS NULL",
                    (self.active_timeline_id,)
                ).fetchone()[0]
            cur = conn.execute(
                "INSERT INTO Category (Title, timelineid, parent_id, sort_order) VALUES (?, ?, ?, ?)",
                (name, self.active_timeline_id, parent_id, max_order + 1)
            )
            return cur.lastrowid

    def rename_category(self, cat_id, new_title):
        """Rename category by ID."""
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("UPDATE Category SET Title=? WHERE CategoryID=?",
                         (new_title, cat_id))
        # Update in-memory events
        for e in self.events:
            if e.get("categoryid") == cat_id:
                e["category"] = new_title

    def delete_category(self, cat_id):
        """Delete category; events and children re-parented one level up (or root)."""
        with sqlite3.connect(self.db_file) as conn:
            row = conn.execute(
                "SELECT parent_id FROM Category WHERE CategoryID=?", (cat_id,)
            ).fetchone()
            if row is None:
                return
            new_parent = row[0]   # None if deleting a root category

            # Re-parent direct children
            conn.execute(
                "UPDATE Category SET parent_id=? WHERE parent_id=?",
                (new_parent, cat_id)
            )
            # Re-parent events
            conn.execute(
                "UPDATE events SET categoryid=? WHERE categoryid=?",
                (new_parent, cat_id)
            )
            conn.execute("DELETE FROM Category WHERE CategoryID=?", (cat_id,))

        # Update in-memory events
        for e in self.events:
            if e.get("categoryid") == cat_id:
                e["categoryid"] = new_parent
                if new_parent and hasattr(self, "cat_by_id") and new_parent in self.cat_by_id:
                    e["category"] = self.cat_by_id[new_parent]["title"]
                else:
                    e["category"] = ""

    def reorder_category(self, cat_id, direction):
        """Move a category up (-1) or down (+1) among its siblings."""
        with sqlite3.connect(self.db_file) as conn:
            node = conn.execute(
                "SELECT CategoryID, parent_id FROM Category WHERE CategoryID=?",
                (cat_id,)
            ).fetchone()
            if node is None:
                return
            pid = node[1]
            if pid:
                siblings = conn.execute(
                    "SELECT CategoryID FROM Category "
                    "WHERE timelineid=? AND parent_id=? ORDER BY sort_order, CategoryID",
                    (self.active_timeline_id, pid)
                ).fetchall()
            else:
                siblings = conn.execute(
                    "SELECT CategoryID FROM Category "
                    "WHERE timelineid=? AND parent_id IS NULL ORDER BY sort_order, CategoryID",
                    (self.active_timeline_id,)
                ).fetchall()
            ids = [r[0] for r in siblings]
            # Normalise sort_order to sequential values so NULLs don't prevent swapping
            for i, sid in enumerate(ids):
                conn.execute("UPDATE Category SET sort_order=? WHERE CategoryID=?",
                             (i, sid))
            idx = next((i for i, sid in enumerate(ids) if sid == cat_id), None)
            if idx is None:
                return
            swap = idx + direction
            if swap < 0 or swap >= len(ids):
                return
            conn.execute("UPDATE Category SET sort_order=? WHERE CategoryID=?",
                         (swap, ids[idx]))
            conn.execute("UPDATE Category SET sort_order=? WHERE CategoryID=?",
                         (idx, ids[swap]))

    # ── schema ────────────────────────────────────────────────────────────────

    def _init_db(self):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Timeline (
                    TimelineID INTEGER PRIMARY KEY AUTOINCREMENT,
                    Title      TEXT NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Category (
                    CategoryID INTEGER PRIMARY KEY AUTOINCREMENT,
                    Title      TEXT NOT NULL,
                    timelineid INTEGER REFERENCES Timeline(TimelineID)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS events (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    title      TEXT NOT NULL,
                    year       INTEGER NOT NULL DEFAULT 0,
                    desc       TEXT,
                    categoryid INTEGER REFERENCES Category(CategoryID),
                    timelineid INTEGER REFERENCES Timeline(TimelineID)
                )
            """)

            # Schema migrations
            tl_cols    = [r[1] for r in conn.execute("PRAGMA table_info(Timeline)")]
            cat_cols   = [r[1] for r in conn.execute("PRAGMA table_info(Category)")]
            event_cols = [r[1] for r in conn.execute("PRAGMA table_info(events)")]

            if "px_per_year" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN px_per_year REAL")
            if "freeze_scale" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN freeze_scale INTEGER DEFAULT 0")
            if "ruler_min" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN ruler_min REAL")
            if "ruler_max" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN ruler_max REAL")
            if "ruler_max_is_present" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN ruler_max_is_present INTEGER DEFAULT 0")
            if "default_icon_short" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN default_icon_short TEXT")
            if "default_icon_long" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN default_icon_long TEXT")
            if "cat_header_style" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN cat_header_style TEXT DEFAULT 'Left'")
            if "cat_header_title_pos" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN cat_header_title_pos TEXT DEFAULT 'Center (View)'")
            if "canvas_bg_color" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN canvas_bg_color TEXT")
            if "bg_image" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN bg_image BLOB")
            if "bg_image_name" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN bg_image_name TEXT")
            if "bg_image_pos" not in tl_cols:
                conn.execute("ALTER TABLE Timeline ADD COLUMN bg_image_pos TEXT DEFAULT 'Top'")

            if "timelineid" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN timelineid INTEGER REFERENCES Timeline(TimelineID)")
            if "parent_id" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN parent_id INTEGER REFERENCES Category(CategoryID)")
            if "sort_order" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN sort_order INTEGER")
                # Seed sort_order from existing insertion order
                conn.execute("""
                    UPDATE Category SET sort_order = CategoryID
                    WHERE sort_order IS NULL
                """)
            if "timelineid" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN timelineid INTEGER REFERENCES Timeline(TimelineID)")
            if "url" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN url TEXT")
            if "image" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN image BLOB")
                conn.execute("ALTER TABLE events ADD COLUMN image_name TEXT")
                conn.execute("ALTER TABLE events ADD COLUMN image_type TEXT")
            if "start_value" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN start_value REAL")
                conn.execute("ALTER TABLE events ADD COLUMN start_display TEXT")
                conn.execute("ALTER TABLE events ADD COLUMN start_unit TEXT")
                conn.execute("ALTER TABLE events ADD COLUMN start_month INTEGER")
                conn.execute("ALTER TABLE events ADD COLUMN start_day INTEGER")
                conn.execute("ALTER TABLE events ADD COLUMN end_value REAL")
                conn.execute("ALTER TABLE events ADD COLUMN end_display TEXT")
                conn.execute("ALTER TABLE events ADD COLUMN end_unit TEXT")
                conn.execute("ALTER TABLE events ADD COLUMN end_month INTEGER")
                conn.execute("ALTER TABLE events ADD COLUMN end_day INTEGER")
                conn.execute("""
                    UPDATE events
                    SET start_value   = CAST(year AS REAL),
                        start_display = CAST(year AS TEXT) || ' CE',
                        start_unit    = 'CE'
                    WHERE year IS NOT NULL
                """)

            if "sort_order" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN sort_order INTEGER")
                conn.execute("UPDATE events SET sort_order = id WHERE sort_order IS NULL")

            if "standalone" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN standalone INTEGER DEFAULT 0")

            if "hidden" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN hidden INTEGER DEFAULT 0")
            if "color" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN color TEXT")
            if "row_bg_color" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN row_bg_color TEXT")
            if "show_row_guide" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN show_row_guide INTEGER DEFAULT 1")
            if "cat_image" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN cat_image BLOB")
            if "cat_image_name" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN cat_image_name TEXT")
            if "cat_image_pos" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN cat_image_pos TEXT DEFAULT 'Row'")
            if "cat_pad_top" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN cat_pad_top INTEGER DEFAULT 0")
            if "cat_pad_bottom" not in cat_cols:
                conn.execute("ALTER TABLE Category ADD COLUMN cat_pad_bottom INTEGER DEFAULT 0")

            if "hidden" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN hidden INTEGER DEFAULT 0")

            if "picture_position" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN picture_position TEXT")
                # Migrate legacy show_picture=1 rows to 'Left of Event'
                if "show_picture" in event_cols:
                    conn.execute(
                        "UPDATE events SET picture_position='Left of Event' WHERE show_picture=1"
                    )

            if "linked_categoryid" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN linked_categoryid INTEGER")

            if "linked_timelineid" not in event_cols:
                conn.execute("ALTER TABLE events ADD COLUMN linked_timelineid INTEGER")

            # Timeline break table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS TimelineBreak (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    timeline_id INTEGER REFERENCES Timeline(TimelineID),
                    break_start REAL NOT NULL,
                    break_end   REAL NOT NULL
                )
            """)

            # Seed "Practice Timeline" and migrate orphaned records
            tl_count = conn.execute("SELECT COUNT(*) FROM Timeline").fetchone()[0]
            if tl_count == 0:
                cur = conn.execute(
                    "INSERT INTO Timeline (Title) VALUES ('Practice Timeline')"
                )
                practice_id = cur.lastrowid
                # Migrate all existing categories and events to Practice Timeline
                conn.execute("UPDATE Category SET timelineid=? WHERE timelineid IS NULL",
                             (practice_id,))
                # Seed default categories if none exist
                cat_count = conn.execute(
                    "SELECT COUNT(*) FROM Category WHERE timelineid=?", (practice_id,)
                ).fetchone()[0]
                if cat_count == 0:
                    for cat in ["War", "Science", "Politics", "Exploration",
                                "Culture", "Religion", "General"]:
                        conn.execute(
                            "INSERT INTO Category (Title, timelineid) VALUES (?, ?)",
                            (cat, practice_id)
                        )
                conn.execute("UPDATE events SET timelineid=? WHERE timelineid IS NULL",
                             (practice_id,))
            else:
                # Migrate any orphaned records to first timeline
                first_id = conn.execute(
                    "SELECT TimelineID FROM Timeline ORDER BY TimelineID LIMIT 1"
                ).fetchone()[0]
                conn.execute("UPDATE Category SET timelineid=? WHERE timelineid IS NULL",
                             (first_id,))
                conn.execute("UPDATE events SET timelineid=? WHERE timelineid IS NULL",
                             (first_id,))

        self._migrate_from_json()

    def _migrate_from_json(self, json_file="timeline.json"):
        if not os.path.exists(json_file):
            return
        with sqlite3.connect(self.db_file) as conn:
            count = conn.execute("SELECT COUNT(*) FROM events").fetchone()[0]
            if count > 0:
                return
            practice_id = conn.execute(
                "SELECT TimelineID FROM Timeline WHERE Title='Practice Timeline'"
            ).fetchone()[0]
            with open(json_file, "r") as f:
                events = json.load(f)
            for e in events:
                row = conn.execute(
                    "SELECT CategoryID FROM Category WHERE LOWER(Title)=LOWER(?) AND timelineid=?",
                    (e.get("category", "general"), practice_id)
                ).fetchone()
                cat_id = row[0] if row else None
                yr = e["year"]
                conn.execute(
                    "INSERT INTO events (title, year, desc, categoryid, timelineid, "
                    "start_value, start_display, start_unit) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (e["title"], yr, e.get("desc", ""), cat_id, practice_id,
                     float(yr), f"{yr} CE", "CE")
                )



    def load(self):
        with sqlite3.connect(self.db_file) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("""
                SELECT e.id, e.title, e.year, e.desc, e.categoryid,
                       c.Title as category, e.url,
                       e.image, e.image_name, e.image_type,
                       e.start_value, e.start_display, e.start_unit, e.start_month, e.start_day,
                       e.end_value,   e.end_display,   e.end_unit,   e.end_month,   e.end_day,
                       e.sort_order, e.standalone, e.hidden, e.picture_position,
                       e.linked_categoryid, e.linked_timelineid
                FROM events e
                LEFT JOIN Category c ON e.categoryid = c.CategoryID
                WHERE e.timelineid=?
                ORDER BY e.start_value
            """, (self.active_timeline_id,)).fetchall()
        self.events = [dict(row) for row in rows]
        for e in self.events:
            for prefix in ("start", "end"):
                # Refresh "Present" values to today's date every time events load
                if e.get(f"{prefix}_unit") == "Present":
                    e[f"{prefix}_value"]   = _today_value()
                    e[f"{prefix}_display"] = "Present"
                else:
                    self._regenerate_display(e, prefix)

    def _regenerate_display(self, e, prefix):
        """Recompute display string from stored components."""
        unit  = e.get(f"{prefix}_unit")
        value = e.get(f"{prefix}_value")
        month = e.get(f"{prefix}_month") or 0
        day   = e.get(f"{prefix}_day") or 0
        if value is None or unit is None:
            return
        if unit == "CE":
            num_str = str(int(value))
        elif unit == "BCE":
            num_str = str(int(-value))
        elif unit == "MYA":
            num_str = f"{(-value / 1_000_000):g}"
        elif unit == "BYA":
            num_str = f"{(-value / 1_000_000_000):g}"
        else:
            return
        e[f"{prefix}_display"] = _date_display(num_str, unit, month, day)

    def save(self):
        with sqlite3.connect(self.db_file) as conn:
            for e in self.events:
                if "id" in e:
                    # Use stored categoryid directly; only resolve by name as fallback
                    cat_id = e.get("categoryid") if e.get("categoryid") is not None \
                        else self._ensure_category(conn, e.get("category", "General"))
                    conn.execute(
                        "UPDATE events SET title=?, year=?, desc=?, categoryid=?, url=?, "
                        "image=?, image_name=?, image_type=?, "
                        "start_value=?, start_display=?, start_unit=?, start_month=?, start_day=?, "
                        "end_value=?,   end_display=?,   end_unit=?,   end_month=?,   end_day=?, "
                        "standalone=?, sort_order=?, hidden=?, linked_categoryid=?, linked_timelineid=? "
                        "WHERE id=?",
                        (e["title"], e.get("year", 0), e["desc"], cat_id, e.get("url", ""),
                         e.get("image"), e.get("image_name"), e.get("image_type"),
                         e.get("start_value"), e.get("start_display"), e.get("start_unit"),
                         e.get("start_month"), e.get("start_day"),
                         e.get("end_value"), e.get("end_display"), e.get("end_unit"),
                         e.get("end_month"), e.get("end_day"),
                         e.get("standalone", 0), e.get("sort_order"), e.get("hidden", 0),
                         e.get("linked_categoryid"), e.get("linked_timelineid"),
                         e["id"])
                    )

    def add(self, title, desc, category="General", url="",
            image=None, image_name=None, image_type=None,
            start_value=None, start_display="", start_unit="CE",
            start_month=None, start_day=None,
            end_value=None, end_display="", end_unit="CE",
            end_month=None, end_day=None, standalone=0, linked_categoryid=None,
            linked_timelineid=None):
        with sqlite3.connect(self.db_file) as conn:
            cat_id = self._ensure_category(conn, category)
            year = int(start_value) if start_value is not None else 0
            next_order = conn.execute(
                "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM events "
                "WHERE categoryid=? AND timelineid=?",
                (cat_id, self.active_timeline_id)
            ).fetchone()[0]
            cursor = conn.execute(
                "INSERT INTO events (title, year, desc, categoryid, timelineid, url, "
                "image, image_name, image_type, "
                "start_value, start_display, start_unit, start_month, start_day, "
                "end_value, end_display, end_unit, end_month, end_day, standalone, sort_order, hidden, "
                "linked_categoryid, linked_timelineid) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (title, year, desc, cat_id, self.active_timeline_id, url,
                 image, image_name, image_type,
                 start_value, start_display, start_unit, start_month, start_day,
                 end_value, end_display, end_unit, end_month, end_day, standalone, next_order, 0,
                 linked_categoryid, linked_timelineid)
            )
            new_id = cursor.lastrowid
        self.events.append({
            "id": new_id, "title": title, "year": year, "desc": desc,
            "category": category, "categoryid": cat_id, "url": url,
            "image": image, "image_name": image_name, "image_type": image_type,
            "start_value": start_value, "start_display": start_display,
            "start_unit": start_unit, "start_month": start_month, "start_day": start_day,
            "end_value": end_value, "end_display": end_display,
            "end_unit": end_unit, "end_month": end_month, "end_day": end_day,
            "standalone": standalone, "sort_order": next_order, "hidden": False,
            "linked_categoryid": linked_categoryid, "linked_timelineid": linked_timelineid,
        })

    @staticmethod
    def image_to_blob(path, max_size=(200, 200)):
        img = Image.open(path)
        img.thumbnail(max_size, Image.LANCZOS)
        fmt = img.format or "PNG"
        buf = io.BytesIO()
        img.save(buf, format=fmt)
        return buf.getvalue(), os.path.basename(path), fmt

    def delete_event(self, event):
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("DELETE FROM events WHERE id=?", (event["id"],))
        self.events[:] = [e for e in self.events if e["id"] != event["id"]]

    def reorder_event(self, event_id, direction):
        """Move an event up (-1) or down (+1) within its category custom sort."""
        event = next((e for e in self.events if e["id"] == event_id), None)
        if event is None:
            return
        cat = event.get("category") or "General"
        # Events in same category sorted by current sort_order
        siblings = sorted(
            [e for e in self.events if (e.get("category") or "General") == cat],
            key=lambda e: (e.get("sort_order") or 0, e["id"])
        )
        idx = next((i for i, e in enumerate(siblings) if e["id"] == event_id), None)
        if idx is None:
            return
        swap_idx = idx + direction
        if swap_idx < 0 or swap_idx >= len(siblings):
            return
        a, b = siblings[idx], siblings[swap_idx]
        a_order = a.get("sort_order") or a["id"]
        b_order = b.get("sort_order") or b["id"]
        with sqlite3.connect(self.db_file) as conn:
            conn.execute("UPDATE events SET sort_order=? WHERE id=?", (b_order, a["id"]))
            conn.execute("UPDATE events SET sort_order=? WHERE id=?", (a_order, b["id"]))
        a["sort_order"] = b_order
        b["sort_order"] = a_order



class TimelineView:
    """Standalone timeline canvas dialog."""

    RULER_HEIGHT  = 50
    LABEL_WIDTH   = 130
    LANE_HEIGHT   = 30
    LANE_PAD      = 3
    BAR_H         = 14
    IMG_PAD       = 10   # extra vertical padding (5px top + 5px bottom) for image rows
    LABEL_TIER_H  = 16   # height per label tier when staggering overlapping titles
    TICK_H        = 20
    CANVAS_MARGIN  = 20   # px of blank space at each horizontal end
    BREAK_STUB_PX  = 20   # pixels allocated to each break marker
    CAT_HEADER_H   = 36   # height of the category title bar row

    # Category colour palette (used for label strips and lane backgrounds)
    CAT_COLORS = [
        "#4e79a7", "#f28e2b", "#e15759", "#76b7b2",
        "#59a14f", "#edc948", "#b07aa1", "#ff9da7",
        "#9c755f", "#bab0ac",
    ]

    # Tint factors cycled per-event within a category.
    # 0.0 = base color, 1.0 = white.  Values kept light enough to
    # stay readable against the canvas background.
    EVENT_SHADE_FACTORS = [0.45, 0.20, 0.62, 0.33]

    def __init__(self, parent, timeline_db):
        self.db          = timeline_db
        self.px_per_year = 8.0
        self._drag_x     = None
        self._last_mouse_x = None
        self._collapsed_cats = set()

        self._init_data()
        self._build_window(parent)

    def _init_data(self):
        """(Re)compute all data-derived state from self.db.events/categories."""
        import math
        self.db.load_categories()

        # Hidden categories and events are excluded from the canvas entirely.
        # If a parent category is hidden, all its subcategories/events are also hidden.
        _hidden_cat_ids = set()

        def _mark_hidden(node, parent_hidden=False):
            is_hidden = parent_hidden or node.get("hidden")
            if is_hidden:
                _hidden_cat_ids.add(node["id"])
            for child in node.get("children", []):
                _mark_hidden(child, is_hidden)

        for _root in self.db.cat_tree:
            _mark_hidden(_root)

        _hidden_cat_names = {self.db.cat_by_id[cid]["title"] for cid in _hidden_cat_ids}
        _visible_events   = [e for e in self.db.events
                             if not e.get("hidden")
                             and (e.get("category") or "General") not in _hidden_cat_names]

        vals     = [e["start_value"] for e in _visible_events if e.get("start_value") is not None]
        end_vals = [e["end_value"]   for e in _visible_events if e.get("end_value")   is not None]
        raw_min  = min(vals + end_vals) if vals else 0
        raw_max  = max(vals + end_vals) if vals else 100
        raw_span = raw_max - raw_min or 1

        interval       = self._tick_interval(raw_span)
        self.min_date  = math.floor(raw_min / interval) * interval
        self.max_date  = math.ceil(raw_max  / interval) * interval
        if self.max_date == self.min_date:
            self.min_date -= interval
            self.max_date += interval

        # Icon styles hardcoded to Line; selection UI removed
        self._icon_short, self._icon_long = "Line", "Line"
        # self._icon_short, self._icon_long = self.db.load_timeline_icons(self.db.active_timeline_id)

        self._cat_header_style     = self.db.load_timeline_cat_header_style(self.db.active_timeline_id)
        self._cat_header_title_pos = self.db.load_timeline_cat_header_title_pos(self.db.active_timeline_id)
        self._canvas_bg_color      = self.db.load_timeline_canvas_bg(self.db.active_timeline_id) or "#f0ede8"

        # Override with manually saved ruler bounds if set
        ruler_min, ruler_max, ruler_max_present = self.db.load_timeline_ruler(self.db.active_timeline_id)
        if ruler_min is not None:
            self.min_date = ruler_min
        if ruler_max_present:
            self.max_date = _today_value()
        elif ruler_max is not None:
            self.max_date = ruler_max

        self._breaks = self.db.load_timeline_breaks(self.db.active_timeline_id)

        _known_cats      = {n["title"] for n in self.db.cat_nodes}
        event_cats       = {e.get("category") or "General" for e in _visible_events}
        self._seen_cats  = [n["title"] for n in self.db.cat_nodes
                            if n["title"] in event_cats and not n.get("hidden")]
        self._cat_depth  = {n["title"]: n["depth"] for n in self.db.cat_nodes}
        for cat in event_cats:
            if cat not in self._seen_cats and cat not in _known_cats:
                # Only surface orphan categories that truly have no DB entry
                # (e.g. imported data with a category name not yet created).
                # "General" and other fallback names that appear solely because
                # an event has a null category are suppressed.
                if cat != "General":
                    self._seen_cats.append(cat)
                    self._cat_depth[cat] = 0

        _saved_colors = {n["title"]: n["color"] for n in self.db.cat_nodes if n.get("color")}
        self.cat_color = {
            c: _saved_colors.get(c) or self.CAT_COLORS[i % len(self.CAT_COLORS)]
            for i, c in enumerate(self._seen_cats)
        }
        self.cat_row_bg = {n["title"]: n["row_bg_color"]
                           for n in self.db.cat_nodes if n.get("row_bg_color")}
        self.cat_show_guide = {n["title"]: n["show_row_guide"]
                               for n in self.db.cat_nodes}
        self.cat_image      = {n["title"]: n["cat_image"]
                               for n in self.db.cat_nodes if n.get("cat_image")}
        self.cat_image_pos  = {n["title"]: (n.get("cat_image_pos") or "Row")
                               for n in self.db.cat_nodes}
        self.cat_pad_top    = {n["title"]: int(n.get("cat_pad_top")    or 0)
                               for n in self.db.cat_nodes}
        self.cat_pad_bottom = {n["title"]: int(n.get("cat_pad_bottom") or 0)
                               for n in self.db.cat_nodes}

        _bg = self.db.load_timeline_bg_image(self.db.active_timeline_id)
        self._tl_bg_image_blob = _bg["image"]
        self._tl_bg_image_pos  = _bg["pos"]
        self._tl_bg_photo      = None   # PhotoImage kept alive across draw calls

        self._all_cat_events = {
            cat: sorted(
                [e for e in _visible_events if (e.get("category") or "General") == cat],
                key=lambda e: (e.get("sort_order") or e["id"])
            )
            for cat in self._seen_cats
        }

        self._event_color_map = {}
        for cat, events in self._all_cat_events.items():
            base = self.cat_color[cat]
            for e in events:
                self._event_color_map[e["id"]] = base

        self._build_layout()

    def _ancestor_collapsed(self, cat_title):
        """Return True if any ancestor category of cat_title is collapsed."""
        node = next((n for n in self.db.cat_nodes if n["title"] == cat_title), None)
        if node is None:
            return False
        # Walk up via parent_id
        pid = node.get("parent_id")
        while pid:
            parent = self.db.cat_by_id.get(pid)
            if parent is None:
                break
            if parent["title"] in self._collapsed_cats:
                return True
            pid = parent.get("parent_id")
        return False

    def _build_layout(self):
        """Rebuild self.rows and self.cat_row_spans respecting collapsed categories.

        Each entry in self.rows is a list of events sharing the same horizontal lane.
        Non-overlapping events without the standalone flag are packed greedily.
        Standalone events always occupy their own lane.
        """
        self.rows = []
        self.cat_row_spans = {}
        for cat in self._seen_cats:
            if self._ancestor_collapsed(cat):
                continue
            events = self._all_cat_events[cat]
            first = len(self.rows)
            if self._cat_header_style in ("Top", "Both") and cat not in self._collapsed_cats:
                self.rows.append([{"_cat_header": True, "category": cat}])
            pad_top = getattr(self, "cat_pad_top", {}).get(cat, 0)
            if pad_top > 0 and cat not in self._collapsed_cats:
                self.rows.append([{"_cat_pad": True, "category": cat, "_pad_h": pad_top}])
            if cat in self._collapsed_cats:
                total = sum(
                    len(self._all_cat_events.get(d, []))
                    for d in [cat] + self._descendant_titles(
                        next((n for n in self.db.cat_nodes if n["title"] == cat), {"children": []})
                    )
                )
                self.rows.append([{"_collapsed": True, "category": cat, "_count": total}])
            else:
                # Greedy interval packing — preserve user sort_order so lane assignment
                # matches the navigation panel order.
                # slots: list of {'max_end': float|None, 'events': list, 'standalone': bool}
                slots = []
                pack_order = sorted(
                    events,
                    key=lambda e: (e.get("sort_order") or 0, e.get("id") or 0)
                )
                for e in pack_order:
                    sv     = e.get("start_value")
                    ev_end = e.get("end_value") if e.get("end_value") is not None else sv
                    is_standalone = bool(e.get("standalone"))

                    if is_standalone or sv is None:
                        slots.append({"max_end": ev_end, "events": [e], "standalone": True})
                        continue

                    placed = False
                    for slot in slots:
                        if slot["standalone"]:
                            continue
                        if slot["max_end"] is None or sv >= slot["max_end"]:
                            slot["events"].append(e)
                            slot["max_end"] = max(
                                slot["max_end"] if slot["max_end"] is not None else sv,
                                ev_end if ev_end is not None else sv
                            )
                            placed = True
                            break

                    if not placed:
                        slots.append({"max_end": ev_end, "events": [e], "standalone": False})

                for slot in slots:
                    self.rows.append(slot["events"])

            pad_bottom = getattr(self, "cat_pad_bottom", {}).get(cat, 0)
            if pad_bottom > 0 and cat not in self._collapsed_cats:
                self.rows.append([{"_cat_pad": True, "category": cat, "_pad_h": pad_bottom}])
            self.cat_row_spans[cat] = (first, len(self.rows) - 1)

    # ── window ────────────────────────────────────────────────────────────────

    def _build_window(self, parent):
        win = parent
        win.title("Timeline Editor")
        win.minsize(900, 400)
        win.state("zoomed")
        self.win = win
        win.protocol("WM_DELETE_WINDOW", self._on_close)

        # Toolbar
        tb = tk.Frame(win, bd=1, relief=tk.RAISED, bg="#e0e0e0")
        tb.pack(side=tk.TOP, fill=tk.X)
        self._tl_var = tk.StringVar()
        self._tl_combo = ttk.Combobox(tb, textvariable=self._tl_var,
                                      state="readonly", font=("Arial", 9), width=30)
        self._tl_combo.pack(side=tk.LEFT, padx=(8, 4), pady=3)
        ts_box = tk.LabelFrame(tb, text="Time Scale", font=("Arial", 7), bg="#e0e0e0",
                               padx=4, pady=1)
        ts_box.pack(side=tk.LEFT, padx=6, pady=2)
        self._btn_fit_all = tk.Button(ts_box, text="Fit All", command=self._fit_all, width=6)
        self._btn_fit_all.pack(side=tk.LEFT, padx=(0, 4))
        self._scale_var = tk.StringVar()
        self._scale_entry = tk.Entry(ts_box, textvariable=self._scale_var, width=10,
                                     state="normal", relief=tk.SUNKEN,
                                     font=("Arial", 8), justify=tk.CENTER)
        self._scale_entry.pack(side=tk.LEFT, padx=(0, 2))
        self._scale_entry.bind("<Return>",   lambda e: self._apply_manual_scale())
        self._scale_entry.bind("<FocusOut>", lambda e: self._apply_manual_scale())
        self._scale_unit_var = tk.StringVar(value="px/yr")
        tk.Label(ts_box, textvariable=self._scale_unit_var,
                 font=("Arial", 8), bg="#e0e0e0").pack(side=tk.LEFT, padx=(0, 6))
        self._freeze_scale_var = tk.BooleanVar(value=False)
        tk.Checkbutton(ts_box, text="Freeze", variable=self._freeze_scale_var,
                       bg="#e0e0e0", command=self._on_freeze_toggle).pack(side=tk.LEFT, padx=(0, 2))
        tk.Button(ts_box, text="Save", command=self._save_view_state,
                  width=5).pack(side=tk.LEFT)
        tk.Label(tb, textvariable=self._tl_var, font=("Arial", 13, "bold"),
                 fg="#34495e", bg="#e0e0e0").pack(side=tk.LEFT, expand=True)
        tk.Button(tb, text="Help",     command=self._open_help,           width=7).pack(side=tk.RIGHT, padx=(4, 6), pady=3)
        tk.Button(tb, text="Save PDF", command=self._save_pdf,            width=9).pack(side=tk.RIGHT, padx=4,      pady=3)
        tk.Button(tb, text="Publish",  command=self._publish_to_timelinehub,
                  width=9, bg="#2980b9", fg="white",
                  activebackground="#1a6fa0", activeforeground="white",
                  relief=tk.RAISED).pack(side=tk.RIGHT, padx=4, pady=3)

        # Status bar
        self.status_var = tk.StringVar()
        tk.Label(win, textvariable=self.status_var, anchor=tk.W,
                 relief=tk.SUNKEN, font=("Arial", 8), padx=5).pack(
                 side=tk.BOTTOM, fill=tk.X)

        # Main area: nav panel + canvas
        main_frame = tk.Frame(win)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ── Left nav panel ───────────────────────────────────────────────────
        nav_frame = tk.Frame(main_frame, width=220, bg="#2c3e50")
        nav_frame.pack(side=tk.LEFT, fill=tk.Y)
        nav_frame.pack_propagate(False)

        self._sync_tl_combo()
        self._tl_combo.bind("<<ComboboxSelected>>", self._on_tl_selected)

        # Manage buttons below the combobox
        tl_manage_frame = tk.Frame(nav_frame, bg="#2c3e50")
        tl_manage_frame.pack(fill=tk.X, padx=6, pady=(0, 2))
        tk.Button(tl_manage_frame, text="Manage Timelines", font=("Arial", 8),
                  command=self._open_manage_timelines).pack(fill=tk.X, expand=True)

        cat_manage_frame = tk.Frame(nav_frame, bg="#2c3e50")
        cat_manage_frame.pack(fill=tk.X, padx=6, pady=(0, 2))
        tk.Button(cat_manage_frame, text="Manage Categories", font=("Arial", 8),
                  command=lambda: self._manage_categories_dialog(self.win)).pack(fill=tk.X, expand=True)

        add_event_frame = tk.Frame(nav_frame, bg="#2c3e50")
        add_event_frame.pack(fill=tk.X, padx=6, pady=(0, 4))
        tk.Button(add_event_frame, text="Add Event", font=("Arial", 8),
                  command=self._open_add_event_dialog).pack(fill=tk.X, expand=True)

        # Expand / Collapse all buttons
        nav_btn_frame = tk.Frame(nav_frame, bg="#2c3e50")
        nav_btn_frame.pack(fill=tk.X, padx=6, pady=(0, 6))

        tk.Button(nav_btn_frame, text="Expand All", command=self._expand_all,
                  font=("Arial", 9), width=10).pack(side=tk.LEFT, padx=(0, 4), expand=True)
        tk.Button(nav_btn_frame, text="Collapse All", command=self._collapse_all,
                  font=("Arial", 9), width=11).pack(side=tk.LEFT, expand=True)

        # Nav tree
        nav_scroll = tk.Scrollbar(nav_frame, orient=tk.VERTICAL)
        self.nav_tree = ttk.Treeview(nav_frame, show="tree",
                                     yscrollcommand=nav_scroll.set,
                                     selectmode="browse")
        nav_scroll.config(command=self.nav_tree.yview)
        nav_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.nav_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Style nav tree
        style = ttk.Style()
        style.configure("Nav.Treeview", background="#34495e", foreground="white",
                        fieldbackground="#34495e", font=("Arial", 9),
                        rowheight=24)
        style.configure("Nav.Treeview.Item", padding=(4, 0))
        style.map("Nav.Treeview", background=[("selected", "#4a6fa5")],
                  foreground=[("selected", "white")])
        self.nav_tree.configure(style="Nav.Treeview")

        # Drop indicator line — placed over the treeview during drag
        self._drop_line = tk.Frame(self.nav_tree, height=2, bg="#2ecc71")

        self._build_nav_tree()
        self.nav_tree.bind("<<TreeviewSelect>>", self._on_nav_select)
        self.nav_tree.bind("<<TreeviewClose>>",  self._on_nav_collapse)
        self.nav_tree.bind("<<TreeviewOpen>>",   self._on_nav_expand)
        self.nav_tree.bind("<Button-3>",         self._on_nav_right_click)
        self.nav_tree.bind("<ButtonPress-1>",    self._on_nav_drag_start)
        self.nav_tree.bind("<B1-Motion>",        self._on_nav_drag_motion)
        self.nav_tree.bind("<ButtonRelease-1>",  self._on_nav_drag_end)
        self._drag_item     = None   # iid being dragged
        self._drag_target   = None   # iid currently highlighted as drop target
        self._drag_moved    = False  # did the mouse actually move enough to count as a drag

        # ── Canvas + scrollbars ──────────────────────────────────────────────
        canvas_frame = tk.Frame(main_frame)
        canvas_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.hscroll = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        self.vscroll = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)

        # ── Top strip: dark corner + frozen ruler ────────────────────────────
        top_strip = tk.Frame(canvas_frame, height=self.RULER_HEIGHT, bg="#2c3e50")
        top_strip.pack(side=tk.TOP, fill=tk.X)
        top_strip.pack_propagate(False)
        # Corner square above the label column — same dark colour as ruler
        self._label_corner = tk.Frame(top_strip, width=self.LABEL_WIDTH, bg="#2c3e50")
        self._label_corner.pack(side=tk.LEFT, fill=tk.Y)
        # Ruler canvas: horizontally scrollable, never vertically scrollable
        self.ruler_canvas = tk.Canvas(top_strip, bg="#2c3e50",
                                      highlightthickness=0)
        self.ruler_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── Body: frozen label column + scrollable event canvas ─────────────
        self.vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.hscroll.pack(side=tk.BOTTOM, fill=tk.X)
        body_frame = tk.Frame(canvas_frame)
        body_frame.pack(fill=tk.BOTH, expand=True)

        # Label column: synced vertically, never scrolls horizontally
        self.label_canvas = tk.Canvas(body_frame, width=self.LABEL_WIDTH,
                                      bg="#f0ede8", highlightthickness=0)
        self.label_canvas.pack(side=tk.LEFT, fill=tk.Y)

        # Main event canvas
        self.canvas = tk.Canvas(body_frame, bg=self._canvas_bg_color, cursor="fleur",
                                highlightthickness=0)

        # Horizontal scroll: ruler + event canvas move together
        def _xscroll_both(*args):
            self.ruler_canvas.xview(*args)
            self.canvas.xview(*args)
        self.hscroll.config(command=_xscroll_both)
        def _on_xscroll(*args):
            self.hscroll.set(*args)
            pos = getattr(self, "_cat_header_title_pos", "Center (View)")
            if pos in ("Left (View)", "Center (View)"):
                self._update_floating_headers()
        self.canvas.config(xscrollcommand=_on_xscroll)

        # Vertical scroll: label column + event canvas move together
        def _yscroll_both(*args):
            self.label_canvas.yview(*args)
            self.canvas.yview(*args)
        self.vscroll.config(command=_yscroll_both)
        self.canvas.config(yscrollcommand=self.vscroll.set)

        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Arrow-key / page-key bindings (bound on the window so focus doesn't matter)
        win.bind("<Left>",     lambda e: self._arrow_scroll_h(-1))
        win.bind("<Right>",    lambda e: self._arrow_scroll_h(1))
        win.bind("<Up>",       lambda e: self._zoom_at_mouse(1.5))
        win.bind("<Down>",     lambda e: self._zoom_at_mouse(1 / 1.5))
        win.bind("<Prior>",    lambda e: self._arrow_scroll_v(-1))  # Page Up
        win.bind("<Next>",     lambda e: self._arrow_scroll_v(1))   # Page Down
        win.bind("<Home>",     lambda e: self._scroll_home())
        win.bind("<End>",      lambda e: self._scroll_end())

        # Bindings — main event canvas
        self.canvas.bind("<ButtonPress-1>",   self._drag_start)
        self.canvas.bind("<B1-Motion>",        self._drag_move)
        self.canvas.bind("<MouseWheel>",       self._mouse_wheel)
        self.canvas.bind("<Button-4>",         self._mouse_wheel)
        self.canvas.bind("<Button-5>",         self._mouse_wheel)
        self.canvas.bind("<Motion>",           self._on_hover)
        self.canvas.bind("<Button-3>",         self._on_canvas_right_click)
        self.canvas.bind("<Leave>",            lambda e: self._tooltip_hide_soon())

        # Ruler — drag pans horizontally; wheel zooms
        self.ruler_canvas.bind("<ButtonPress-1>", self._drag_start_ruler)
        self.ruler_canvas.bind("<B1-Motion>",      self._drag_move_ruler)
        self.ruler_canvas.bind("<Motion>",         self._on_ruler_motion)
        self.ruler_canvas.bind("<MouseWheel>",     self._mouse_wheel)
        self.ruler_canvas.bind("<Button-4>",       self._mouse_wheel)
        self.ruler_canvas.bind("<Button-5>",       self._mouse_wheel)
        self.ruler_canvas.bind("<Button-3>",       self._on_ruler_right_click)

        # Label column — drag scrolls vertically; wheel zooms
        self.label_canvas.bind("<ButtonPress-1>", self._drag_start_label)
        self.label_canvas.bind("<B1-Motion>",      self._drag_move_label)
        self.label_canvas.bind("<MouseWheel>",     self._mouse_wheel)
        self.label_canvas.bind("<Button-4>",       self._mouse_wheel)
        self.label_canvas.bind("<Button-5>",       self._mouse_wheel)

        self._nav_selected_iid = None
        self._tip_win   = None
        self._tip_event = None
        self._tip_after = None
        self._link_icon_map     = {}   # canvas item id -> event dict
        self._timeline_icon_map = {}   # canvas item id -> event dict
        self.canvas.tag_bind("link_icon",     "<Button-1>", self._on_link_icon_click)
        self.canvas.tag_bind("timeline_icon", "<Button-1>", self._on_timeline_icon_click)
        # Defer until the window is fully rendered, then expand all and apply
        # the saved timescale (or fit all if none has been saved).
        def _startup():
            self._expand_all(draw=False)
            self.win.update_idletasks()
            state = self.db.load_timeline_view_state(self.db.active_timeline_id)
            if state:
                self.px_per_year = state["px_per_year"]
                self._freeze_scale_var.set(state["freeze_scale"])
                self._on_freeze_toggle()
                self._draw()
                self._scroll_to_visible()
            else:
                self._freeze_scale_var.set(False)
                self._on_freeze_toggle()
                self._fit_all()
        self.win.after(50, _startup)

    def _build_nav_tree(self):
        """Populate the nav tree respecting category parent/child hierarchy."""
        self.nav_tree.delete(*self.nav_tree.get_children())
        self._nav_iid_map = {}   # iid -> event dict or None (for category nodes)

        def _should_show_in_nav(node, par_hidden=False):
            """True if this node or any descendant has visible events or is hidden."""
            h = par_hidden or node.get("hidden")
            if node["title"] in self._all_cat_events or h:
                return True
            return any(_should_show_in_nav(c, h) for c in node.get("children", []))

        def _insert_cat(node, parent_iid="", inherited_hidden=False):
            cat = node["title"]
            is_hidden = inherited_hidden or node.get("hidden")
            if not _should_show_in_nav(node, inherited_hidden):
                return
            cat_iid   = f"cat_{node['id']}"
            cat_tag   = "hidden_cat" if is_hidden else "category"
            cat_label = f"🚫 {cat}" if node.get("hidden") else cat
            self.nav_tree.insert(parent_iid, tk.END, iid=cat_iid,
                                 text=cat_label, open=False, tags=(cat_tag,))
            self._nav_iid_map[cat_iid] = None

            # Always insert events (including hidden) so user can see and manage them.
            all_events_for_cat = self.db.events  # full list, not filtered
            for e in sorted(
                [ev for ev in all_events_for_cat
                 if (ev.get("category") or "General") == cat],
                key=lambda ev: (ev.get("start_value") or 0, ev.get("sort_order") or ev["id"])
            ):
                ev_hidden = e.get("hidden") or is_hidden
                ev_tag    = "hidden_event" if ev_hidden else "event"
                ev_label  = f"🚫 {e['title']}" if e.get("hidden") else e["title"]
                ev_iid = self.nav_tree.insert(cat_iid, tk.END,
                                              text=ev_label,
                                              tags=(ev_tag,))
                self._nav_iid_map[ev_iid] = e

            # Recurse into child categories, propagating hidden state
            for child in node.get("children", []):
                _insert_cat(child, cat_iid, inherited_hidden=is_hidden)

        for root_node in self.db.cat_tree:
            _insert_cat(root_node)

        self.nav_tree.tag_configure("category",    foreground="#f0c040",
                                    font=("Arial", 9, "bold"))
        self.nav_tree.tag_configure("hidden_cat",  foreground="#cc3333",
                                    font=("Arial", 9, "bold"))
        self.nav_tree.tag_configure("event",       foreground="white",
                                    font=("Arial", 9))
        self.nav_tree.tag_configure("hidden_event",foreground="#cc3333",
                                    font=("Arial", 9))
        self.nav_tree.tag_configure("drop_here",   foreground="white",
                                    background="#2980b9", font=("Arial", 9))

    def _descendant_titles(self, node):
        """Return all descendant category titles of a node."""
        titles = []
        for child in node.get("children", []):
            titles.append(child["title"])
            titles.extend(self._descendant_titles(child))
        return titles

    def _on_nav_select(self, _event=None):
        """Scroll the canvas to bring the selected nav item into view."""
        sel = self.nav_tree.selection()
        if not sel:
            return
        iid = sel[0]
        entry = self._nav_iid_map.get(iid)

        if entry is None:
            # Category node — extract title from iid or text
            cat_text = self._nav_iid_to_cat(iid)
            if cat_text and cat_text in self.cat_row_spans:
                row_index = self.cat_row_spans[cat_text][0]
            else:
                return
        else:
            # Event node — find its row index in self.rows
            ev_id = entry.get("id")
            row_index = next(
                (i for i, slot in enumerate(self.rows)
                 if any(e.get("id") == ev_id for e in slot)),
                None
            )
            if row_index is None:
                return

        self._nav_selected_iid = iid
        self._highlight_row(row_index)

        # ── Vertical scroll — bring row into view ────────────────────────────
        rh      = self._row_heights[row_index] if getattr(self, "_row_heights", None) else self.LANE_HEIGHT
        ry      = self._lane_y(row_index)
        total_h = self._total_height()
        vp_h    = self.canvas.winfo_height()
        vy_top  = self.canvas.yview()[0] * total_h
        vy_bot  = vy_top + vp_h

        if ry < vy_top or ry + rh > vy_bot:
            frac = max(0.0, (ry - 10) / total_h)
            self.canvas.yview_moveto(frac)
            self.label_canvas.yview_moveto(frac)

        # ── Horizontal scroll — bring event into view ────────────────────────
        if entry is not None:
            sv = entry.get("start_value")
            ev = entry.get("end_value")
            if sv is not None:
                x1      = self._x(sv)
                x2      = self._x(ev) if ev and ev != sv else x1
                W       = self._canvas_width()
                vp_w    = self.canvas.winfo_width()
                vx_left = self.canvas.xview()[0] * W
                vx_right = vx_left + vp_w
                if x1 < vx_left or x2 > vx_right:
                    center = (x1 + x2) / 2
                    frac   = max(0.0, min(1.0, (center - vp_w / 2) / W))
                    self.canvas.xview_moveto(frac)
                    self.ruler_canvas.xview_moveto(frac)

    def _highlight_row(self, row_index):
        """Highlight a row permanently until another selection replaces it."""
        self.canvas.delete("nav_highlight")
        rh = self._row_heights[row_index] if getattr(self, "_row_heights", None) else self.LANE_HEIGHT
        W  = self._canvas_width()
        y  = self._lane_y(row_index)
        self.canvas.create_rectangle(
            0, y, W, y + rh,
            fill="gray", stipple="gray12",
            outline="#f0c040", width=2, tags="nav_highlight"
        )

    def _nav_iid_to_cat(self, iid):
        """Return category title for a cat_ iid, or None."""
        if not iid.startswith("cat_"):
            return None
        return self.nav_tree.item(iid, "text").strip()

    def _update_window_title(self):
        title = self._tl_var.get()
        self.win.title(f"Timeline Editor — {title}" if title else "Timeline Editor")

    def _sync_tl_combo(self):
        """Populate the timeline combobox and select the active timeline."""
        timelines = self.db.load_timelines()   # [(id, title), ...]
        self._tl_id_map = {title: tid for tid, title in timelines}
        self._tl_combo["values"] = [t[1] for t in timelines]
        active_title = next((t for tid, t in timelines
                             if tid == self.db.active_timeline_id), None)
        if active_title:
            self._tl_var.set(active_title)
        self._update_window_title()

    def _on_tl_selected(self, *_):
        self.canvas.focus_set()
        title = self._tl_var.get()
        tid   = self._tl_id_map.get(title)
        if tid is None or tid == self.db.active_timeline_id:
            return
        self._reload(tid)

    def _open_manage_timelines(self):
        """Open the Manage Timelines dialog from the viewer."""
        def _on_change(new_active_id):
            self._sync_tl_combo()
            self._reload(new_active_id)

        def _on_import():
            self._sync_tl_combo()
            self._reload(self.db.active_timeline_id)

        EditTimelineDialog(self.win, self.db, view=self, on_change=_on_change, on_import=_on_import)

    def _reload(self, timeline_id):
        """Switch to a different timeline and redraw the viewer."""
        self.db.active_timeline_id = timeline_id
        self.db.load()
        self._collapsed_cats.clear()
        self._init_data()
        self.canvas.config(bg=self._canvas_bg_color)
        self._build_nav_tree()
        self._expand_all(draw=False)
        self._update_window_title()
        state = self.db.load_timeline_view_state(timeline_id)
        if state:
            self.px_per_year = state["px_per_year"]
            self._freeze_scale_var.set(state["freeze_scale"])
            self._on_freeze_toggle()
            self.win.update_idletasks()
            self._draw()
            self._scroll_to_visible()
        else:
            self._freeze_scale_var.set(False)
            self._on_freeze_toggle()
            self._fit_all()

    def _expand_all(self, draw=True):
        self._collapsed_cats.clear()
        self._build_layout()
        self._build_nav_tree()
        for iid in list(self._nav_iid_map):
            if iid.startswith("cat_"):
                self.nav_tree.item(iid, open=True)
        if draw:
            self._draw()

    def _collapse_all(self):
        self._collapsed_cats.update(self._all_cat_events.keys())
        self._build_layout()
        self._build_nav_tree()
        for iid in list(self._nav_iid_map):
            if iid.startswith("cat_"):
                self.nav_tree.item(iid, open=False)
        self._draw()

    def _on_nav_collapse(self, *_):
        """Category collapsed in nav tree — collapse its rows on canvas too."""
        iid = self.nav_tree.focus()
        cat = self._nav_iid_to_cat(iid)
        if cat and cat in self._all_cat_events:
            self._collapsed_cats.add(cat)
            self._build_layout()
            self._draw()

    def _on_nav_expand(self, *_):
        """Category expanded in nav tree — restore its rows on canvas."""
        iid = self.nav_tree.focus()
        cat = self._nav_iid_to_cat(iid)
        if cat:
            self._collapsed_cats.discard(cat)
            self._build_layout()
            self._draw()

    def _on_nav_right_click(self, event):
        """Right-click on nav tree: Edit event or Manage Categories depending on node type."""
        iid = self.nav_tree.identify_row(event.y)
        if not iid:
            return
        self.nav_tree.selection_set(iid)
        evt = self._nav_iid_map.get(iid)
        if evt is None:
            # Category node — open category manager, preselect the clicked category
            cat_id = int(iid.split("_", 1)[1]) if iid.startswith("cat_") else None
            self._manage_categories_dialog(self.win, preselect_id=cat_id)
        else:
            self._open_edit_event_dialog(evt)

    # ── Nav tree drag-to-reorder ───────────────────────────────────────────────

    def _on_nav_drag_start(self, event):
        iid = self.nav_tree.identify_row(event.y)
        if not iid or self._nav_iid_map.get(iid) is None:
            self._drag_item = None
            return
        # Toggle: clicking an already-selected item deselects it
        if iid == self._nav_selected_iid:
            self.nav_tree.selection_remove(iid)
            self.canvas.delete("nav_highlight")
            self._nav_selected_iid = None
            self._drag_item = None
            return "break"
        self._drag_item        = iid
        self._drag_moved       = False
        self._drag_target      = None
        self._drag_insert_after = False

    def _on_nav_drag_motion(self, event):
        if not self._drag_item:
            return
        self._drag_moved = True
        target = self.nav_tree.identify_row(event.y)
        if not target or target == self._drag_item:
            self._clear_drop_indicator()
            return
        target_evt = self._nav_iid_map.get(target)
        if target_evt is None:
            self._clear_drop_indicator()
            return
        drag_evt = self._nav_iid_map[self._drag_item]
        if (drag_evt.get("category") or "General") != (target_evt.get("category") or "General"):
            self._clear_drop_indicator()
            return

        # Direction: compare positions in the ordered list
        cat     = drag_evt.get("category") or "General"
        ordered = self._all_cat_events.get(cat, [])
        drag_idx = next((i for i, e in enumerate(ordered) if e["id"] == drag_evt["id"]), 0)
        tgt_idx  = next((i for i, e in enumerate(ordered) if e["id"] == target_evt["id"]), 0)
        insert_after = drag_idx < tgt_idx   # dragging down → place below target

        self._drag_target       = target
        self._drag_insert_after = insert_after

        # Draw the green insertion line above or below the target row
        bbox = self.nav_tree.bbox(target)
        if bbox:
            _, row_y, row_w, row_h = bbox
            line_y = (row_y + row_h) if insert_after else row_y
            self._drop_line.place(x=0, y=line_y - 1, width=row_w, height=2)
            self._drop_line.lift()

    def _on_nav_drag_end(self, event):
        if not self._drag_item or not self._drag_moved:
            self._drag_item = None
            self._clear_drop_indicator()
            return
        target       = self._drag_target
        insert_after = self._drag_insert_after
        drag_iid     = self._drag_item
        self._drag_item = None
        self._clear_drop_indicator()
        if not target:
            return
        drag_evt   = self._nav_iid_map.get(drag_iid)
        target_evt = self._nav_iid_map.get(target)
        if not drag_evt or not target_evt:
            return
        cat = drag_evt.get("category") or "General"
        if (target_evt.get("category") or "General") != cat:
            return
        ordered  = list(self._all_cat_events.get(cat, []))
        drag_idx = next((i for i, e in enumerate(ordered) if e["id"] == drag_evt["id"]), None)
        tgt_idx  = next((i for i, e in enumerate(ordered) if e["id"] == target_evt["id"]), None)
        if drag_idx is None or tgt_idx is None or drag_idx == tgt_idx:
            return
        item = ordered.pop(drag_idx)
        # Adjust target index for the removed item
        adj_tgt = tgt_idx - (1 if tgt_idx > drag_idx else 0)
        ordered.insert(adj_tgt + (1 if insert_after else 0), item)
        with sqlite3.connect(self.db.db_file) as conn:
            for i, e in enumerate(ordered):
                new_order = i + 1
                if (e.get("sort_order") or 0) != new_order:
                    conn.execute("UPDATE events SET sort_order=? WHERE id=?", (new_order, e["id"]))
                    e["sort_order"] = new_order
        self._reload(self.db.active_timeline_id)

    def _clear_drop_indicator(self):
        self._drop_line.place_forget()
        self._drag_target = None

    def _manage_categories_dialog(self, parent, on_close=None, preselect_id=None):
        """Category management dialog — mirrors TimelineApp.manage_categories."""
        win = tk.Toplevel(parent)
        win.title("Manage Categories")
        win.resizable(True, True)
        win.grab_set()
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)   # tree row stretches

        # ── Tree ──────────────────────────────────────────────────────────────
        tree_frame = tk.Frame(win, padx=16)
        tree_frame.grid(row=0, column=0, sticky="nsew", pady=(8, 4))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        vsb = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        vsb.grid(row=0, column=1, sticky="ns")
        cat_tree = ttk.Treeview(tree_frame, show="tree", yscrollcommand=vsb.set,
                                selectmode="browse", height=8)
        vsb.config(command=cat_tree.yview)
        cat_tree.grid(row=0, column=0, sticky="nsew")

        # ── Name row ──────────────────────────────────────────────────────────
        entry_frame = tk.Frame(win, padx=16)
        entry_frame.grid(row=1, column=0, sticky="ew", pady=(0, 2))
        tk.Label(entry_frame, text="ID:").pack(side=tk.LEFT)
        id_var = tk.StringVar(value="")
        tk.Entry(entry_frame, textvariable=id_var, width=6,
                 state="readonly").pack(side=tk.LEFT, padx=(2, 10))
        tk.Label(entry_frame, text="Name:").pack(side=tk.LEFT)
        name_var = tk.StringVar()
        name_entry = tk.Entry(entry_frame, textvariable=name_var, width=26)
        name_entry.pack(side=tk.LEFT, padx=6)

        # ── Parent row ────────────────────────────────────────────────────────
        parent_frame = tk.Frame(win, padx=16)
        parent_frame.grid(row=2, column=0, sticky="ew", pady=(0, 2))
        tk.Label(parent_frame, text="Parent:").pack(side=tk.LEFT)
        parent_var = tk.StringVar(value="<Root>")
        parent_combo = ttk.Combobox(parent_frame, textvariable=parent_var,
                                    state="readonly", width=30)
        parent_combo.pack(side=tk.LEFT, padx=6)

        # ── Hidden checkbox ───────────────────────────────────────────────────
        hidden_frame = tk.Frame(win, padx=16)
        hidden_frame.grid(row=3, column=0, sticky="ew", pady=(0, 2))
        hidden_cat_var = tk.IntVar(value=0)
        hidden_cat_chk = tk.Checkbutton(hidden_frame, text="Hidden (exclude from timeline)",
                                        variable=hidden_cat_var)
        hidden_cat_chk.pack(side=tk.LEFT)
        show_guide_var = tk.IntVar(value=1)
        tk.Checkbutton(hidden_frame, text="Show row guide lines",
                       variable=show_guide_var).pack(side=tk.LEFT, padx=(16, 0))

        # ── Color picker ──────────────────────────────────────────────────────
        color_frame = tk.Frame(win, padx=16)
        color_frame.grid(row=4, column=0, sticky="ew", pady=(0, 2))
        tk.Label(color_frame, text="Event color:").pack(side=tk.LEFT)
        color_var = tk.StringVar(value="")
        color_swatch = tk.Label(color_frame, width=4, relief="sunken", bd=1)
        color_swatch.pack(side=tk.LEFT, padx=(6, 4))

        def _apply_color(hex_color):
            color_var.set(hex_color)
            color_swatch.config(bg=hex_color if hex_color else win.cget("bg"))

        def _pick_color():
            from tkinter import colorchooser
            result = colorchooser.askcolor(color=color_var.get() or None, parent=win,
                                           title="Event Color")
            if result and result[1]:
                _apply_color(result[1])

        tk.Button(color_frame, text="Choose...", command=_pick_color).pack(side=tk.LEFT)
        tk.Button(color_frame, text="Clear",
                  command=lambda: _apply_color("")).pack(side=tk.LEFT, padx=(4, 0))

        # ── Row background color ───────────────────────────────────────────────
        row_bg_frame = tk.Frame(win, padx=16)
        row_bg_frame.grid(row=5, column=0, sticky="ew", pady=(0, 4))
        tk.Label(row_bg_frame, text="Row background:").pack(side=tk.LEFT)
        row_bg_var = tk.StringVar(value="")
        row_bg_swatch = tk.Label(row_bg_frame, width=4, relief="sunken", bd=1)
        row_bg_swatch.pack(side=tk.LEFT, padx=(6, 4))

        def _apply_row_bg(hex_color):
            row_bg_var.set(hex_color)
            row_bg_swatch.config(bg=hex_color if hex_color else win.cget("bg"))

        def _pick_row_bg():
            from tkinter import colorchooser
            result = colorchooser.askcolor(color=row_bg_var.get() or None, parent=win,
                                           title="Row Background Color")
            if result and result[1]:
                _apply_row_bg(result[1])

        tk.Button(row_bg_frame, text="Choose...", command=_pick_row_bg).pack(side=tk.LEFT)
        tk.Button(row_bg_frame, text="Clear",
                  command=lambda: _apply_row_bg("")).pack(side=tk.LEFT, padx=(4, 0))

        # ── Padding ───────────────────────────────────────────────────────────
        pad_frame = tk.Frame(win, padx=16)
        pad_frame.grid(row=6, column=0, sticky="ew", pady=(0, 4))
        tk.Label(pad_frame, text="Padding:").pack(side=tk.LEFT)
        cat_pad_top_var = tk.IntVar(value=0)
        cat_pad_bottom_var = tk.IntVar(value=0)
        tk.Label(pad_frame, text="Top").pack(side=tk.LEFT, padx=(8, 2))
        tk.Spinbox(pad_frame, from_=0, to=200, increment=4, width=5,
                   textvariable=cat_pad_top_var).pack(side=tk.LEFT)
        tk.Label(pad_frame, text="px", fg="gray").pack(side=tk.LEFT, padx=(2, 12))
        tk.Label(pad_frame, text="Bottom").pack(side=tk.LEFT, padx=(0, 2))
        tk.Spinbox(pad_frame, from_=0, to=200, increment=4, width=5,
                   textvariable=cat_pad_bottom_var).pack(side=tk.LEFT)
        tk.Label(pad_frame, text="px", fg="gray").pack(side=tk.LEFT, padx=(2, 0))

        # ── Category image ────────────────────────────────────────────────────
        THUMB_SIZE = 96

        # ── Image panel ───────────────────────────────────────────────────────
        img_panel = tk.LabelFrame(win, text="Category Image", padx=8, pady=6)
        img_panel.grid(row=7, column=0, sticky="ew", padx=16, pady=(0, 4))

        _cat_img_bytes   = [None]   # raw blob
        _cat_thumb_ref   = [None]   # PhotoImage ref for thumbnail (kept alive here)
        _cat_tmp_path    = [None]   # temp file written for external editor
        cat_img_name_var = tk.StringVar(value="")   # kept for preview title / reload
        img_size_var     = tk.StringVar(value="")   # pixel dimensions display

        # Single row: thumbnail column (left) + button column (right)
        thumb_row_frame = tk.Frame(img_panel)
        thumb_row_frame.pack(fill=tk.X, pady=(0, 6))

        # Thumbnail + size label stacked in left column
        thumb_col = tk.Frame(thumb_row_frame)
        thumb_col.pack(side=tk.LEFT, anchor=tk.N)
        thumb_container = tk.Frame(thumb_col, width=THUMB_SIZE, height=THUMB_SIZE,
                                   relief="sunken", bd=1)
        thumb_container.pack()
        thumb_container.pack_propagate(False)
        thumb_lbl = tk.Label(thumb_container, text="No image", fg="gray",
                             compound=tk.CENTER)
        thumb_lbl.pack(fill=tk.BOTH, expand=True)
        tk.Label(thumb_col, textvariable=img_size_var,
                 fg="gray", font=("Arial", 8)).pack(pady=(3, 0))

        # All buttons stacked to the right of the thumbnail
        edit_btn_frame = tk.Frame(thumb_row_frame)
        edit_btn_frame.pack(side=tk.LEFT, padx=(10, 0), anchor=tk.N)
        btn_choose_img    = tk.Button(edit_btn_frame, text="Choose...",          width=16)
        btn_clear_img     = tk.Button(edit_btn_frame, text="Clear",              width=16)
        btn_preview       = tk.Button(edit_btn_frame, text="Preview",            width=16)
        btn_open_editor   = tk.Button(edit_btn_frame, text="Open in Editor",     width=16)
        btn_reload_editor = tk.Button(edit_btn_frame, text="Reload from Editor",
                                      width=16, state=tk.DISABLED)
        btn_choose_img.pack(pady=(0, 4))
        btn_clear_img.pack(pady=(0, 4))
        btn_preview.pack(pady=(0, 4))
        btn_open_editor.pack(pady=(0, 4))
        btn_reload_editor.pack()

        _thumb_box_size = [THUMB_SIZE]   # current rendered box size

        # Keep the thumbnail box square, sized to the full height of the button column
        def _sync_thumb_size(event=None):
            h = edit_btn_frame.winfo_reqheight()
            if h > 10:
                _thumb_box_size[0] = h
                thumb_container.config(width=h, height=h)
                _render_thumb()
        edit_btn_frame.bind("<Configure>", _sync_thumb_size)

        def _render_thumb():
            """Re-render the thumbnail scaled to fill the current box size."""
            if not _cat_img_bytes[0]:
                return
            box = _thumb_box_size[0]
            try:
                img   = Image.open(io.BytesIO(_cat_img_bytes[0]))
                scale = box / max(img.width, img.height)
                new_w = max(1, int(img.width  * scale))
                new_h = max(1, int(img.height * scale))
                thumb = img.resize((new_w, new_h), Image.LANCZOS)
                photo = ImageTk.PhotoImage(thumb)
                _cat_thumb_ref[0] = photo
                thumb_lbl.config(image=photo, text="")
            except Exception:
                pass

        def _set_cat_image(data, name):
            _cat_img_bytes[0] = data
            cat_img_name_var.set(name or "")
            if data:
                try:
                    img = Image.open(io.BytesIO(data))
                    img_size_var.set(f"{img.width} \u00d7 {img.height} px")
                    _render_thumb()
                except Exception:
                    img_size_var.set("")
                    _cat_thumb_ref[0] = None
                    thumb_lbl.config(image="", text="(error)", fg="red")
            else:
                img_size_var.set("")
                _cat_thumb_ref[0] = None
                thumb_lbl.config(image="", text="No image", fg="gray")

        def _preview_cat_image():
            if not _cat_img_bytes[0]:
                return
            try:
                img = Image.open(io.BytesIO(_cat_img_bytes[0]))
            except Exception:
                return
            pw, ph = img.width, img.height
            pwin = tk.Toplevel(win)
            pwin.title(cat_img_name_var.get() or "Image Preview")
            pwin.resizable(True, True)
            # Cap initial window at 80% of screen
            sw = win.winfo_screenwidth()
            sh = win.winfo_screenheight()
            max_w = int(sw * 0.8)
            max_h = int(sh * 0.8)
            win_w = min(pw + 20, max_w)
            win_h = min(ph + 20, max_h)
            pwin.geometry(f"{win_w}x{win_h}+{(sw - win_w) // 2}+{(sh - win_h) // 2}")
            pwin.columnconfigure(0, weight=1)
            pwin.rowconfigure(0, weight=1)
            hbar = tk.Scrollbar(pwin, orient=tk.HORIZONTAL)
            vbar = tk.Scrollbar(pwin, orient=tk.VERTICAL)
            pcanvas = tk.Canvas(pwin, xscrollcommand=hbar.set,
                                yscrollcommand=vbar.set, bg="#888888")
            hbar.config(command=pcanvas.xview)
            vbar.config(command=pcanvas.yview)
            hbar.grid(row=1, column=0, sticky="ew")
            vbar.grid(row=0, column=1, sticky="ns")
            pcanvas.grid(row=0, column=0, sticky="nsew")
            photo = ImageTk.PhotoImage(img)
            pwin._preview_photo = photo   # keep ref alive with the window
            pcanvas.create_image(0, 0, anchor=tk.NW, image=photo)
            pcanvas.config(scrollregion=(0, 0, pw, ph))

        def _open_in_editor():
            import os, tempfile
            if not _cat_img_bytes[0]:
                messagebox.showwarning("No Image", "Load an image first.", parent=win)
                return
            try:
                img = Image.open(io.BytesIO(_cat_img_bytes[0]))
                tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                img.save(tmp.name, "PNG")
                tmp.close()
                _cat_tmp_path[0] = tmp.name
                os.startfile(tmp.name)
                btn_reload_editor.config(state=tk.NORMAL)
            except Exception as ex:
                messagebox.showerror("Error", str(ex), parent=win)

        def _reload_from_editor():
            import os
            path = _cat_tmp_path[0]
            if not path or not os.path.exists(path):
                messagebox.showwarning("Reload",
                    "No temp file found. Use 'Open in Editor' first.", parent=win)
                return
            try:
                with open(path, "rb") as fh:
                    data = fh.read()
                name = cat_img_name_var.get() or os.path.basename(path)
                _set_cat_image(data, name)
            except Exception as ex:
                messagebox.showerror("Error", str(ex), parent=win)

        btn_open_editor.config(command=_open_in_editor)
        btn_reload_editor.config(command=_reload_from_editor)

        def _pick_cat_image():
            from tkinter import filedialog
            import os
            path = filedialog.askopenfilename(
                title="Choose Category Image",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.tiff"),
                           ("All files", "*.*")],
                parent=win
            )
            if path:
                with open(path, "rb") as fh:
                    _set_cat_image(fh.read(), os.path.basename(path))

        btn_choose_img.config(command=_pick_cat_image)
        btn_clear_img.config(command=lambda: _set_cat_image(None, ""))
        btn_preview.config(command=_preview_cat_image)

        # Bottom rows of panel: image position + canvas height hint
        img_pos_frame = tk.Frame(img_panel)
        img_pos_frame.pack(fill=tk.X, pady=(0, 4))
        tk.Label(img_pos_frame, text="Image position:").pack(side=tk.LEFT)
        cat_img_pos_var = tk.StringVar(value="Row")
        ttk.Combobox(img_pos_frame, textvariable=cat_img_pos_var,
                     values=["Row", "Left Header", "Top Header"],
                     width=12, state="readonly").pack(side=tk.LEFT, padx=(6, 0))

        cat_height_var = tk.StringVar(value="")
        tk.Label(img_panel, textvariable=cat_height_var,
                 fg="gray", font=("Arial", 8), anchor=tk.W).pack(fill=tk.X)

        def _update_cat_height_hint(cat_name):
            spans = getattr(self, "cat_row_spans", {})
            rh    = getattr(self, "_row_heights", None)
            if not cat_name or cat_name not in spans or not rh:
                cat_height_var.set("")
                return
            first, last = spans[cat_name]
            rows  = [i for i in range(first, last + 1) if i < len(rh)]
            total = sum(rh[i] for i in rows)
            n     = len(rows)
            cat_height_var.set(
                f"Canvas height: {total} px  ({n} row{'s' if n != 1 else ''}"
                f",  {rh[rows[0]]} px each)" if n > 0 else ""
            )

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_frame = tk.Frame(win)
        btn_frame.grid(row=8, column=0, pady=8)

        def _selected_node():
            sel = cat_tree.selection()
            return self.db.cat_by_id.get(int(sel[0])) if sel else None

        def _all_cat_options():
            """Return [(display_label, cat_id_or_None), ...] for parent dropdown."""
            opts = [("<Root>", None)]
            def _collect(nodes):
                for n in nodes:
                    opts.append(("  " * n["depth"] + n["title"], n["id"]))
                    _collect(n["children"])
            _collect(self.db.cat_tree)
            return opts

        _parent_opts = []

        def _refresh_parent_combo(exclude_id=None):
            nonlocal _parent_opts
            _parent_opts = [(lbl, cid) for lbl, cid in _all_cat_options()
                            if cid != exclude_id]
            parent_combo["values"] = [lbl for lbl, _ in _parent_opts]

        _refresh_parent_combo()

        def refresh_tree(reselect_id=None):
            self.db.load_categories()
            cat_tree.delete(*cat_tree.get_children())
            def _insert(nodes, parent_iid=""):
                for n in nodes:
                    iid = str(n["id"])
                    cat_tree.insert(parent_iid, tk.END, iid=iid,
                                    text="  " * n["depth"] + n["title"], open=True)
                    _insert(n["children"], iid)
            _insert(self.db.cat_tree)
            if reselect_id:
                iid = str(reselect_id)
                if cat_tree.exists(iid):
                    cat_tree.selection_set(iid)
                    cat_tree.see(iid)

        refresh_tree(reselect_id=preselect_id)

        def new_cat():
            cat_tree.selection_remove(cat_tree.selection())
            id_var.set("0")
            name_var.set("")
            parent_var.set("<Root>")
            hidden_cat_var.set(0)
            show_guide_var.set(1)
            _apply_color("")
            _apply_row_bg("")
            _set_cat_image(None, "")
            cat_img_pos_var.set("Row")
            cat_pad_top_var.set(0)
            cat_pad_bottom_var.set(0)
            _refresh_parent_combo()
            _update_btn_states()

        def save_cat():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showwarning("Save", "Name cannot be empty.", parent=win); return
            chosen_label = parent_var.get()
            chosen_parent_id = next(
                (cid for lbl, cid in _parent_opts if lbl == chosen_label), None
            )
            chosen_color     = color_var.get() or None
            chosen_row_bg    = row_bg_var.get() or None
            chosen_show_guide = show_guide_var.get()
            chosen_img        = _cat_img_bytes[0]
            chosen_img_name   = cat_img_name_var.get() or None
            chosen_img_pos    = cat_img_pos_var.get() or "Row"
            chosen_pad_top    = int(cat_pad_top_var.get() or 0)
            chosen_pad_bottom = int(cat_pad_bottom_var.get() or 0)
            if id_var.get() in ("0", ""):
                # Insert new category
                new_id = self.db.add_category(new_name, parent_id=chosen_parent_id)
                with sqlite3.connect(self.db.db_file) as conn:
                    conn.execute(
                        "UPDATE Category SET color=?, row_bg_color=?, show_row_guide=?, "
                        "cat_image=?, cat_image_name=?, cat_image_pos=?, "
                        "cat_pad_top=?, cat_pad_bottom=? WHERE CategoryID=?",
                        (chosen_color, chosen_row_bg, chosen_show_guide,
                         chosen_img, chosen_img_name, chosen_img_pos,
                         chosen_pad_top, chosen_pad_bottom, new_id))
                refresh_tree(new_id)
            else:
                # Update existing category
                node = _selected_node()
                if not node:
                    return
                with sqlite3.connect(self.db.db_file) as conn:
                    if new_name != node["title"]:
                        self.db.rename_category(node["id"], new_name)
                    current_parent_id = node.get("parent_id")
                    if chosen_parent_id != current_parent_id:
                        conn.execute(
                            "UPDATE Category SET parent_id=? WHERE CategoryID=?",
                            (chosen_parent_id, node["id"])
                        )
                    conn.execute(
                        "UPDATE Category SET hidden=?, color=?, row_bg_color=?, show_row_guide=?, "
                        "cat_image=?, cat_image_name=?, cat_image_pos=?, "
                        "cat_pad_top=?, cat_pad_bottom=? WHERE CategoryID=?",
                        (hidden_cat_var.get(), chosen_color, chosen_row_bg, chosen_show_guide,
                         chosen_img, chosen_img_name, chosen_img_pos,
                         chosen_pad_top, chosen_pad_bottom, node["id"])
                    )
                refresh_tree(node["id"])
            btn_save_cat.config(text="Saved ✓", fg="green")
            win.after(2000, lambda: btn_save_cat.config(text="Save", fg=""))

        def delete_cat():
            node = _selected_node()
            if not node:
                messagebox.showwarning("Delete", "Select a category to delete.", parent=win); return
            count       = sum(1 for e in self.db.events if e.get("categoryid") == node["id"])
            child_count = len(node["children"])
            msg = f"Delete '{node['title']}'?"
            if count:       msg += f"\n\n{count} event(s) will move to the parent category."
            if child_count: msg += f"\n{child_count} sub-categor{'y' if child_count==1 else 'ies'} will move up one level."
            if not messagebox.askyesno("Confirm Delete", msg, parent=win): return
            self.db.delete_category(node["id"])
            name_var.set(""); refresh_tree()

        def move_up():
            node = _selected_node()
            if node: self.db.reorder_category(node["id"], -1); refresh_tree(node["id"])

        def move_down():
            node = _selected_node()
            if node: self.db.reorder_category(node["id"],  1); refresh_tree(node["id"])

        def open_events_dialog():
            initial_node = _selected_node()
            if not initial_node:
                return
            ewin = tk.Toplevel(win)
            ewin.title(f"Events — {initial_node['title']}")
            ewin.resizable(True, True)

            # current_node is mutable via nonlocal in nested functions
            current_node = initial_node

            # ── Category selector ─────────────────────────────────────────────
            cat_bar = tk.Frame(ewin)
            cat_bar.pack(fill=tk.X, padx=12, pady=(8, 0))
            tk.Label(cat_bar, text="Category:").pack(side=tk.LEFT)
            cat_paths    = [n["path"] for n in self.db.cat_nodes]
            cat_path_map = {n["path"]: n for n in self.db.cat_nodes}
            cat_var      = tk.StringVar(value=initial_node["path"])
            cat_combo    = ttk.Combobox(cat_bar, textvariable=cat_var,
                                        values=cat_paths, state="readonly", width=34)
            cat_combo.pack(side=tk.LEFT, padx=(6, 0))

            # ── Event list ────────────────────────────────────────────────────
            list_frame = tk.Frame(ewin, padx=12, pady=8)
            list_frame.pack(fill=tk.BOTH, expand=True)
            list_frame.columnconfigure(0, weight=1)
            list_frame.rowconfigure(0, weight=1)

            vsb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
            vsb.grid(row=0, column=1, sticky="ns")
            ev_tree = ttk.Treeview(
                list_frame,
                columns=("title", "start", "end"),
                show="headings",
                yscrollcommand=vsb.set,
                height=14,
                selectmode="browse",
            )
            vsb.config(command=ev_tree.yview)
            ev_tree.heading("title", text="Title")
            ev_tree.heading("start", text="Start Date")
            ev_tree.heading("end",   text="End Date")
            ev_tree.column("title", width=220, stretch=True)
            ev_tree.column("start", width=130, stretch=False)
            ev_tree.column("end",   width=130, stretch=False)
            ev_tree.grid(row=0, column=0, sticky="nsew")

            def _cat_events():
                return sorted(
                    [e for e in self.db.events if e.get("categoryid") == current_node["id"]],
                    key=lambda e: (e.get("sort_order") or 0, e.get("start_value") or 0),
                )

            def _refresh(reselect_id=None):
                ev_tree.delete(*ev_tree.get_children())
                for e in _cat_events():
                    ev_tree.insert("", tk.END, iid=str(e["id"]),
                                   values=(e.get("title", ""),
                                           e.get("start_display") or "",
                                           e.get("end_display") or ""))
                if reselect_id and ev_tree.exists(str(reselect_id)):
                    ev_tree.selection_set(str(reselect_id))
                    ev_tree.see(str(reselect_id))

            def _on_cat_change(*_):
                nonlocal current_node
                new_node = cat_path_map.get(cat_var.get())
                if new_node and new_node["id"] != current_node["id"]:
                    current_node = new_node
                    ewin.title(f"Events — {current_node['title']}")
                    _refresh()
                    _on_ev_select()

            cat_combo.bind("<<ComboboxSelected>>", _on_cat_change)

            _refresh()

            def _selected_ev():
                sel = ev_tree.selection()
                if not sel:
                    return None
                eid = int(sel[0])
                return next((e for e in self.db.events if e["id"] == eid), None)

            def _move(direction):
                e = _selected_ev()
                if not e:
                    return
                events = _cat_events()
                idx = next((i for i, ev in enumerate(events) if ev["id"] == e["id"]), None)
                if idx is None:
                    return
                other_idx = idx + direction
                if other_idx < 0 or other_idx >= len(events):
                    return
                with sqlite3.connect(self.db.db_file) as conn:
                    # Normalize sort_orders first so no two events share a value or are NULL
                    for i, ev in enumerate(events):
                        if ev.get("sort_order") != i:
                            conn.execute("UPDATE events SET sort_order=? WHERE id=?", (i, ev["id"]))
                            ev["sort_order"] = i
                    a, b = events[idx], events[other_idx]
                    conn.execute("UPDATE events SET sort_order=? WHERE id=?", (b["sort_order"], a["id"]))
                    conn.execute("UPDATE events SET sort_order=? WHERE id=?", (a["sort_order"], b["id"]))
                a["sort_order"], b["sort_order"] = b["sort_order"], a["sort_order"]
                _refresh(reselect_id=e["id"])
                self._draw()

            def _edit_event():
                e = _selected_ev()
                if not e:
                    return
                dlg = self._open_edit_event_dialog(e)
                ewin.wait_window(dlg)
                _refresh()

            def _add_event():
                blank = {
                    "id": 0, "title": "", "categoryid": current_node["id"],
                    "start_unit": "CE", "start_value": None,
                    "start_month": None, "start_day": None,
                    "end_unit": "CE", "end_value": None,
                    "end_month": None, "end_day": None,
                    "desc": "", "url": "", "standalone": 1,
                    "image": None, "image_name": None, "image_type": None,
                }
                dlg = self._open_edit_event_dialog(blank)
                ewin.wait_window(dlg)
                _refresh()

            # ── Buttons ───────────────────────────────────────────────────────
            ebtn_frame = tk.Frame(ewin)
            ebtn_frame.pack(pady=(0, 8))

            btn_ev_add  = tk.Button(ebtn_frame, text="Add Event", width=11, command=_add_event)
            btn_ev_edit = tk.Button(ebtn_frame, text="Edit",      width=11, state=tk.DISABLED, command=_edit_event)
            btn_ev_up   = tk.Button(ebtn_frame, text="↑ Up",      width=11, state=tk.DISABLED, command=lambda: _move(-1))
            btn_ev_down = tk.Button(ebtn_frame, text="↓ Down",    width=11, state=tk.DISABLED, command=lambda: _move(1))
            btn_ev_close= tk.Button(ebtn_frame, text="Close",     width=11, command=ewin.destroy)
            for b in [btn_ev_add, btn_ev_edit, btn_ev_up, btn_ev_down, btn_ev_close]:
                b.pack(side=tk.LEFT, padx=4)

            _ev_sel_btns = [btn_ev_edit, btn_ev_up, btn_ev_down]

            def _on_ev_select(*_):
                state = tk.NORMAL if ev_tree.selection() else tk.DISABLED
                for b in _ev_sel_btns:
                    b.config(state=state)

            ev_tree.bind("<<TreeviewSelect>>", _on_ev_select)
            ev_tree.bind("<Double-1>", lambda *_: _edit_event())

            ewin.update_idletasks()
            sw, sh = ewin.winfo_screenwidth(), ewin.winfo_screenheight()
            w = max(ewin.winfo_reqwidth(), 540)
            h = max(ewin.winfo_reqheight(), 320)
            ewin.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        def _close():
            win.destroy()
            if on_close:
                on_close()
            self._reload(self.db.active_timeline_id)

        btn_row = tk.Frame(btn_frame); btn_row.pack(pady=2)
        tk.Button(btn_row,  text="New",    command=new_cat,         width=11).pack(side=tk.LEFT, padx=4)
        btn_save_cat = tk.Button(btn_row, text="Save", command=save_cat, width=11)
        btn_save_cat.pack(side=tk.LEFT, padx=4)
        btn_delete_cat = tk.Button(btn_row, text="Delete", command=delete_cat, width=11); btn_delete_cat.pack(side=tk.LEFT, padx=4)
        btn_events_cat = tk.Button(btn_row, text="Events", command=open_events_dialog, width=11, state=tk.DISABLED); btn_events_cat.pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row,  text="Close",  command=_close,          width=11).pack(side=tk.LEFT, padx=4)

        btn_row2 = tk.Frame(btn_frame); btn_row2.pack(pady=(0, 2))
        btn_up   = tk.Button(btn_row2, text="↑ Up",   command=move_up,   width=11); btn_up.pack(side=tk.LEFT, padx=4)
        btn_down = tk.Button(btn_row2, text="↓ Down", command=move_down, width=11); btn_down.pack(side=tk.LEFT, padx=4)

        _sel_btns = [btn_delete_cat, btn_events_cat, btn_up, btn_down]

        def _update_btn_states(*_):
            state = tk.NORMAL if cat_tree.selection() else tk.DISABLED
            for b in _sel_btns:
                b.config(state=state)

        def _on_tree_select(*_):
            node = _selected_node()
            if node:
                id_var.set(str(node["id"]))
                name_var.set(node["title"])
                _refresh_parent_combo(exclude_id=node["id"])
                parent_id = node.get("parent_id")
                label = next((lbl for lbl, cid in _parent_opts if cid == parent_id), "<Root>")
                parent_var.set(label)
                hidden_cat_var.set(1 if node.get("hidden") else 0)
                show_guide_var.set(1 if node.get("show_row_guide", True) else 0)
                _apply_color(node.get("color") or "")
                _apply_row_bg(node.get("row_bg_color") or "")
                _set_cat_image(node.get("cat_image"), node.get("cat_image_name") or "")
                cat_img_pos_var.set(node.get("cat_image_pos") or "Row")
                cat_pad_top_var.set(int(node.get("cat_pad_top") or 0))
                cat_pad_bottom_var.set(int(node.get("cat_pad_bottom") or 0))
                _update_cat_height_hint(node["title"])
            else:
                id_var.set("")
                name_var.set("")
                parent_var.set("<Root>")
                hidden_cat_var.set(0)
                show_guide_var.set(1)
                _apply_color("")
                _apply_row_bg("")
                _set_cat_image(None, "")
                cat_img_pos_var.set("Row")
                cat_pad_top_var.set(0)
                cat_pad_bottom_var.set(0)
                _update_cat_height_hint(None)
                _refresh_parent_combo()
            _update_btn_states()

        _update_btn_states()
        cat_tree.bind("<<TreeviewSelect>>", _on_tree_select)

        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        w = min(win.winfo_reqwidth(),  sw - 40)
        h = min(win.winfo_reqheight(), sh - 80)
        x = (sw - w) // 2
        y = max(20, (sh - h) // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")

    def _open_add_event_dialog(self):
        """Open the event editor pre-filled with blank/default values for a new event."""
        first_cat_id = self.db.cat_nodes[0]["id"] if self.db.cat_nodes else None
        blank = {
            "id": 0, "title": "", "categoryid": first_cat_id,
            "start_unit": "CE", "start_value": None, "start_month": None, "start_day": None,
            "end_unit":   "CE", "end_value":   None, "end_month":   None, "end_day":   None,
            "desc": "", "url": "", "standalone": 1,
            "image": None, "image_name": None, "image_type": None,
        }
        self._open_edit_event_dialog(blank)

    def _open_edit_event_dialog(self, evt):
        """Open a full-featured Toplevel edit dialog matching the main editor panel."""
        dlg = tk.Toplevel(self.win)
        is_new = evt.get("id", 0) == 0
        dlg.title("Add Event" if is_new else f"Edit Event — {evt['title']}")
        dlg.resizable(True, True)
        dlg.grab_set()

        f = tk.LabelFrame(dlg, text="Event", padx=10, pady=10)
        f.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        pad = dict(padx=5, pady=3)

        # ── ID (readonly) ──────────────────────────────────────────────────────
        tk.Label(f, text="ID:").grid(row=0, column=0, sticky=tk.W, **pad)
        id_var = tk.StringVar(value=str(evt.get("id", "")))
        tk.Entry(f, textvariable=id_var, width=10, state="readonly").grid(row=0, column=1, sticky=tk.W, **pad)

        # ── Title ──────────────────────────────────────────────────────────────
        tk.Label(f, text="Title:").grid(row=1, column=0, sticky=tk.W, **pad)
        title_var = tk.StringVar(value=evt["title"])
        title_entry = tk.Entry(f, textvariable=title_var, width=50)
        title_entry.grid(row=1, column=1, sticky=tk.W, **pad)

        # ── Date row builder ───────────────────────────────────────────────────
        def make_date_row(grid_row, label, prefix):
            tk.Label(f, text=label).grid(row=grid_row, column=0, sticky=tk.W, **pad)
            frame = tk.Frame(f)
            frame.grid(row=grid_row, column=1, sticky=tk.W, **pad)

            num_var   = tk.StringVar()
            unit_var  = tk.StringVar(value="CE")
            month_var = tk.StringVar(value="")
            day_var   = tk.StringVar(value="")

            num_entry = tk.Entry(frame, textvariable=num_var, width=10)
            num_entry.pack(side=tk.LEFT, padx=(0, 4))
            unit_combo = ttk.Combobox(frame, textvariable=unit_var,
                                      values=["CE", "BCE", "MYA", "BYA", "Present"],
                                      width=7, state="readonly")
            unit_combo.pack(side=tk.LEFT, padx=(0, 4))
            month_combo = ttk.Combobox(frame, textvariable=month_var,
                                       values=MONTHS, width=10, state="readonly")
            month_combo.pack(side=tk.LEFT, padx=(0, 4))
            day_combo = ttk.Combobox(frame, textvariable=day_var,
                                     values=[""] + [str(i) for i in range(1, 32)],
                                     width=4, state="readonly")
            day_combo.pack(side=tk.LEFT)

            def on_unit_change(*_):
                u = unit_var.get()
                if u == "Present":
                    num_entry.pack_forget(); month_combo.pack_forget(); day_combo.pack_forget()
                    num_var.set(""); month_var.set(""); day_var.set("")
                elif u in ("MYA", "BYA"):
                    num_entry.pack(side=tk.LEFT, padx=(0, 4))
                    month_combo.pack_forget(); day_combo.pack_forget()
                    month_var.set(""); day_var.set("")
                else:
                    num_entry.pack(side=tk.LEFT, padx=(0, 4))
                    month_combo.pack(side=tk.LEFT, padx=(0, 4))
                    day_combo.pack(side=tk.LEFT)
            unit_var.trace_add("write", on_unit_change)

            unit = evt.get(f"{prefix}_unit") or "CE"
            unit_var.set(unit)
            if unit != "Present":
                raw = evt.get(f"{prefix}_value")
                if raw is not None:
                    if unit == "CE":    num_var.set(str(int(raw)))
                    elif unit == "BCE": num_var.set(str(int(-raw)))
                    elif unit == "MYA": num_var.set(f"{(-raw/1_000_000):g}")
                    elif unit == "BYA": num_var.set(f"{(-raw/1_000_000_000):g}")
                m = evt.get(f"{prefix}_month") or 0
                month_var.set(MONTHS[m] if 0 < m < len(MONTHS) else "")
                d = evt.get(f"{prefix}_day") or 0
                day_var.set(str(d) if d else "")
            on_unit_change()
            return num_var, unit_var, month_var, day_var

        start_vars = make_date_row(2, "Start:", "start")
        end_vars   = make_date_row(3, "End:",   "end")

        # ── Category ──────────────────────────────────────────────────────────
        tk.Label(f, text="Category:").grid(row=4, column=0, sticky=tk.W, **pad)
        cat_row_frame = tk.Frame(f)
        cat_row_frame.grid(row=4, column=1, sticky=tk.W, **pad)

        cat_nodes    = self.db.cat_nodes
        display_list = [n["path"] for n in cat_nodes]
        cat_id_map   = {n["path"]: n["id"] for n in cat_nodes}
        cat_var      = tk.StringVar()
        current_node = self.db.cat_by_id.get(evt.get("categoryid"))
        cat_var.set(current_node["path"] if current_node else (display_list[0] if display_list else ""))
        cat_combo = ttk.Combobox(cat_row_frame, textvariable=cat_var,
                                 values=display_list, state="readonly", width=30)
        cat_combo.pack(side=tk.LEFT)

        def _refresh_cat_combo():
            """Reload categories into the combo after the manage dialog closes."""
            self.db.load_categories()
            new_nodes    = self.db.cat_nodes
            new_display  = [n["path"] for n in new_nodes]
            cat_id_map.clear()
            cat_id_map.update({n["path"]: n["id"] for n in new_nodes})
            cat_combo["values"] = new_display
            # Keep current selection if still valid, else pick first
            if cat_var.get() not in cat_id_map:
                cat_var.set(new_display[0] if new_display else "")

        tk.Button(cat_row_frame, text="Edit", width=6,
                  command=lambda: self._manage_categories_dialog(dlg, _refresh_cat_combo)
                  ).pack(side=tk.LEFT, padx=(4, 0))

        # ── Description ───────────────────────────────────────────────────────
        tk.Label(f, text="Desc:").grid(row=5, column=0, sticky=tk.NW, **pad)
        desc_frame = tk.Frame(f)
        desc_frame.grid(row=5, column=1, sticky=tk.W, **pad)
        desc_text = tk.Text(desc_frame, width=48, height=5, wrap=tk.WORD)
        desc_scroll = tk.Scrollbar(desc_frame, command=desc_text.yview)
        desc_text.config(yscrollcommand=desc_scroll.set)
        desc_text.pack(side=tk.LEFT)
        desc_scroll.pack(side=tk.LEFT, fill=tk.Y)
        desc_text.insert("1.0", evt.get("desc") or "")

        # ── URL ───────────────────────────────────────────────────────────────
        tk.Label(f, text="URL:").grid(row=6, column=0, sticky=tk.W, **pad)
        url_var = tk.StringVar(value=evt.get("url") or "")
        url_frame = tk.Frame(f)
        url_frame.grid(row=6, column=1, sticky=tk.W, **pad)
        tk.Entry(url_frame, textvariable=url_var, width=40).pack(side=tk.LEFT)
        test_btn   = tk.Button(url_frame, text="Test",
                               command=lambda: webbrowser.open(url_var.get().strip()), width=6)
        search_btn = tk.Button(url_frame, text="Search",
                               command=lambda: webbrowser.open("https://www.wikipedia.org"), width=6)

        def _update_url_buttons(*_):
            if url_var.get().strip():
                search_btn.pack_forget()
                test_btn.pack(side=tk.LEFT, padx=(4, 0))
            else:
                test_btn.pack_forget()
                search_btn.pack(side=tk.LEFT, padx=(4, 0))
        url_var.trace_add("write", _update_url_buttons)
        _update_url_buttons()

        # ── Standalone checkbox ────────────────────────────────────────────────
        standalone_var = tk.IntVar(value=1 if evt.get("standalone") else 0)
        tk.Checkbutton(f, text="Own row in timeline",
                       variable=standalone_var).grid(row=7, column=1, sticky=tk.W, pady=(4, 0))

        # ── Hidden checkbox ────────────────────────────────────────────────────
        hidden_evt_var = tk.IntVar(value=1 if evt.get("hidden") else 0)
        tk.Checkbutton(f, text="Hidden (exclude from timeline)",
                       variable=hidden_evt_var).grid(row=8, column=1, sticky=tk.W, pady=(2, 0))

        # ── Linked Category ───────────────────────────────────────────────────
        tk.Label(f, text="Linked Category:").grid(row=9, column=0, sticky=tk.W, **pad)
        link_cat_frame = tk.Frame(f)
        link_cat_frame.grid(row=9, column=1, sticky=tk.W, **pad)
        link_display_list = ["(none)"] + [n["path"] for n in cat_nodes]
        link_cat_id_map   = {n["path"]: n["id"] for n in cat_nodes}
        link_cat_var = tk.StringVar()
        current_linked = evt.get("linked_categoryid")
        if current_linked and current_linked in self.db.cat_by_id:
            link_cat_var.set(self.db.cat_by_id[current_linked]["path"])
        else:
            link_cat_var.set("(none)")
        ttk.Combobox(link_cat_frame, textvariable=link_cat_var,
                     values=link_display_list, state="readonly", width=30).pack(side=tk.LEFT)
        tk.Label(link_cat_frame, text="  Icon shows when event has a linked category",
                 font=("Arial", 8), fg="gray").pack(side=tk.LEFT)

        # ── Linked Timeline ───────────────────────────────────────────────────
        tk.Label(f, text="Linked Timeline:").grid(row=10, column=0, sticky=tk.W, **pad)
        link_tl_frame = tk.Frame(f)
        link_tl_frame.grid(row=10, column=1, sticky=tk.W, **pad)
        all_timelines     = self.db.load_timelines()   # [(id, title), ...]
        link_tl_id_map    = {title: tid for tid, title in all_timelines}
        link_tl_display   = ["(none)"] + [title for _, title in all_timelines]
        link_tl_var = tk.StringVar()
        current_linked_tl = evt.get("linked_timelineid")
        linked_tl_title   = next((t for tid, t in all_timelines if tid == current_linked_tl), None)
        link_tl_var.set(linked_tl_title if linked_tl_title else "(none)")
        ttk.Combobox(link_tl_frame, textvariable=link_tl_var,
                     values=link_tl_display, state="readonly", width=30).pack(side=tk.LEFT)
        tk.Label(link_tl_frame, text="  Arrow icon navigates to linked timeline",
                 font=("Arial", 8), fg="gray").pack(side=tk.LEFT)

        # ── Picture position ──────────────────────────────────────────────────
        tk.Label(f, text="Show Picture:").grid(row=11, column=0, sticky=tk.W, **pad)
        PIC_POSITIONS = ["", "Left of Event", "Center of Event", "Right of Event"]
        pic_pos_var = tk.StringVar(value=evt.get("picture_position") or "")
        ttk.Combobox(f, textvariable=pic_pos_var, values=PIC_POSITIONS,
                     state="readonly", width=18).grid(row=11, column=1, sticky=tk.W, **pad)

        # ── Image ─────────────────────────────────────────────────────────────
        tk.Label(f, text="Image:").grid(row=12, column=0, sticky=tk.NW, **pad)
        img_ctrl = tk.Frame(f)
        img_ctrl.grid(row=12, column=1, sticky=tk.W, **pad)

        img_blob      = [evt.get("image")]
        img_name      = [evt.get("image_name")]
        img_type      = [evt.get("image_type")]
        img_name_var  = tk.StringVar(value=evt.get("image_name") or "")
        tk.Label(img_ctrl, textvariable=img_name_var,
                 font=("Arial", 8), fg="gray").pack(side=tk.TOP, anchor=tk.W)
        img_btn_row = tk.Frame(img_ctrl)
        img_btn_row.pack(side=tk.TOP, anchor=tk.W)

        blank_photo  = [ImageTk.PhotoImage(Image.new("RGB", (200, 200), "#f0f0f0"))]
        preview_photo = [None]

        tk.Label(f, text="").grid(row=13, column=0)   # spacer
        preview_lbl = tk.Label(f, relief=tk.SUNKEN, image=blank_photo[0], bg="#f0f0f0")
        preview_lbl.grid(row=13, column=1, sticky=tk.W, **pad)

        def _show_preview(blob):
            if not blob:
                preview_lbl.config(image=blank_photo[0])
                return
            img = Image.open(io.BytesIO(blob))
            preview_photo[0] = ImageTk.PhotoImage(img)
            preview_lbl.config(image=preview_photo[0], width=img.width, height=img.height)

        _show_preview(img_blob[0])

        def _browse_image():
            path = filedialog.askopenfilename(
                parent=dlg, title="Select Image",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.webp"),
                           ("All files", "*.*")])
            if not path:
                return
            try:
                blob, name, fmt = TimelineDB.image_to_blob(path)
                img_blob[0] = blob; img_name[0] = name; img_type[0] = fmt
                img_name_var.set(name)
                _show_preview(blob)
            except Exception as e:
                messagebox.showerror("Error", f"Could not load image:\n{e}", parent=dlg)

        def _clear_image():
            img_blob[0] = img_name[0] = img_type[0] = None
            preview_photo[0] = None
            img_name_var.set("")
            preview_lbl.config(image=blank_photo[0])

        tk.Button(img_btn_row, text="Browse...", command=_browse_image, width=10).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(img_btn_row, text="Clear",     command=_clear_image,  width=6).pack(side=tk.LEFT)

        # ── Save / Cancel ──────────────────────────────────────────────────────
        btn_frame = tk.Frame(f, pady=6)
        btn_frame.grid(row=14, column=0, columnspan=2, pady=8)

        def _read_date(nv, uv, mv, dv):
            unit = uv.get()
            if unit == "Present":
                return _today_value(), "Present", "Present", None, None
            num_str    = nv.get().strip()
            month_name = mv.get()
            month = MONTHS.index(month_name) if month_name in MONTHS and month_name else 0
            day   = int(dv.get().strip()) if dv.get().strip() else 0
            return (_date_value(num_str, unit, month, day),
                    _date_display(num_str, unit, month, day),
                    unit, month or None, day or None)

        def _save():
            sv, sd, su, sm, sday  = _read_date(*start_vars)
            ev2, ed, eu, em, eday = _read_date(*end_vars)
            combo_val = cat_var.get()
            cat_id    = cat_id_map.get(combo_val, evt.get("categoryid"))
            new_title = title_var.get().strip()
            if not new_title:
                messagebox.showwarning("Save Event", "Title cannot be empty.", parent=dlg)
                return
            # End date before start date
            if sv is not None and ev2 is not None and ev2 < sv:
                messagebox.showwarning(
                    "Save Event",
                    "End date is earlier than the start date.\n"
                    "Please correct the dates before saving.",
                    parent=dlg)
                return
            # Warn if event falls outside the timeline's ruler range
            if not hidden_evt_var.get() and sv is not None:
                ruler_min, ruler_max, ruler_max_present = \
                    self.db.load_timeline_ruler(self.db.active_timeline_id)
                eff_ruler_max = _today_value() if ruler_max_present else ruler_max
                if ruler_min is not None or eff_ruler_max is not None:
                    issues = []
                    if ruler_min is not None and sv < ruler_min:
                        issues.append(f"  \u2022 Start ({sd}) is before ruler start "
                                      f"({_date_val_to_display(ruler_min)}).")
                    if eff_ruler_max is not None and sv > eff_ruler_max:
                        issues.append(f"  \u2022 Start ({sd}) is after ruler end "
                                      f"({'Present' if ruler_max_present else _date_val_to_display(eff_ruler_max)}).")
                    if ev2 is not None:
                        if eff_ruler_max is not None and ev2 > eff_ruler_max:
                            issues.append(f"  \u2022 End ({ed}) extends past ruler end "
                                          f"({'Present' if ruler_max_present else _date_val_to_display(eff_ruler_max)}).")
                        if ruler_min is not None and ev2 < ruler_min:
                            issues.append(f"  \u2022 End ({ed}) is before ruler start "
                                          f"({_date_val_to_display(ruler_min)}).")
                    if issues:
                        msg = ("This event falls outside the timeline's ruler range "
                               "and may not be visible:\n\n"
                               + "\n".join(issues)
                               + "\n\nSave anyway?")
                        if not messagebox.askyesno("Event Outside Ruler Range", msg, parent=dlg):
                            return
            desc = desc_text.get("1.0", tk.END).strip()
            url  = url_var.get().strip()
            lc_val = link_cat_var.get()
            linked_catid = link_cat_id_map.get(lc_val) if lc_val != "(none)" else None
            lt_val = link_tl_var.get()
            linked_tlid  = link_tl_id_map.get(lt_val) if lt_val != "(none)" else None
            with sqlite3.connect(self.db.db_file) as conn:
                if is_new:
                    next_order = conn.execute(
                        "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM events "
                        "WHERE categoryid=? AND timelineid=?",
                        (cat_id, self.db.active_timeline_id)
                    ).fetchone()[0]
                    conn.execute(
                        "INSERT INTO events "
                        "(title, year, desc, categoryid, timelineid, url, "
                        "image, image_name, image_type, "
                        "start_value, start_display, start_unit, start_month, start_day, "
                        "end_value,   end_display,   end_unit,   end_month,   end_day,   "
                        "standalone, sort_order, hidden, picture_position, "
                        "linked_categoryid, linked_timelineid) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (new_title, int(sv) if sv else 0,
                         desc, cat_id, self.db.active_timeline_id, url,
                         img_blob[0], img_name[0], img_type[0],
                         sv, sd, su, sm, sday,
                         ev2, ed, eu, em, eday,
                         standalone_var.get(), next_order, hidden_evt_var.get(),
                         pic_pos_var.get() or None, linked_catid, linked_tlid)
                    )
                else:
                    conn.execute(
                        "UPDATE events SET title=?, year=?, desc=?, categoryid=?, url=?, "
                        "image=?, image_name=?, image_type=?, "
                        "start_value=?, start_display=?, start_unit=?, start_month=?, start_day=?, "
                        "end_value=?, end_display=?, end_unit=?, end_month=?, end_day=?, "
                        "standalone=?, hidden=?, picture_position=?, "
                        "linked_categoryid=?, linked_timelineid=? WHERE id=?",
                        (new_title, int(sv) if sv else evt.get("year", 0),
                         desc, cat_id, url,
                         img_blob[0], img_name[0], img_type[0],
                         sv, sd, su, sm, sday,
                         ev2, ed, eu, em, eday,
                         standalone_var.get(), hidden_evt_var.get(),
                         pic_pos_var.get() or None, linked_catid, linked_tlid, evt["id"])
                    )
            dlg.destroy()
            self._reload(self.db.active_timeline_id)

        def _delete():
            if not messagebox.askyesno(
                "Delete Event",
                f"Delete \"{evt.get('title', '')}\"?\n\nThis cannot be undone.",
                parent=dlg,
            ):
                return
            with sqlite3.connect(self.db.db_file) as conn:
                conn.execute("DELETE FROM events WHERE id=?", (evt["id"],))
            dlg.destroy()
            self._reload(self.db.active_timeline_id)

        tk.Button(btn_frame, text="Save",   width=12, command=_save).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="Delete", width=12, command=_delete,
                  state=tk.DISABLED if is_new else tk.NORMAL).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="Close", width=12, command=dlg.destroy).pack(side=tk.LEFT, padx=6)

        dlg.update_idletasks()
        wx = self.win.winfo_x() + (self.win.winfo_width()  - dlg.winfo_width())  // 2
        wy = self.win.winfo_y() + (self.win.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{max(0,wx)}+{max(0,wy)}")
        title_entry.focus_set()
        return dlg

    # ── coordinate helpers ────────────────────────────────────────────────────

    def _compressed_offset(self, d):
        """Pixels saved by all breaks whose start is before date d."""
        total = 0.0
        for b in getattr(self, '_breaks', []):
            bs, be = b['start'], b['end']
            span = be - bs
            if span <= 0 or d <= bs:
                continue
            if d >= be:
                total += max(0.0, span * self.px_per_year - self.BREAK_STUB_PX)
            else:
                # d is inside this break — compress the portion crossed
                total += (d - bs) * self.px_per_year
        return total

    def _total_break_savings(self):
        """Total pixels removed by all breaks."""
        total = 0.0
        for b in getattr(self, '_breaks', []):
            span = b['end'] - b['start']
            if span > 0:
                total += max(0.0, span * self.px_per_year - self.BREAK_STUB_PX)
        return total

    def _x(self, date_val):
        raw = self.CANVAS_MARGIN + (date_val - self.min_date) * self.px_per_year
        return raw - self._compressed_offset(date_val)

    def _canvas_width(self):
        return (self.CANVAS_MARGIN
                + (self.max_date - self.min_date) * self.px_per_year
                - self._total_break_savings()
                + self.CANVAS_MARGIN)

    def _date_from_x(self, cx):
        """Invert _x(): canvas pixel x → date value, respecting breaks."""
        breaks = sorted(getattr(self, '_breaks', []), key=lambda b: b['start'])
        d_cursor = self.min_date
        for b in breaks:
            bs, be = b['start'], b['end']
            if bs <= d_cursor:
                d_cursor = max(d_cursor, be)
                continue
            x_seg_start = self._x(d_cursor)
            x_seg_end   = self._x(bs)
            if cx <= x_seg_end:
                if x_seg_end <= x_seg_start:
                    return d_cursor
                frac = (cx - x_seg_start) / (x_seg_end - x_seg_start)
                return d_cursor + frac * (bs - d_cursor)
            x_stub_end = x_seg_end + self.BREAK_STUB_PX
            if cx <= x_stub_end:
                return (bs + be) / 2
            d_cursor = be
        # Final segment after all breaks
        x_seg_start = self._x(d_cursor)
        x_seg_end   = self._x(self.max_date)
        if x_seg_end <= x_seg_start:
            return d_cursor
        frac = (cx - x_seg_start) / max(x_seg_end - x_seg_start, 1)
        return d_cursor + max(0.0, min(1.0, frac)) * (self.max_date - d_cursor)

    def _lane_y(self, lane_index):
        rh = getattr(self, "_row_heights", None)
        if not rh:
            return lane_index * (self.LANE_HEIGHT + self.LANE_PAD)
        return sum(rh[i] + self.LANE_PAD for i in range(lane_index))

    def _total_height(self):
        return self._lane_y(len(self.rows)) + 20

    def _compute_label_tiers(self):
        """Assign vertical tiers to event labels in rows where titles would overlap."""
        import tkinter.font as tkfont
        self._label_tiers     = {}   # event_id  -> tier index
        self._row_label_tiers = {}   # row_idx   -> number of tiers used
        try:
            font = tkfont.Font(family="Arial", size=9)
        except Exception:
            return
        GAP = 6   # minimum horizontal gap between adjacent labels (px)
        for row_idx, slot in enumerate(self.rows):
            # Only stagger rows with multiple non-image events
            if (slot[0].get("_cat_header") or slot[0].get("_cat_pad")
                    or slot[0].get("_collapsed") or len(slot) <= 1
                    or any(e.get("picture_position") for e in slot)):
                self._row_label_tiers[row_idx] = 1
                continue
            # Compute the x-range each label will occupy
            ranges = []
            for e in slot:
                sv = e.get("start_value")
                if sv is None:
                    continue
                ev_val = e.get("end_value")
                x1 = self._x(sv)
                tw = font.measure(e.get("title") or "")
                if ev_val is not None and ev_val != sv:
                    cx = (x1 + self._x(ev_val)) / 2
                else:
                    cx = x1
                ranges.append((e, cx - tw / 2, cx + tw / 2))
            if not ranges:
                self._row_label_tiers[row_idx] = 1
                continue
            # Sort by label left edge, then greedily assign tiers
            ranges.sort(key=lambda t: t[1])
            tier_ends = []   # rightmost x used so far in each tier
            for e, lx1, lx2 in ranges:
                placed = False
                for t, t_end in enumerate(tier_ends):
                    if lx1 >= t_end + GAP:
                        tier_ends[t] = lx2
                        self._label_tiers[e.get("id")] = t
                        placed = True
                        break
                if not placed:
                    self._label_tiers[e.get("id")] = len(tier_ends)
                    tier_ends.append(lx2)
            self._row_label_tiers[row_idx] = len(tier_ends)

    def _compute_row_heights(self):
        """Set self._row_heights: per-row pixel height (≥ LANE_HEIGHT, grows to fit any shown image)."""
        self._compute_label_tiers()
        self._row_heights = []
        for row_idx, slot in enumerate(self.rows):
            if slot[0].get("_cat_pad"):
                self._row_heights.append(slot[0]["_pad_h"])
                continue
            if slot[0].get("_cat_header"):
                cat_name = slot[0].get("category", "")
                cat_img_blob = getattr(self, "cat_image", {}).get(cat_name)
                cat_img_pos  = getattr(self, "cat_image_pos", {}).get(cat_name, "Row")
                hdr_h = self.CAT_HEADER_H
                if cat_img_blob and cat_img_pos == "Top Header":
                    try:
                        img_obj = Image.open(io.BytesIO(cat_img_blob))
                        hdr_h = max(self.CAT_HEADER_H, img_obj.height)
                    except Exception:
                        pass
                self._row_heights.append(hdr_h)
                continue
            h = self.LANE_HEIGHT
            for e in slot:
                if e.get("picture_position") and e.get("image"):
                    try:
                        img = Image.open(io.BytesIO(e["image"]))
                        pad = self.IMG_PAD * 2
                        h = max(h, img.height + pad)
                    except Exception:
                        pass
            # Extra height for staggered label tiers.
            # For LINE icons cap_top sits LANE_HEIGHT//2 below bar_y (empty space),
            # so we can reclaim that space and place the tiers closer to the icon.
            n_tiers = self._row_label_tiers.get(row_idx, 1)
            if n_tiers >= 2:
                line_offset = self.LANE_HEIGHT // 2  # always Line style
                h = max(h, n_tiers * self.LABEL_TIER_H + self.LANE_HEIGHT - line_offset + 11)
            self._row_heights.append(h)

    # ── drawing ───────────────────────────────────────────────────────────────

    def _draw_header_text(self, ry, cat_name, hdr_h=None):
        """Draw the category name inside a header bar at row y=ry."""
        c   = self.canvas
        ty  = ry + (hdr_h if hdr_h is not None else self.CAT_HEADER_H) // 2
        pos = self._cat_header_title_pos
        if pos == "Left":
            c.create_text(8, ty, text=cat_name, fill="white",
                          font=("Arial", 11, "bold"), anchor=tk.W,
                          tags=("cat_header_text",))
        elif pos == "Left (View)":
            vl = self.canvas.xview()[0] * self._canvas_width()
            c.create_text(vl + 8, ty, text=cat_name, fill="white",
                          font=("Arial", 11, "bold"), anchor=tk.W,
                          tags=("cat_header_text",))
        elif pos == "Center":
            c.create_text(self._canvas_width() / 2, ty, text=cat_name,
                          fill="white", font=("Arial", 11, "bold"),
                          anchor=tk.CENTER, tags=("cat_header_text",))
        else:  # Center (View)
            vl = self.canvas.xview()[0] * self._canvas_width()
            vw = self.canvas.winfo_width()
            c.create_text(vl + vw / 2, ty, text=cat_name, fill="white",
                          font=("Arial", 11, "bold"), anchor=tk.CENTER,
                          tags=("cat_header_text",))

    def _update_floating_headers(self):
        """Re-stamp floating header texts at the current viewport position."""
        if not getattr(self, "_header_row_data", None):
            return
        self.canvas.delete("cat_header_text")
        for entry in self._header_row_data:
            ry, cat_name = entry[0], entry[1]
            hdr_h = entry[3] if len(entry) > 3 else None
            self._draw_header_text(ry, cat_name, hdr_h)

    def _update_label_column(self):
        """Show or hide the left label column based on _cat_header_style."""
        show   = (self._cat_header_style in ("Left", "Both"))
        mapped = self.label_canvas.winfo_ismapped()
        if show and not mapped:
            self._label_corner.pack(side=tk.LEFT, fill=tk.Y,
                                    before=self.ruler_canvas)
            self.label_canvas.pack(side=tk.LEFT, fill=tk.Y,
                                   before=self.canvas)
        elif not show and mapped:
            self._label_corner.pack_forget()
            self.label_canvas.pack_forget()

    def _draw(self):
        self._update_label_column()
        import math
        disp_val, disp_unit = _px_per_year_to_display(self.px_per_year)
        self._scale_var.set(f"{disp_val:.3g}")
        self._scale_unit_var.set(disp_unit)
        c = self.canvas
        c.delete("all")
        self._event_images = []   # keep PhotoImage refs alive for this draw cycle
        self._cat_images   = []   # keep category PhotoImage refs alive for this draw cycle
        self._compute_row_heights()
        W = self._canvas_width()
        H = self._total_height()
        c.configure(scrollregion=(0, 0, W, H))

        # Compute tick interval once so ruler and grid lines are guaranteed identical
        try:
            viewport_px = self.ruler_canvas.winfo_width()
        except Exception:
            viewport_px = 800
        visible_span = max(viewport_px / self.px_per_year, 0.0001)
        self._tick_interval_cache = self._tick_interval(visible_span)

        # Ruler canvas: same width, horizontally scrollable only
        self.ruler_canvas.delete("all")
        self.ruler_canvas.configure(scrollregion=(0, 0, W, self.RULER_HEIGHT))
        self._draw_ruler(W)

        # Label canvas: fixed width, vertically scrollable only
        self.label_canvas.delete("all")
        self.label_canvas.configure(scrollregion=(0, 0, self.LABEL_WIDTH, H))

        self._draw_timeline_bg_image(H)
        self._draw_lane_backgrounds(W)
        self._draw_grid_lines(H)

        self._draw_events()
        self._draw_lane_labels()
        self._update_status()

    def _draw_timeline_bg_image(self, total_h):
        """Draw the timeline background image anchored left: top, bottom, or tiled."""
        blob = getattr(self, "_tl_bg_image_blob", None)
        if not blob:
            return
        try:
            img_obj = Image.open(io.BytesIO(blob))
            photo   = ImageTk.PhotoImage(img_obj)
            self._tl_bg_photo = photo   # keep ref alive
            c   = self.canvas
            pos = getattr(self, "_tl_bg_image_pos", "Top")
            iw, ih = img_obj.width, img_obj.height
            if pos == "Tile":
                W = self._canvas_width()
                y = 0
                while y < total_h:
                    x = 0
                    while x < W:
                        c.create_image(x, y, anchor=tk.NW, image=photo, tags="bg")
                        x += iw
                    y += ih
            elif pos == "Bottom":
                c.create_image(0, max(0, total_h - ih), anchor=tk.NW,
                               image=photo, tags="bg")
            else:   # Top
                c.create_image(0, 0, anchor=tk.NW, image=photo, tags="bg")
        except Exception:
            pass

    def _draw_lane_backgrounds(self, W):
        c = self.canvas
        # Alternate shading per category group
        cat_index = 0
        for cat, (first, last) in self.cat_row_spans.items():
            default_bg   = "#ffffff" if cat_index % 2 == 0 else "#ebebeb"
            cat_index   += 1
            custom_bg    = getattr(self, "cat_row_bg", {}).get(cat)
            bg           = custom_bg or default_bg
            gap_color    = "#ebebeb" if default_bg == "#ffffff" else "#ffffff"
            show_guide   = getattr(self, "cat_show_guide", {}).get(cat, True)
            for row in range(first, last + 1):
                y  = self._lane_y(row)
                rh = self._row_heights[row]
                if show_guide:
                    # Gap strip uses alternating default shade so separators remain visible
                    c.create_rectangle(0, y + rh, W, y + rh + self.LANE_PAD,
                                       fill=gap_color, outline="", tags="bg")
                    # Row background
                    c.create_rectangle(0, y, W, y + rh,
                                       fill=bg, outline="", tags="bg")
            # Solid category separator
            sep_y = self._lane_y(last) + self._row_heights[last]
            c.create_line(0, sep_y, W, sep_y, fill="#aaaaaa", width=1, tags="bg")

            # Category image (Row position) — scaled to fit the category's total row height
            cat_img_blob = getattr(self, "cat_image", {}).get(cat)
            cat_img_pos  = getattr(self, "cat_image_pos", {}).get(cat, "Row")
            if cat_img_blob and cat_img_pos == "Row":
                try:
                    img_obj  = Image.open(io.BytesIO(cat_img_blob))
                    cat_y    = self._lane_y(first)
                    cat_h    = self._lane_y(last) + self._row_heights[last] - cat_y
                    if img_obj.height != cat_h and cat_h > 0:
                        scale   = cat_h / img_obj.height
                        new_w   = max(1, int(img_obj.width * scale))
                        img_obj = img_obj.resize((new_w, cat_h), Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img_obj)
                    self._cat_images.append(photo)
                    c.create_image(0, cat_y, anchor=tk.NW,
                                   image=photo, tags="bg")
                except Exception:
                    pass

    def _draw_grid_lines(self, H):
        import math
        c = self.canvas
        interval = self._tick_interval_cache
        start_tick = math.floor(self.min_date / interval) * interval
        tick = start_tick
        while tick <= self.max_date:
            x = self._x(tick)
            in_break = any(
                self._x(b['start']) <= x <= self._x(b['start']) + self.BREAK_STUB_PX
                for b in self._breaks
            )
            if not in_break:
                c.create_line(x, 0, x, H, fill="#cccccc", width=1,
                              dash=(3, 6), tags="bg")
            tick += interval

    def _zigzag_pts(self, x, y_top, y_bot, amp=4, period=10):
        """Return flat coordinate list for a vertical zig-zag line at x."""
        pts = [x, y_top]
        y = y_top
        direction = 1
        half = period / 2
        while y + half < y_bot:
            y += half
            pts.extend([x + direction * amp, y])
            direction = -direction
        pts.extend([x, y_bot])
        return pts

    def _draw_break_markers(self, H):
        """Draw zig-zag break indicators on the main canvas."""
        c = self.canvas
        for b in self._breaks:
            bs, be = b['start'], b['end']
            if be <= self.min_date or bs >= self.max_date:
                continue
            x1 = self._x(bs)
            x2 = x1 + self.BREAK_STUB_PX
            c.create_rectangle(x1, 0, x2, H, fill="#c8c8c8", outline="", tags="bg")
            pts_l = self._zigzag_pts(x1, 0, H, amp=5, period=12)
            c.create_line(*pts_l, fill="#666666", width=2, tags="bg")
            pts_r = self._zigzag_pts(x2, 0, H, amp=-5, period=12)
            c.create_line(*pts_r, fill="#666666", width=2, tags="bg")

    def _draw_ruler(self, W):
        c = self.ruler_canvas
        H = self.RULER_HEIGHT
        # Ruler background
        c.create_rectangle(0, 0, W, H, fill="#2c3e50", outline="")
        # Bottom edge
        c.create_line(0, H, W, H, fill="#1a252f", width=2)

        # Pass 1: draw break backgrounds and zig-zags
        for b in self._breaks:
            bs, be = b['start'], b['end']
            if be <= self.min_date or bs >= self.max_date:
                continue
            x1 = self._x(bs)
            x2 = x1 + self.BREAK_STUB_PX
            c.create_rectangle(x1, 0, x2, H, fill="#1a2535", outline="")
            pts_l = self._zigzag_pts(x1, 0, H, amp=4, period=10)
            c.create_line(*pts_l, fill="#aaaaaa", width=1)
            pts_r = self._zigzag_pts(x2, 0, H, amp=-4, period=10)
            c.create_line(*pts_r, fill="#aaaaaa", width=1)

        # Pass 2: draw tick marks and labels on top
        import math
        interval = self._tick_interval_cache
        start_tick = math.floor(self.min_date / interval) * interval
        tick = start_tick
        while tick <= self.max_date:
            x = self._x(tick)
            in_break = any(
                self._x(b['start']) <= x <= self._x(b['start']) + self.BREAK_STUB_PX
                for b in self._breaks
            )
            if not in_break:
                c.create_line(x, H - 10, x, H, fill="#ffffff", width=1)
                label = self._format_tick(tick)
                c.create_text(x, H - 16, text=label, fill="#ffffff",
                              font=("Arial", 8, "bold"), anchor=tk.S)
            tick += interval

        # Pass 1 already drew zig-zags; width=1, no // label needed

    def _draw_lane_labels(self):
        if self._cat_header_style not in ("Left", "Both"):
            return
        c = self.label_canvas
        for cat, (first, last) in self.cat_row_spans.items():
            color = self.cat_color[cat]
            depth = self._cat_depth.get(cat, 0)
            indent = depth * 10   # 10px per nesting level
            # Exclude leading/trailing padding rows from the label strip
            non_pad = [i for i in range(first, last + 1)
                       if not self.rows[i][0].get("_cat_pad")]
            if not non_pad:
                continue
            y1 = self._lane_y(non_pad[0])
            y2 = self._lane_y(non_pad[-1]) + self._row_heights[non_pad[-1]]
            c.create_rectangle(indent, y1, self.LABEL_WIDTH - 2, y2,
                               fill=color, outline="")
            # Left depth indicator bar for nested categories
            if indent > 0:
                c.create_rectangle(0, y1, indent, y2, fill="#1a252f", outline="")
            avail_w = self.LABEL_WIDTH - indent - 8
            cat_img_blob = getattr(self, "cat_image", {}).get(cat)
            cat_img_pos  = getattr(self, "cat_image_pos", {}).get(cat, "Row")
            if cat_img_blob and cat_img_pos == "Left Header":
                try:
                    img_obj = Image.open(io.BytesIO(cat_img_blob))
                    col_h   = y2 - y1
                    if img_obj.height != col_h and col_h > 0:
                        scale   = col_h / img_obj.height
                        new_w   = max(1, int(img_obj.width * scale))
                        img_obj = img_obj.resize((new_w, col_h), Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img_obj)
                    self._cat_images.append(photo)
                    c.create_image(indent, y1, anchor=tk.NW, image=photo)
                except Exception:
                    pass
            if self._cat_header_style in ("Left", "Both"):
                c.create_text(indent + avail_w // 2, (y1 + y2) // 2,
                              text=cat, fill="white",
                              font=("Arial", 9, "bold"),
                              width=max(avail_w, 10),
                              anchor=tk.CENTER)
        # Right border separating label column from event area
        c.create_line(self.LABEL_WIDTH - 1, 0,
                      self.LABEL_WIDTH - 1, self._total_height(),
                      fill="#555555", width=1)

    def _draw_events(self):
        import tkinter.font as tkfont
        c = self.canvas
        self._event_rects = {}
        self._link_icon_map      = {}
        self._timeline_icon_map  = {}
        self._header_row_data    = []   # (ry, cat_name, color) for floating redraws
        _evt_font = tkfont.Font(family="Arial", size=9)
        _LINK_ICON_SIZE = 10   # px square for the linked-category toggle icon

        for row_idx, slot in enumerate(self.rows):
            pseudo = slot[0]

            # Padding row — empty space with category background colour
            if pseudo.get("_cat_pad"):
                continue

            # Category header row — full-width title bar
            if pseudo.get("_cat_header"):
                cat_name = pseudo["category"]
                color  = self.cat_color.get(cat_name, "#888888")
                ry     = self._lane_y(row_idx)
                hdr_h  = self._row_heights[row_idx]
                W      = self._canvas_width()
                c.create_rectangle(0, ry, W, ry + hdr_h,
                                   fill=color, outline="", tags=("bg",))
                self._header_row_data.append((ry, cat_name, color, hdr_h))
                self._draw_header_text(ry, cat_name, hdr_h)
                # Category image in Top Header position
                cat_img_blob = getattr(self, "cat_image", {}).get(cat_name)
                cat_img_pos  = getattr(self, "cat_image_pos", {}).get(cat_name, "Row")
                if cat_img_blob and cat_img_pos == "Top Header":
                    try:
                        img_obj = Image.open(io.BytesIO(cat_img_blob))
                        photo = ImageTk.PhotoImage(img_obj)
                        self._cat_images.append(photo)
                        c.create_image(0, ry, anchor=tk.NW,
                                       image=photo, tags=("bg",))
                    except Exception:
                        pass
                continue

            # Collapsed placeholder — slot holds a single pseudo-event
            if pseudo.get("_collapsed"):
                color  = self.cat_color.get(pseudo["category"], "#888888")
                ry     = self._lane_y(row_idx)
                lane_h = self._row_heights[row_idx]
                W      = self._canvas_width()
                cy     = ry + lane_h // 2
                c.create_line(0, cy, W, cy,
                              fill=color, width=2,
                              dash=(6, 4), tags=("event",))
                c.create_text(14, cy,
                              text=f"▶  {pseudo['_count']} events  (collapsed)",
                              fill=color, font=("Arial", 8, "italic"),
                              anchor=tk.W, tags=("event_label",))
                continue

            row_y  = self._lane_y(row_idx)
            lane_h = self._row_heights[row_idx]
            n_tiers = self._row_label_tiers.get(row_idx, 1)
            stagger = (n_tiers >= 2 and
                       not any(ev.get("picture_position") for ev in slot))
            if stagger:
                line_offset = self.LANE_HEIGHT // 2  # always Line style
                extra_h = n_tiers * self.LABEL_TIER_H - line_offset + 8
            else:
                extra_h = 0
            bar_y   = row_y + extra_h   # top of the event shape area

            for e in slot:
                sv = e.get("start_value")
                if sv is None:
                    continue
                color  = self._event_color_map.get(
                             e.get("id"),
                             self.cat_color.get(e.get("category") or "General", "#888888")
                         )
                x1      = self._x(sv)
                ev_val  = e.get("end_value")
                pic_pos = e.get("picture_position") or ""
                cy_mid  = bar_y + self.LANE_HEIGHT // 2
                if stagger:
                    tier    = self._label_tiers.get(e.get("id"), 0)
                    label_y = row_y + 3 + tier * self.LABEL_TIER_H + self.LABEL_TIER_H // 2
                else:
                    label_y = cy_mid

                # For left/right image positions the shape fills the full row height
                if pic_pos in ("Left of Event", "Center of Event", "Right of Event") and e.get("image"):
                    shape_y = row_y
                    shape_h = lane_h
                else:
                    shape_y = bar_y
                    shape_h = self.LANE_HEIGHT
                shape_cy = shape_y + shape_h // 2

                # Pre-load image if a position is set
                img_obj = photo = None
                if pic_pos and e.get("image"):
                    try:
                        img_obj = Image.open(io.BytesIO(e["image"]))
                        photo   = ImageTk.PhotoImage(img_obj)
                        self._event_images.append(photo)
                    except Exception:
                        img_obj = photo = None

                if ev_val is not None and ev_val != sv:
                    # ── Duration bar ───────────────────────────────────────
                    x2    = self._x(ev_val)
                    min_w = self.LANE_HEIGHT // 2 - 2
                    if x2 - x1 < min_w:
                        mid = (x1 + x2) / 2
                        x1b, x2b = mid - min_w / 2, mid + min_w / 2
                    else:
                        x1b, x2b = x1, x2

                    if self._icon_long == "Line":
                        _vcap = self.LANE_HEIGHT // 2           # fixed cap height
                        if photo and img_obj:
                            # Event has graphic — center line and caps at row midpoint
                            line_y  = shape_cy
                            cap_top = shape_cy - _vcap // 2
                            cap_bot = shape_cy + _vcap // 2
                        elif lane_h > self.LANE_HEIGHT and not stagger:
                            # Row is expanded by another event's graphic — align to row center
                            _row_cy = row_y + lane_h // 2
                            cap_top = _row_cy - _vcap // 2
                            cap_bot = _row_cy + _vcap // 2
                            line_y  = _row_cy
                        else:
                            cap_top = shape_y + shape_h - _vcap     # top of vertical caps
                            cap_bot = shape_y + shape_h             # bottom of vertical caps
                            line_y  = cap_top + _vcap // 2          # horizontal line at mid-cap
                        rid = c.create_rectangle(x1b, cap_top, x2b, cap_bot,
                                                 fill="", outline="", tags=("event",))
                        c.create_line(x1b, line_y, x2b, line_y,
                                      fill=color, width=3, tags=("event",))
                        c.create_line(x1b, cap_top, x1b, cap_bot,
                                      fill=color, width=2, tags=("event",))
                        c.create_line(x2b, cap_top, x2b, cap_bot,
                                      fill=color, width=2, tags=("event",))
                        # Title: staggered tiers use CENTER anchor (matching BOX);
                        # non-staggered uses S anchor to sit just above cap_top.
                        if not (photo and img_obj):
                            lx = (x1b + x2b) / 2
                            tw = _evt_font.measure(e["title"])
                            if stagger:
                                c.create_text(lx, label_y,
                                              text=e["title"], fill="#222222",
                                              font=("Arial", 9), anchor=tk.CENTER,
                                              tags=("event_label",))
                                uly = label_y + 7
                                c.create_line(lx - tw / 2, uly, lx + tw / 2, uly,
                                              fill=color, width=1, tags=("event_label",))
                                c.create_line(lx, uly, lx, cap_top,
                                              fill=color, width=1, tags=("event_label",))
                            else:
                                c.create_text(lx, cap_top,
                                              text=e["title"], fill="#222222",
                                              font=("Arial", 9), anchor=tk.S,
                                              tags=("event_label",))
                    # else:  # Box (default) — commented out; all events use Line
                    #     rid = c.create_rectangle(x1b, shape_y, x2b, shape_y + shape_h,
                    #                              fill=color, outline="#333333",
                    #                              width=1, tags=("event",))
                    #     if not (photo and img_obj):
                    #         if not e.get("standalone"):
                    #             lx = (x1b + x2b) / 2
                    #             tw = _evt_font.measure(e["title"])
                    #             if stagger and tw > (x2b - x1b) - 6:
                    #                 c.create_text(lx, label_y,
                    #                               text=e["title"], fill="#222222",
                    #                               font=("Arial", 9), tags=("event_label",))
                    #                 uly = label_y + 7
                    #                 c.create_line(lx - tw / 2, uly, lx + tw / 2, uly,
                    #                               fill=color, width=1, tags=("event_label",))
                    #                 c.create_line(lx, uly, lx, bar_y,
                    #                               fill=color, width=1, tags=("event_label",))
                    #             else:
                    #                 c.create_text(lx, shape_cy,
                    #                               text=e["title"], fill="#222222",
                    #                               font=("Arial", 9), tags=("event_label",))
                    #         else:
                    #             text_w = _evt_font.measure(e["title"])
                    #             if x1b == x1 and text_w <= (x2b - x1b) - 6:
                    #                 lx = (x1b + x2b) / 2
                    #                 c.create_text(lx, label_y,
                    #                               text=e["title"], fill="#000000",
                    #                               font=("Arial", 9), tags=("event_label",))
                    #                 if stagger:
                    #                     uly = label_y + 7
                    #                     c.create_line(lx - text_w / 2, uly, lx + text_w / 2, uly,
                    #                                   fill=color, width=1, tags=("event_label",))
                    #                     c.create_line(lx, uly, lx, bar_y,
                    #                                   fill=color, width=1, tags=("event_label",))
                    #             else:
                    #                 lx = x2b + 3
                    #                 c.create_text(lx, label_y,
                    #                               text=e["title"], fill="#222222",
                    #                               font=("Arial", 9), anchor=tk.W,
                    #                               tags=("event_label",))
                    #                 if stagger:
                    #                     uly = label_y + 7
                    #                     c.create_line(lx, uly, lx + text_w, uly,
                    #                                   fill=color, width=1, tags=("event_label",))
                    #                     c.create_line(lx, uly, lx, bar_y,
                    #                                   fill=color, width=1, tags=("event_label",))

                    if photo and img_obj:
                        iw, ih = img_obj.width, img_obj.height
                        iy = row_y + lane_h - ih          # image sits on the bottom
                        if pic_pos == "Left of Event":
                            ix = x1b - iw - 10
                        elif pic_pos == "Right of Event":
                            ix = x2b + 10
                        elif pic_pos == "Center of Event":
                            ix = (x1b + x2b) / 2 - iw // 2
                        iid = c.create_image(ix, iy, anchor=tk.NW,
                                             image=photo, tags=("event",))
                        self._event_rects[iid] = e
                        # Title above the image, horizontally centered on it
                        tx = ix + iw // 2
                        ty = row_y + (lane_h - ih) // 2   # centered in space above image
                        c.create_text(tx, ty, text=e["title"],
                                      fill="#222222", font=("Arial", 9),
                                      tags=("event_label",))

                else:
                    # ── Point event ────────────────────────────────────────
                    r  = self.LANE_HEIGHT // 2 - 2
                    rw = r // 2
                    # horizontal half-width used for label offset & hit-rect
                    shape_hw = rw  # always Line style; was: r if Circle else rw

                    # if self._icon_short == "Circle":  # commented out; all events use Line
                    #     rid = c.create_oval(
                    #         x1 - r, shape_cy - r, x1 + r, shape_cy + r,
                    #         fill=color, outline="#333333", width=1, tags=("event",)
                    #     )
                    if True:  # _icon_short hardcoded to Line
                        _vcap = self.LANE_HEIGHT // 2   # fixed tick height regardless of row size
                        if photo and img_obj:
                            # Event has graphic — center tick at row midpoint
                            line_top = shape_cy - _vcap // 2
                            line_bot = shape_cy + _vcap // 2
                        elif lane_h > self.LANE_HEIGHT and not stagger:
                            # Row is expanded by another event's graphic — align to row center
                            _row_cy  = row_y + lane_h // 2
                            line_top = _row_cy - _vcap // 2
                            line_bot = _row_cy + _vcap // 2
                        else:
                            line_top = shape_y + shape_h - _vcap
                            line_bot = shape_y + shape_h
                        rid = c.create_rectangle(x1 - rw, line_top,
                                                 x1 + rw, line_bot,
                                                 fill="", outline="", tags=("event",))
                        c.create_line(x1, line_top, x1, line_bot,
                                      fill=color, width=3, tags=("event",))
                        # Title: staggered tiers use CENTER anchor (matching BOX);
                        # non-staggered uses S anchor to sit just above line_top.
                        if not (photo and img_obj):
                            tw = _evt_font.measure(e["title"])
                            if stagger:
                                c.create_text(x1, label_y,
                                              text=e["title"], fill="#222222",
                                              font=("Arial", 9), anchor=tk.CENTER,
                                              tags=("event_label",))
                                uly = label_y + 7
                                c.create_line(x1 - tw / 2, uly, x1 + tw / 2, uly,
                                              fill=color, width=1, tags=("event_label",))
                                c.create_line(x1, uly, x1, line_top,
                                              fill=color, width=1, tags=("event_label",))
                            else:
                                c.create_text(x1, line_top,
                                              text=e["title"], fill="#222222",
                                              font=("Arial", 9), anchor=tk.S,
                                              tags=("event_label",))
                    # else:  # Diamond (default) — commented out; all events use Line
                    #     rid = c.create_polygon(
                    #         x1,      shape_cy - r,
                    #         x1 + rw, shape_cy,
                    #         x1,      shape_cy + r,
                    #         x1 - rw, shape_cy,
                    #         fill=color, outline="#333333", width=1, tags=("event",)
                    #     )
                    # if self._icon_short != "Line" and not (photo and img_obj):  # always False now
                    #     if not e.get("standalone"):
                    #         tw = _evt_font.measure(e["title"])
                    #         c.create_text(x1, label_y,
                    #                       text=e["title"], fill="#222222",
                    #                       font=("Arial", 9), anchor=tk.CENTER,
                    #                       tags=("event_label",))
                    #         if stagger:
                    #             uly = label_y + 7
                    #             c.create_line(x1 - tw / 2, uly, x1 + tw / 2, uly,
                    #                           fill=color, width=1, tags=("event_label",))
                    #             c.create_line(x1, uly, x1, shape_cy - r,
                    #                           fill=color, width=1, tags=("event_label",))
                    #     else:
                    #         tw = _evt_font.measure(e["title"])
                    #         lx = x1 + shape_hw + 3
                    #         c.create_text(lx, label_y,
                    #                       text=e["title"], fill="#222222",
                    #                       font=("Arial", 9), anchor=tk.W,
                    #                       tags=("event_label",))
                    #         if stagger:
                    #             uly = label_y + 7
                    #             c.create_line(lx, uly, lx + tw, uly,
                    #                           fill=color, width=1, tags=("event_label",))
                    #             c.create_line(lx, uly, x1 + shape_hw + 1, shape_cy - r,
                    #                           fill=color, width=1, tags=("event_label",))

                    if photo and img_obj:
                        iw, ih = img_obj.width, img_obj.height
                        iy = row_y + lane_h - ih          # image sits on the bottom
                        if pic_pos == "Left of Event":
                            ix = x1 - rw - iw - 10
                        elif pic_pos == "Right of Event":
                            ix = x1 + rw + 10
                        elif pic_pos == "Center of Event":
                            ix = x1 - iw // 2
                        iid = c.create_image(ix, iy, anchor=tk.NW,
                                             image=photo, tags=("event",))
                        self._event_rects[iid] = e
                        # Title above the image, horizontally centered on it
                        tx = ix + iw // 2
                        ty = row_y + (lane_h - ih) // 2   # centered in space above image
                        c.create_text(tx, ty, text=e["title"],
                                      fill="#222222", font=("Arial", 9),
                                      tags=("event_label",))

                self._event_rects[rid] = e

                # ── Icons to the right of the event title ─────────────────
                linked_id    = e.get("linked_categoryid")
                linked_tlid  = e.get("linked_timelineid")
                if linked_id or linked_tlid:
                    # Compute title anchor and icon baseline (shared by all icons)
                    _tw = _evt_font.measure(e["title"])
                    if photo and img_obj:
                        _title_lx = tx - _tw / 2
                        _icon_y   = ty
                    elif ev_val is not None and ev_val != sv:
                        # Duration event: Line-style title always centered at bar midpoint.
                        # stagger → CENTER anchor: label_y is text center.
                        # non-stagger → S anchor: label_y is text bottom; shift icon up
                        # so its bottom aligns with the text bottom.
                        _icon_y   = label_y if stagger else label_y - _LINK_ICON_SIZE // 2
                        _title_lx = (x1b + x2b) / 2 - _tw / 2
                    else:
                        # Point event: Line-style title always centered at x1.
                        _icon_y   = label_y if stagger else label_y - _LINK_ICON_SIZE // 2
                        _title_lx = x1 - _tw / 2

                    IS = _LINK_ICON_SIZE
                    # Right edge of title text + 3px gap.
                    # For Line style, titles are always centered over the event
                    # position, so we always compute _title_lx from the center.
                    _next_icon_x = _title_lx + _tw + 3

                    # ── +/− linked-category toggle icon ───────────────────
                    if linked_id:
                        linked_node = self.db.cat_by_id.get(linked_id)
                        if linked_node:
                            cat_hidden = bool(linked_node.get("hidden"))
                            ic   = "#4a7fc1" if cat_hidden else "#4a5568"
                            sign = "+" if cat_hidden else "\u2212"
                            _ix1 = _next_icon_x
                            _iy1 = _icon_y - IS // 2
                            _ix2 = _ix1 + IS
                            _iy2 = _iy1 + IS
                            icon_id = c.create_rectangle(
                                _ix1, _iy1, _ix2, _iy2,
                                fill=ic, outline=ic, width=1,
                                tags=("link_icon",)
                            )
                            c.create_text(
                                (_ix1 + _ix2) / 2, (_iy1 + _iy2) / 2,
                                text=sign, fill="white",
                                font=("Arial", 9, "bold"),
                                tags=("link_icon",)
                            )
                            self._link_icon_map[icon_id] = e
                            _next_icon_x = _ix2 + 3

                    # ── → linked-timeline navigation icon ─────────────────
                    if linked_tlid:
                        _ix1 = _next_icon_x
                        _iy1 = _icon_y - IS // 2
                        _ix2 = _ix1 + IS
                        _iy2 = _iy1 + IS
                        tl_icon_id = c.create_rectangle(
                            _ix1, _iy1, _ix2, _iy2,
                            fill="#2e7d32", outline="#2e7d32", width=1,
                            tags=("timeline_icon",)
                        )
                        c.create_text(
                            (_ix1 + _ix2) / 2, (_iy1 + _iy2) / 2,
                            text="\u2192", fill="white",
                            font=("Arial", 9, "bold"),
                            tags=("timeline_icon",)
                        )
                        self._timeline_icon_map[tl_icon_id] = e

    # ── ruler helpers ─────────────────────────────────────────────────────────

    def _tick_interval(self, span, target_ticks=None):
        if span <= 0:
            return 1.0
        if target_ticks is None:
            # Derive target from visible canvas width: aim for one tick per ~80px
            try:
                px = self.ruler_canvas.winfo_width()
            except Exception:
                px = 800
            target_ticks = max(5, px // 80)
        # Candidates from finest to coarsest: months, years, then geological scales.
        # Using exact fractions avoids floating-point drift in tick loops.
        candidates = [
            1/12, 2/12, 3/12, 6/12,                               # 1, 2, 3, 6 months
            1, 2, 5, 10, 25, 50, 100, 250, 500, 1_000,            # whole years
            5_000, 10_000, 50_000, 100_000, 500_000,              # thousands of years
            1_000_000, 5_000_000, 10_000_000, 50_000_000,         # millions of years (MYA)
            100_000_000, 250_000_000, 500_000_000,                 # hundreds of millions
            1_000_000_000,                                         # 1 billion years (BYA)
        ]
        for interval in candidates:
            if span / interval <= target_ticks:
                return interval
        # Fallback: compute an interval that yields at most target_ticks ticks
        import math
        raw = span / target_ticks
        magnitude = 10 ** math.floor(math.log10(raw))
        for step in (1, 2, 5):
            if magnitude * step >= raw:
                return magnitude * step
        return magnitude * 10

    def _format_tick(self, val):
        if abs(val) >= 1_000_000_000:
            return f"{-val/1e9:g} BYA"
        if abs(val) >= 1_000_000:
            return f"{-val/1e6:g} MYA"
        interval = getattr(self, "_tick_interval_cache", 1.0)
        if interval < 1.0:
            _MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
                       "Jul","Aug","Sep","Oct","Nov","Dec"]
            year = int(val) if val >= 0 else int(val) - 1
            month_idx = min(11, max(0, round((val - year) * 12)))
            era = " BCE" if year < 0 else ""
            return f"{_MONTHS[month_idx]} {abs(year)}{era}"
        year = int(val)
        if year < 0:
            return f"{-year} BCE"
        return str(year)

    # ── interactions ──────────────────────────────────────────────────────────

    def _scroll_home(self):
        self.canvas.xview_moveto(0.0)
        self.ruler_canvas.xview_moveto(0.0)
        self.canvas.yview_moveto(0.0)
        self.label_canvas.yview_moveto(0.0)

    def _scroll_end(self):
        self.canvas.xview_moveto(1.0)
        self.ruler_canvas.xview_moveto(1.0)
        self.canvas.yview_moveto(1.0)
        self.label_canvas.yview_moveto(1.0)

    def _arrow_scroll_h(self, direction):
        """Scroll the timeline horizontally by ~100 px per arrow keypress."""
        STEP = 100
        total_w = self._canvas_width()
        if total_w <= 0:
            return
        delta = direction * STEP / total_w
        new_x = max(0.0, min(1.0, self.canvas.xview()[0] + delta))
        self.canvas.xview_moveto(new_x)
        self.ruler_canvas.xview_moveto(new_x)

    def _arrow_scroll_v(self, direction):
        """Scroll the timeline vertically by ~50 px per arrow keypress."""
        STEP = 50
        total_h = self.canvas.winfo_height()
        if total_h <= 0:
            return
        delta = direction * STEP / total_h
        new_y = max(0.0, min(1.0, self.canvas.yview()[0] + delta))
        self.canvas.yview_moveto(new_y)
        self.label_canvas.yview_moveto(new_y)

    def _drag_start(self, event):
        self.canvas.scan_mark(event.x, event.y)
        self.ruler_canvas.scan_mark(event.x, 0)
        self.label_canvas.scan_mark(0, event.y)

    def _drag_move(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)
        self.ruler_canvas.scan_dragto(event.x, 0, gain=1)
        self.label_canvas.scan_dragto(0, event.y, gain=1)

    def _drag_start_ruler(self, event):
        self.ruler_canvas.scan_mark(event.x, 0)
        self.canvas.scan_mark(event.x, 0)

    def _drag_move_ruler(self, event):
        self.ruler_canvas.scan_dragto(event.x, 0, gain=1)
        self.canvas.scan_dragto(event.x, 0, gain=1)

    def _drag_start_label(self, event):
        self.label_canvas.scan_mark(0, event.y)
        self.canvas.scan_mark(0, event.y)

    def _drag_move_label(self, event):
        self.label_canvas.scan_dragto(0, event.y, gain=1)
        self.canvas.scan_dragto(0, event.y, gain=1)

    def _mouse_wheel(self, event):
        factor = 1.5 if (event.num == 4 or event.delta > 0) else 1 / 1.5
        self._zoom(factor, event)

    def _save_view_state(self):
        self.db.save_timeline_view_state(
            self.db.active_timeline_id, self.px_per_year, self._freeze_scale_var.get()
        )

    def _on_freeze_toggle(self):
        frozen = self._freeze_scale_var.get()
        state  = tk.DISABLED if frozen else tk.NORMAL
        self._btn_fit_all.config(state=state)
        self._scale_entry.config(state=state)

    def _apply_manual_scale(self):
        if self._freeze_scale_var.get():
            return
        try:
            val = float(self._scale_var.get())
            if val > 0:
                px = _display_to_px_per_year(val, self._scale_unit_var.get())
                self.px_per_year = max(min(px, 50000), 1e-15)
                self._draw()
        except ValueError:
            pass
        self.canvas.focus_set()

    def _zoom_at_mouse(self, factor):
        """Zoom centred on the last known mouse x position, like the scroll wheel."""
        if self._last_mouse_x is not None:
            class _Evt:
                pass
            ev = _Evt()
            ev.x = self._last_mouse_x
            self._zoom(factor, ev)
        else:
            self._zoom(factor)


    def _zoom(self, factor, event=None):
        if self._freeze_scale_var.get():
            return
        # Record the date under the cursor before changing the scale
        if event is not None:
            cx = self.canvas.canvasx(event.x)
            date_at_cursor = self._date_from_x(cx)
            screen_x = event.x   # pixel position within the canvas widget

        self.px_per_year = max(min(self.px_per_year * factor, 50000), 0.000001)
        self._draw()

        # Scroll so the date that was under the cursor stays at the same screen x
        if event is not None:
            new_cx = self._x(date_at_cursor)
            new_offset = new_cx - screen_x
            W = self._canvas_width()
            if W > 0:
                fraction = max(0.0, min(1.0, new_offset / W))
                self.canvas.xview_moveto(fraction)
                self.ruler_canvas.xview_moveto(fraction)

    def _scroll_to_visible(self):
        """Scroll horizontally so the earliest visible event is near the left edge."""
        visible_vals = []
        for cat, events in self._all_cat_events.items():
            if cat in self._collapsed_cats or self._ancestor_collapsed(cat):
                continue
            for e in events:
                if e.get("start_value") is not None:
                    visible_vals.append(e["start_value"])
                if e.get("end_value") is not None:
                    visible_vals.append(e["end_value"])
        if visible_vals:
            vis_min = min(visible_vals)
            cx = self._x(vis_min)
            W  = self._canvas_width()
            if W > 0:
                fraction = max(0.0, min(1.0, (cx - self.CANVAS_MARGIN) / W))
                self.canvas.xview_moveto(fraction)
                self.ruler_canvas.xview_moveto(fraction)

    def _fit_all(self):
        if self._freeze_scale_var.get():
            return
        self.win.update_idletasks()
        w = self.canvas.winfo_width() - 2 * self.CANVAS_MARGIN

        visible_vals = []
        for cat, events in self._all_cat_events.items():
            if cat in self._collapsed_cats or self._ancestor_collapsed(cat):
                continue
            for e in events:
                if e.get("start_value") is not None:
                    visible_vals.append(e["start_value"])
                if e.get("end_value") is not None:
                    visible_vals.append(e["end_value"])
        if visible_vals:
            vis_min = min(visible_vals)
            vis_max = max(visible_vals)
            span = (vis_max - vis_min) or 1
        else:
            vis_min = self.min_date
            vis_max = self.max_date
            span = (vis_max - vis_min) or 1

        # Subtract break spans that fall within the visible range so the
        # solved px_per_year accounts for the compression.
        break_span_in_range = 0.0
        n_breaks_in_range   = 0
        for b in getattr(self, '_breaks', []):
            bs = max(b['start'], vis_min)
            be = min(b['end'],   vis_max)
            if be > bs:
                break_span_in_range += be - bs
                n_breaks_in_range   += 1

        effective_span = span - break_span_in_range
        if effective_span > 0:
            self.px_per_year = max(
                (w - n_breaks_in_range * self.BREAK_STUB_PX) / effective_span,
                1e-15)
        else:
            self.px_per_year = max(w / span, 1e-15)
        self._draw()
        self._scroll_to_visible()

    def _save_pdf(self):
        from PIL import ImageGrab

        path = filedialog.asksaveasfilename(
            parent=self.win,
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Save Timeline as PDF",
            initialfile=f"{self._tl_var.get() or 'timeline'}.pdf",
        )
        if not path:
            return

        c  = self.canvas
        lc = self.label_canvas
        rc = self.ruler_canvas
        self.win.attributes("-topmost", True)
        self.win.lift()
        self.win.update_idletasks()

        # Full canvas dimensions from scrollregion
        sr = list(map(float, c.cget("scrollregion").split()))
        full_W, full_H = int(sr[2]), int(sr[3])
        if full_W <= 0 or full_H <= 0:
            messagebox.showwarning("Save PDF", "Nothing to save.", parent=self.win)
            return

        # ── DPI scale factor (Windows high-DPI fix) ────────────────────────
        # winfo_root* returns logical pixels; ImageGrab.grab uses physical pixels.
        # Detect the scale so we can convert logical → physical for the grab bbox,
        # then resize the captured tile back to logical size before pasting.
        try:
            import ctypes
            dpi_scale = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100.0
        except Exception:
            dpi_scale = 1.0

        def _grab(lx, ly, lw, lh):
            """Grab a logical-pixel region and return a logical-pixel-sized image."""
            # Round origin, then use exact physical size to avoid per-edge rounding drift.
            px0  = round(lx * dpi_scale)
            py0  = round(ly * dpi_scale)
            px_w = round(lw * dpi_scale)
            px_h = round(lh * dpi_scale)
            img  = ImageGrab.grab(bbox=(px0, py0, px0 + px_w, py0 + px_h))
            if dpi_scale != 1.0:
                img = img.resize((lw, lh), Image.LANCZOS)
            return img

        # Viewport sizes and screen positions (all in logical pixels)
        vp_w  = c.winfo_width();   vp_h  = c.winfo_height()
        sx    = c.winfo_rootx();   sy    = c.winfo_rooty()
        show_labels = self._cat_header_style in ("Left", "Both")
        lc_w  = lc.winfo_width() if show_labels else 0
        lsx   = lc.winfo_rootx();  lsy   = lc.winfo_rooty()
        rcx   = rc.winfo_rootx();  rcy   = rc.winfo_rooty()
        rc_h  = self.RULER_HEIGHT

        # Save current scroll positions
        saved_x = c.xview()[0]
        saved_y = c.yview()[0]

        try:
            ruler_bg = (44, 62, 80)   # matches #2c3e50
            full_img = Image.new("RGB", (lc_w + full_W, rc_h + full_H), "white")

            # ── Corner square above label column ──────────────────────────
            full_img.paste(Image.new("RGB", (lc_w, rc_h), ruler_bg), (0, 0))

            def _scroll_grab_row(y_paste, grab_x, grab_y, grab_w, grab_h,
                                 img_x_offset, full_row_h):
                """Tile one horizontal strip seamlessly into full_img."""
                x_pos = 0
                while x_pos < full_W:
                    frac = x_pos / full_W
                    c.xview_moveto(frac)
                    rc.xview_moveto(frac)
                    self.win.update()
                    # Use int() (floor) not round() so actual_x never exceeds x_pos.
                    actual_x  = int(c.canvasx(0))
                    tile_skip = max(0, x_pos - actual_x)
                    crop_w    = min(grab_w - tile_skip, full_W - x_pos)
                    if crop_w <= 0:
                        break
                    tile = _grab(grab_x, grab_y, grab_w, grab_h)
                    full_img.paste(
                        tile.crop((tile_skip, 0, tile_skip + crop_w, full_row_h)),
                        (img_x_offset + x_pos, y_paste)
                    )
                    x_pos += crop_w

            # ── Ruler strip (top row) ─────────────────────────────────────
            _scroll_grab_row(0, rcx, rcy, vp_w, rc_h, lc_w, rc_h)

            # ── Body rows ─────────────────────────────────────────────────
            y_pos = 0
            while y_pos < full_H:
                frac_y = y_pos / full_H
                c.yview_moveto(frac_y)
                lc.yview_moveto(frac_y)
                self.win.update()
                actual_y  = int(c.canvasy(0))
                row_skip  = max(0, y_pos - actual_y)
                crop_h    = min(vp_h - row_skip, full_H - y_pos)
                if crop_h <= 0:
                    break

                # Label column (only when visible)
                if show_labels:
                    lbl_tile = _grab(lsx, lsy, lc_w, vp_h)
                    full_img.paste(lbl_tile.crop((0, row_skip, lc_w, row_skip + crop_h)),
                                   (0, rc_h + y_pos))

                # Event canvas — horizontal tiles for this row
                x_pos = 0
                while x_pos < full_W:
                    frac_x = x_pos / full_W
                    c.xview_moveto(frac_x)
                    self.win.update()
                    actual_x  = int(c.canvasx(0))
                    tile_skip = max(0, x_pos - actual_x)
                    crop_w    = min(vp_w - tile_skip, full_W - x_pos)
                    if crop_w <= 0:
                        break
                    tile = _grab(sx, sy, vp_w, vp_h)
                    full_img.paste(
                        tile.crop((tile_skip, row_skip,
                                   tile_skip + crop_w, row_skip + crop_h)),
                        (lc_w + x_pos, rc_h + y_pos)
                    )
                    x_pos += crop_w

                y_pos += crop_h

            full_img.save(path, "PDF", resolution=96.0)
            os.startfile(path)

        except Exception as ex:
            messagebox.showerror("Save PDF Error", str(ex), parent=self.win)
        finally:
            self.win.attributes("-topmost", False)
            c.xview_moveto(saved_x)
            c.yview_moveto(saved_y)
            lc.yview_moveto(saved_y)
            self.win.update_idletasks()

    def _publish_to_timelinehub(self):
        try:
            from publish_dialog import PublishDialog, check_supabase_installed
        except ImportError:
            messagebox.showerror(
                "Missing file",
                "publish_dialog.py not found next to timeline_db.py.",
                parent=self.win,
            )
            return

        if not check_supabase_installed():
            messagebox.showerror(
                "Missing dependency",
                "The 'supabase' package is required.\n\n"
                "Run this in your terminal and restart the app:\n\n"
                "    pip install supabase",
                parent=self.win,
            )
            return

        title = self._tl_var.get() or "Timeline"

        def generate_pdf(path: str):
            """Render the full timeline canvas and save it as a PDF at *path*."""
            from PIL import ImageGrab, Image as _Image

            c  = self.canvas
            lc = self.label_canvas
            rc = self.ruler_canvas
            self.win.attributes("-topmost", True)
            self.win.lift()
            self.win.update_idletasks()

            sr = list(map(float, c.cget("scrollregion").split()))
            full_W, full_H = int(sr[2]), int(sr[3])
            if full_W <= 0 or full_H <= 0:
                raise RuntimeError("Nothing to export — the canvas is empty.")

            try:
                import ctypes
                dpi_scale = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100.0
            except Exception:
                dpi_scale = 1.0

            def _grab(lx, ly, lw, lh):
                px0  = round(lx * dpi_scale)
                py0  = round(ly * dpi_scale)
                px_w = round(lw * dpi_scale)
                px_h = round(lh * dpi_scale)
                img  = ImageGrab.grab(bbox=(px0, py0, px0 + px_w, py0 + px_h))
                if dpi_scale != 1.0:
                    img = img.resize((lw, lh), _Image.LANCZOS)
                return img

            vp_w  = c.winfo_width();   vp_h  = c.winfo_height()
            sx    = c.winfo_rootx();   sy    = c.winfo_rooty()
            show_labels = self._cat_header_style in ("Left", "Both")
            lc_w  = lc.winfo_width() if show_labels else 0
            lsx   = lc.winfo_rootx();  lsy   = lc.winfo_rooty()
            rcx   = rc.winfo_rootx();  rcy   = rc.winfo_rooty()
            rc_h  = self.RULER_HEIGHT
            saved_x = c.xview()[0]
            saved_y = c.yview()[0]

            try:
                ruler_bg = (44, 62, 80)
                full_img = _Image.new("RGB", (lc_w + full_W, rc_h + full_H), "white")
                full_img.paste(_Image.new("RGB", (lc_w, rc_h), ruler_bg), (0, 0))

                def _scroll_grab_row(y_paste, grab_x, grab_y, grab_w, grab_h,
                                     img_x_offset, full_row_h):
                    x_pos = 0
                    while x_pos < full_W:
                        frac = x_pos / full_W
                        c.xview_moveto(frac)
                        rc.xview_moveto(frac)
                        self.win.update()
                        actual_x  = int(c.canvasx(0))
                        tile_skip = max(0, x_pos - actual_x)
                        crop_w    = min(grab_w - tile_skip, full_W - x_pos)
                        if crop_w <= 0:
                            break
                        tile = _grab(grab_x, grab_y, grab_w, grab_h)
                        full_img.paste(
                            tile.crop((tile_skip, 0, tile_skip + crop_w, full_row_h)),
                            (img_x_offset + x_pos, y_paste)
                        )
                        x_pos += crop_w

                _scroll_grab_row(0, rcx, rcy, vp_w, rc_h, lc_w, rc_h)

                y_pos = 0
                while y_pos < full_H:
                    frac_y = y_pos / full_H
                    c.yview_moveto(frac_y)
                    lc.yview_moveto(frac_y)
                    self.win.update()
                    actual_y  = int(c.canvasy(0))
                    row_skip  = max(0, y_pos - actual_y)
                    crop_h    = min(vp_h - row_skip, full_H - y_pos)
                    if crop_h <= 0:
                        break
                    if show_labels:
                        lbl_tile = _grab(lsx, lsy, lc_w, vp_h)
                        full_img.paste(lbl_tile.crop((0, row_skip, lc_w, row_skip + crop_h)),
                                       (0, rc_h + y_pos))
                    x_pos = 0
                    while x_pos < full_W:
                        frac_x = x_pos / full_W
                        c.xview_moveto(frac_x)
                        self.win.update()
                        actual_x  = int(c.canvasx(0))
                        tile_skip = max(0, x_pos - actual_x)
                        crop_w    = min(vp_w - tile_skip, full_W - x_pos)
                        if crop_w <= 0:
                            break
                        tile = _grab(sx, sy, vp_w, vp_h)
                        full_img.paste(
                            tile.crop((tile_skip, row_skip,
                                       tile_skip + crop_w, row_skip + crop_h)),
                            (lc_w + x_pos, rc_h + y_pos)
                        )
                        x_pos += crop_w
                    y_pos += crop_h

                full_img.save(path, "PDF", resolution=96.0)
            finally:
                self.win.attributes("-topmost", False)
                c.xview_moveto(saved_x)
                c.yview_moveto(saved_y)
                lc.yview_moveto(saved_y)
                self.win.update_idletasks()

        PublishDialog(self.win, title, generate_pdf)

    def _on_link_icon_click(self, event):
        """Toggle the hidden state of the linked category for the clicked icon."""
        c = self.canvas
        cx = c.canvasx(event.x)
        cy = c.canvasy(event.y)
        # find_closest may return a text label on top of the rectangle; check both
        for item in c.find_overlapping(cx - 2, cy - 2, cx + 2, cy + 2):
            if "link_icon" not in c.gettags(item):
                continue
            evt = self._link_icon_map.get(item)
            if evt is None:
                continue
            linked_id = evt.get("linked_categoryid")
            if not linked_id:
                continue
            node = self.db.cat_by_id.get(linked_id)
            if not node:
                continue
            new_hidden = 0 if node.get("hidden") else 1
            with sqlite3.connect(self.db.db_file) as conn:
                conn.execute("UPDATE Category SET hidden=? WHERE CategoryID=?",
                             (new_hidden, linked_id))
            node["hidden"] = new_hidden
            self._reload(self.db.active_timeline_id)
            return

    def _on_timeline_icon_click(self, event):
        """Switch to the timeline linked on the clicked event."""
        c = self.canvas
        cx = c.canvasx(event.x)
        cy = c.canvasy(event.y)
        for item in c.find_overlapping(cx - 2, cy - 2, cx + 2, cy + 2):
            if "timeline_icon" not in c.gettags(item):
                continue
            evt = self._timeline_icon_map.get(item)
            if evt is None:
                continue
            linked_tlid = evt.get("linked_timelineid")
            if not linked_tlid:
                continue
            # Resolve target timeline title for the prompt
            target_title = next(
                (t for tid, t in self.db.load_timelines() if tid == linked_tlid),
                f"Timeline {linked_tlid}"
            )
            current_title = self._tl_var.get()

            # Skip prompt if user previously opted out
            if not self.db.load_config().get("skip_timeline_nav_confirm"):
                confirmed = self._confirm_timeline_nav(current_title, target_title)
                if not confirmed:
                    return

            self._reload(linked_tlid)
            self._sync_tl_combo()
            return

    def _confirm_timeline_nav(self, current_title, target_title):
        """Show a Yes/No dialog with a 'Don't ask again' checkbox.
        Returns True if user confirmed, False if cancelled."""
        result = [False]

        dlg = tk.Toplevel(self.win)
        dlg.title("Open Linked Timeline")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self.win)

        tk.Label(dlg,
                 text=f"Close \"{current_title}\" and open \"{target_title}\"?",
                 padx=16, pady=12).pack()

        dont_ask_var = tk.IntVar(value=0)
        tk.Checkbutton(dlg, text="Don't ask me again",
                       variable=dont_ask_var).pack(padx=16, pady=(0, 8))

        btn_row = tk.Frame(dlg)
        btn_row.pack(padx=16, pady=(0, 12))

        def _yes():
            if dont_ask_var.get():
                self.db.save_config({"skip_timeline_nav_confirm": True})
            result[0] = True
            dlg.destroy()

        def _no():
            dlg.destroy()

        tk.Button(btn_row, text="Yes", width=8, command=_yes).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="No",  width=8, command=_no).pack(side=tk.LEFT, padx=4)

        # Centre over parent
        dlg.update_idletasks()
        px = self.win.winfo_rootx() + self.win.winfo_width()  // 2 - dlg.winfo_reqwidth()  // 2
        py = self.win.winfo_rooty() + self.win.winfo_height() // 2 - dlg.winfo_reqheight() // 2
        dlg.wm_geometry(f"+{px}+{py}")

        dlg.wait_window()
        return result[0]

    def _on_close(self):
        self.db.save_config({"last_timeline_id": self.db.active_timeline_id})
        self.win.destroy()

    def _open_help(self):
        manual = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "HistoryTimeline_UserManual.pdf")
        if not os.path.exists(manual):
            messagebox.showinfo(
                "Help",
                "User manual not found.\n\n"
                "Run generate_manual.py to create HistoryTimeline_UserManual.pdf.",
                parent=self.win,
            )
            return
        os.startfile(manual)

    # ── icon tooltip (simple one-liner for +/− icon) ─────────────────────────

    def _icon_tooltip_show(self, text, screen_x, screen_y):
        """Show a minimal text tooltip near the cursor for icon hover hints."""
        if getattr(self, "_icon_tip_win", None):
            if getattr(self, "_icon_tip_text", None) == text:
                return   # already showing the same text
            self._icon_tooltip_hide()
        self._icon_tip_text = text
        tip = tk.Toplevel(self.win)
        tip.wm_overrideredirect(True)
        tip.attributes("-topmost", True)
        tk.Label(tip, text=text, bg="#ffffe0", fg="#222222",
                 font=("Arial", 8), relief=tk.SOLID, bd=1,
                 padx=4, pady=2).pack()
        tip.update_idletasks()
        tw = tip.winfo_reqwidth()
        th = tip.winfo_reqheight()
        tx = screen_x + 14
        ty = screen_y + 14
        if tx + tw > tip.winfo_screenwidth():
            tx = screen_x - tw - 6
        if ty + th > tip.winfo_screenheight():
            ty = screen_y - th - 6
        tip.wm_geometry(f"+{tx}+{ty}")
        self._icon_tip_win = tip

    def _icon_tooltip_hide(self):
        if getattr(self, "_icon_tip_win", None):
            self._icon_tip_win.destroy()
            self._icon_tip_win  = None
            self._icon_tip_text = None

    # ── tooltip ───────────────────────────────────────────────────────────────

    def _tooltip_hide(self):
        if hasattr(self, "_tip_after") and self._tip_after:
            self.canvas.after_cancel(self._tip_after)
            self._tip_after = None
        if hasattr(self, "_tip_win") and self._tip_win:
            self._tip_win.destroy()
            self._tip_win = None
        self._tip_event = None

    def _tooltip_hide_soon(self):
        """Schedule hide with a short delay — cancelled if mouse enters tooltip."""
        if hasattr(self, "_tip_after") and self._tip_after:
            self.canvas.after_cancel(self._tip_after)
        self._tip_after = self.canvas.after(300, self._tooltip_hide)

    def _tooltip_cancel_hide(self):
        """Mouse entered the tooltip window — cancel any pending hide."""
        if hasattr(self, "_tip_after") and self._tip_after:
            self.canvas.after_cancel(self._tip_after)
            self._tip_after = None

    def _tooltip_show(self, e, screen_x, screen_y):
        self._tooltip_hide()
        self._tip_event = e.get("id")

        tip = tk.Toplevel(self.win)
        tip.wm_overrideredirect(True)   # no title bar
        tip.attributes("-topmost", True)
        tip.bind("<Enter>", lambda *_: self._tooltip_cancel_hide())
        tip.bind("<Leave>", lambda *_: self._tooltip_hide_soon())

        # Build content
        sd = e.get("start_display") or ""
        ed = e.get("end_display") or ""
        date_str = sd if (not ed or ed == sd) else f"{sd} – {ed}"
        cat       = e.get("category") or ""
        desc      = (e.get("desc") or "").strip()
        url       = (e.get("url") or "").strip()
        img_blob  = e.get("image")

        frame = tk.Frame(tip, bg="#fffde7", bd=1, relief=tk.SOLID)
        frame.pack(fill=tk.BOTH, expand=True)

        # ── Outer row: image on left, text on right ──────────────────────────
        body = tk.Frame(frame, bg="#fffde7")
        body.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

        # Image column
        if img_blob:
            try:
                pil_img = Image.open(io.BytesIO(img_blob))
                pil_img.thumbnail((120, 120), Image.LANCZOS)
                tk_img = ImageTk.PhotoImage(pil_img)
                img_lbl = tk.Label(body, image=tk_img, bg="#fffde7")
                img_lbl.image = tk_img   # keep reference alive
                img_lbl.pack(side=tk.LEFT, anchor=tk.N, padx=(0, 8))
                # Vertical divider
                tk.Frame(body, bg="#ddddbb", width=1).pack(
                    side=tk.LEFT, fill=tk.Y, padx=(0, 8))
            except Exception:
                pass   # silently skip if image can't be decoded

        # Text column
        text_col = tk.Frame(body, bg="#fffde7")
        text_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, anchor=tk.N)

        tk.Label(text_col, text=e["title"], bg="#fffde7",
                 font=("Arial", 10, "bold"), anchor=tk.W,
                 wraplength=280, justify=tk.LEFT).pack(anchor=tk.W)

        if date_str:
            tk.Label(text_col, text=date_str, bg="#fffde7",
                     font=("Arial", 8), fg="#555555", anchor=tk.W).pack(anchor=tk.W)


        if desc:
            tk.Frame(text_col, bg="#ddddbb", height=1).pack(fill=tk.X, pady=4)
            tk.Label(text_col, text=desc, bg="#fffde7",
                     font=("Arial", 8), fg="#333333", anchor=tk.W,
                     justify=tk.LEFT, wraplength=280).pack(anchor=tk.W)

        if url:
            link = tk.Label(text_col, text="Open Link ↗", bg="#fffde7",
                            font=("Arial", 8, "underline"), fg="#1a6fc4",
                            cursor="hand2", anchor=tk.W)
            link.pack(anchor=tk.W, pady=(4, 0))
            link.bind("<Button-1>", lambda *_, u=url: webbrowser.open(u))
        else:
            tk.Frame(text_col, height=2, bg="#fffde7").pack()

        # Position near cursor but keep on screen
        tip.update_idletasks()
        tw = tip.winfo_reqwidth()
        th = tip.winfo_reqheight()
        sw = tip.winfo_screenwidth()
        sh = tip.winfo_screenheight()
        tx = screen_x + 16
        ty = screen_y + 16
        if tx + tw > sw:
            tx = screen_x - tw - 8
        if ty + th > sh:
            ty = screen_y - th - 8
        tip.wm_geometry(f"+{tx}+{ty}")
        self._tip_win = tip

    def _canvas_event_at(self, x, y):
        """Return the event dict under canvas coordinates (x, y), or None."""
        cx = self.canvas.canvasx(x)
        cy = self.canvas.canvasy(y)
        for item in self.canvas.find_overlapping(cx - 2, cy - 2, cx + 2, cy + 2):
            if item in self._event_rects:
                return self._event_rects[item]
        return None

    def _on_hover(self, event):
        self._last_mouse_x = event.x
        c  = self.canvas
        cx = c.canvasx(event.x)
        cy = c.canvasy(event.y)

        # Check for clickable icons first
        link_tip_text = None
        for item in c.find_overlapping(cx - 2, cy - 2, cx + 2, cy + 2):
            tags = c.gettags(item)
            if "link_icon" in tags:
                evt = self._link_icon_map.get(item)
                if evt:
                    node = self.db.cat_by_id.get(evt.get("linked_categoryid"))
                    if node:
                        action = "Close" if not node.get("hidden") else "Open"
                        link_tip_text = f"{action} {node['title']}"
                break
            if "timeline_icon" in tags:
                evt = self._timeline_icon_map.get(item)
                if evt:
                    linked_tlid = evt.get("linked_timelineid")
                    tl_title = next(
                        (t for tid, t in self.db.load_timelines() if tid == linked_tlid),
                        None
                    )
                    link_tip_text = f"Open {tl_title}" if tl_title else ""
                else:
                    link_tip_text = ""
                break

        if link_tip_text is not None:
            c.config(cursor="hand2")
            self._tooltip_hide_soon()
            if link_tip_text:
                self._icon_tooltip_show(link_tip_text, event.x_root, event.y_root)
            self._update_status()
            return

        self._icon_tooltip_hide()

        e = self._canvas_event_at(event.x, event.y)

        if e:
            self.canvas.config(cursor="hand2")
            # Update status bar
            sd = e.get("start_display", "")
            ed = e.get("end_display", "")
            date_str = sd if (not ed or ed == sd) else f"{sd} – {ed}"
            self.status_var.set(
                f"{e['title']}   {date_str}   [{e.get('category', '')}]")
            # Show tooltip only if hovering a different event
            if self._tip_event != e.get("id"):
                self._tooltip_show(e, event.x_root, event.y_root)
            return

        # Nothing under cursor
        self.canvas.config(cursor="")
        self._tooltip_hide_soon()
        self._update_status()

    def _on_canvas_right_click(self, event):
        e = self._canvas_event_at(event.x, event.y)
        if e:
            self._open_edit_event_dialog(e)
            return

        # Check cursor is inside a lane row (not the gap strip between rows)
        cy = self.canvas.canvasy(event.y)
        row_size = self.LANE_HEIGHT + self.LANE_PAD
        row_index = int(cy / row_size)
        within_lane = (cy % row_size) < self.LANE_HEIGHT

        if not within_lane or row_index < 0 or row_index >= len(self.rows):
            return

        # Find which category owns this row
        cat_name = None
        for cat, (first, last) in self.cat_row_spans.items():
            if first <= row_index <= last:
                cat_name = cat
                break
        if cat_name is None:
            return

        cat_node = next((n for n in self.db.cat_nodes if n["title"] == cat_name), None)
        cat_id = cat_node["id"] if cat_node else None

        # Convert cursor x to a date value and pick appropriate unit
        cx = self.canvas.canvasx(event.x)
        date_val = self._date_from_x(cx)
        if date_val <= -1_000_000_000:
            unit = "BYA"
        elif date_val <= -1_000_000:
            unit = "MYA"
        elif date_val < 0:
            unit = "BCE"
        else:
            unit = "CE"

        blank = {
            "id": 0, "title": "", "categoryid": cat_id,
            "start_unit": unit, "start_value": date_val,
            "start_month": None, "start_day": None,
            "end_unit": "CE", "end_value": None,
            "end_month": None, "end_day": None,
            "desc": "", "url": "", "standalone": 1,
            "image": None, "image_name": None, "image_type": None,
        }
        self._open_edit_event_dialog(blank)

    def _on_ruler_motion(self, event):
        self._last_mouse_x = event.x
        cx = self.ruler_canvas.canvasx(event.x)
        for b in getattr(self, '_breaks', []):
            x1 = self._x(b['start'])
            x2 = x1 + self.BREAK_STUB_PX
            if x1 <= cx <= x2:
                tip = f"Break: {self._format_tick(b['start'])} → {self._format_tick(b['end'])}"
                self._icon_tooltip_show(tip, event.x_root, event.y_root)
                return
        self._icon_tooltip_hide()

    def _on_ruler_right_click(self, event):
        cx = self.ruler_canvas.canvasx(event.x)
        date_at_click = self._date_from_x(cx)

        near_break = None
        for b in self._breaks:
            x1 = self._x(b['start'])
            x2 = x1 + self.BREAK_STUB_PX
            if x1 - 8 <= cx <= x2 + 8:
                near_break = b
                break

        menu = tk.Menu(self.win, tearoff=0)
        if near_break:
            menu.add_command(label="Edit Break...",
                             command=lambda b=near_break: self._open_edit_break_dialog(b))
            menu.add_command(label="Remove Break",
                             command=lambda b=near_break: self._delete_break(b['id']))
        else:
            menu.add_command(label="Add Break Here...",
                             command=lambda: self._open_add_break_dialog(date_at_click))
        menu.tk_popup(event.x_root, event.y_root)

    def _delete_break(self, break_id):
        self.db.delete_timeline_break(break_id)
        self._breaks = self.db.load_timeline_breaks(self.db.active_timeline_id)
        self._draw()

    def _open_edit_break_dialog(self, b):
        dlg = tk.Toplevel(self.win)
        dlg.title("Edit Timeline Break")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self.win)

        def _val_to_parts(val):
            if val <= -1_000_000_000:
                return f"{abs(val)/1_000_000_000:g}", "BYA"
            elif val <= -1_000_000:
                return f"{abs(val)/1_000_000:g}", "MYA"
            elif val < 0:
                return str(int(abs(val))), "BCE"
            else:
                return str(int(val)), "CE"

        def _date_row(grid_row, label_text, init_val):
            tk.Label(dlg, text=label_text).grid(
                row=grid_row, column=0, sticky="e", padx=(12, 4), pady=3)
            frame = tk.Frame(dlg)
            frame.grid(row=grid_row, column=1, sticky="w", pady=3, padx=(0, 12))
            num_s, unit_s = _val_to_parts(init_val)
            num_var  = tk.StringVar(value=num_s)
            unit_var = tk.StringVar(value=unit_s)
            tk.Entry(frame, textvariable=num_var, width=10).pack(side=tk.LEFT, padx=(0, 2))
            ttk.Combobox(frame, textvariable=unit_var,
                         values=["CE", "BCE", "MYA", "BYA"],
                         width=5, state="readonly").pack(side=tk.LEFT)
            return num_var, unit_var

        start_num, start_unit = _date_row(0, "Break Start:", b['start'])
        end_num,   end_unit   = _date_row(1, "Break End:",   b['end'])

        def _ok():
            sv = _date_value(start_num.get().strip(), start_unit.get())
            ev = _date_value(end_num.get().strip(),   end_unit.get())
            if sv is None or ev is None:
                messagebox.showerror("Invalid Date", "Enter valid dates.", parent=dlg)
                return
            if sv >= ev:
                messagebox.showerror("Invalid Range",
                                     "Break Start must be before Break End.", parent=dlg)
                return
            dlg.destroy()
            self.db.update_timeline_break(b['id'], sv, ev)
            self._breaks = self.db.load_timeline_breaks(self.db.active_timeline_id)
            self._draw()

        btn_f = tk.Frame(dlg)
        btn_f.grid(row=2, column=0, columnspan=2, pady=(8, 10))
        tk.Button(btn_f, text="OK",     width=8, command=_ok).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_f, text="Cancel", width=8, command=dlg.destroy).pack(side=tk.LEFT, padx=4)
        dlg.columnconfigure(1, weight=1)
        self.win.update_idletasks()
        wx = self.win.winfo_rootx() + self.win.winfo_width()  // 2 - 210
        wy = self.win.winfo_rooty() + self.win.winfo_height() // 2 - 80
        dlg.geometry(f"+{max(0,wx)}+{max(0,wy)}")

    def _open_add_break_dialog(self, date_hint):
        dlg = tk.Toplevel(self.win)
        dlg.title("Add Timeline Break")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self.win)

        if date_hint <= -1_000_000_000:
            hint_unit, hint_num = "BYA", f"{abs(date_hint)/1_000_000_000:g}"
        elif date_hint <= -1_000_000:
            hint_unit, hint_num = "MYA", f"{abs(date_hint)/1_000_000:g}"
        elif date_hint < 0:
            hint_unit, hint_num = "BCE", str(int(abs(date_hint)))
        else:
            hint_unit, hint_num = "CE", str(int(date_hint))

        tk.Label(dlg,
                 text="A break compresses the gap between two dates on the ruler.",
                 font=("Arial", 9), fg="#555555").grid(
            row=0, column=0, columnspan=2, padx=12, pady=(10, 6), sticky="w")

        def _date_row(grid_row, label_text):
            tk.Label(dlg, text=label_text).grid(
                row=grid_row, column=0, sticky="e", padx=(12, 4), pady=3)
            frame = tk.Frame(dlg)
            frame.grid(row=grid_row, column=1, sticky="w", pady=3, padx=(0, 12))
            num_var  = tk.StringVar(value=hint_num)
            unit_var = tk.StringVar(value=hint_unit)
            tk.Entry(frame, textvariable=num_var, width=10).pack(side=tk.LEFT, padx=(0, 2))
            ttk.Combobox(frame, textvariable=unit_var,
                         values=["CE", "BCE", "MYA", "BYA"],
                         width=5, state="readonly").pack(side=tk.LEFT)
            return num_var, unit_var

        start_num, start_unit = _date_row(1, "Break Start:")
        end_num,   end_unit   = _date_row(2, "Break End:")

        def _ok():
            sv = _date_value(start_num.get().strip(), start_unit.get())
            ev = _date_value(end_num.get().strip(),   end_unit.get())
            if sv is None or ev is None:
                messagebox.showerror("Invalid Date", "Enter valid dates.", parent=dlg)
                return
            if sv >= ev:
                messagebox.showerror("Invalid Range",
                                     "Break Start must be before Break End.", parent=dlg)
                return
            dlg.destroy()
            self.db.add_timeline_break(self.db.active_timeline_id, sv, ev)
            self._breaks = self.db.load_timeline_breaks(self.db.active_timeline_id)
            self._draw()

        btn_f = tk.Frame(dlg)
        btn_f.grid(row=3, column=0, columnspan=2, pady=(6, 10))
        tk.Button(btn_f, text="OK",     width=8, command=_ok).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_f, text="Cancel", width=8, command=dlg.destroy).pack(side=tk.LEFT, padx=4)
        dlg.columnconfigure(1, weight=1)
        self.win.update_idletasks()
        wx = self.win.winfo_rootx() + self.win.winfo_width()  // 2 - 210
        wy = self.win.winfo_rooty() + self.win.winfo_height() // 2 - 90
        dlg.geometry(f"+{max(0,wx)}+{max(0,wy)}")

    def _open_manage_breaks_dialog(self):
        dlg = tk.Toplevel(self.win)
        dlg.title("Manage Timeline Breaks")
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.transient(self.win)

        tk.Label(dlg, text="Timeline Breaks:",
                 font=("Arial", 10, "bold")).pack(padx=12, pady=(10, 4), anchor="w")

        frame = tk.Frame(dlg)
        frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 4))

        tree = ttk.Treeview(frame, columns=("start", "end"), show="headings", height=8)
        tree.heading("start", text="Break Start")
        tree.heading("end",   text="Break End")
        tree.column("start", width=160)
        tree.column("end",   width=160)
        sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        def _refresh():
            tree.delete(*tree.get_children())
            for b in self._breaks:
                tree.insert("", "end", iid=str(b['id']),
                            values=(self._format_tick(b['start']),
                                    self._format_tick(b['end'])))
        _refresh()

        def _delete_selected():
            for iid in tree.selection():
                self.db.delete_timeline_break(int(iid))
            self._breaks = self.db.load_timeline_breaks(self.db.active_timeline_id)
            _refresh()
            self._draw()

        def _add():
            dlg.destroy()
            self._open_add_break_dialog((self.min_date + self.max_date) / 2)

        btn_f = tk.Frame(dlg)
        btn_f.pack(padx=12, pady=(0, 10), anchor="w")
        tk.Button(btn_f, text="Delete Selected", command=_delete_selected).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(btn_f, text="Add Break...",    command=_add).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(btn_f, text="Close",           command=dlg.destroy).pack(side=tk.LEFT)

        self.win.update_idletasks()
        wx = self.win.winfo_rootx() + self.win.winfo_width()  // 2 - 200
        wy = self.win.winfo_rooty() + self.win.winfo_height() // 2 - 140
        dlg.geometry(f"440x280+{max(0,wx)}+{max(0,wy)}")

    def _update_status(self):
        W = self._canvas_width()
        H = self._total_height()
        self.status_var.set(f"Zoom: {self.px_per_year:.2f} px/yr   "
                            f"Range: {self._format_tick(self.min_date)} → "
                            f"{self._format_tick(self.max_date)}   "
                            f"Canvas: {int(round(W))} \u00d7 {int(round(H))} px")



class EditTimelineDialog:
    """Modal dialog for managing timelines — list, add, rename, delete."""

    def __init__(self, root, db, view=None, on_change=None, on_import=None):
        self.db        = db
        self.view      = view
        self.on_change = on_change
        self.on_import = on_import
        self._selected_id  = None
        self._timeline_ids = []

        self.win = tk.Toplevel(root)
        self.win.title("Manage Timelines")
        self.win.resizable(True, True)
        self.win.grab_set()
        self.win.protocol("WM_DELETE_WINDOW", self.win.destroy)
        self.win.columnconfigure(0, weight=1)
        self.win.rowconfigure(1, weight=1)

        tk.Label(self.win, text="Timelines", font=("Arial", 11, "bold")).grid(
            row=0, column=0, pady=(12, 4))

        # ── Listbox ───────────────────────────────────────────────────────────
        list_frame = tk.Frame(self.win, padx=16)
        list_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 4))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        vsb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        vsb.grid(row=0, column=1, sticky="ns")
        self.listbox = tk.Listbox(list_frame, yscrollcommand=vsb.set, height=8,
                                  selectmode=tk.SINGLE, font=("Arial", 10),
                                  exportselection=False)
        vsb.config(command=self.listbox.yview)
        self.listbox.grid(row=0, column=0, sticky="nsew")

        # ── ID + Name ─────────────────────────────────────────────────────────
        entry_frame = tk.Frame(self.win, padx=16)
        entry_frame.grid(row=2, column=0, sticky="ew", pady=(0, 2))
        tk.Label(entry_frame, text="ID:").pack(side=tk.LEFT)
        self.id_var = tk.StringVar(value="")
        tk.Entry(entry_frame, textvariable=self.id_var, width=6,
                 state="readonly").pack(side=tk.LEFT, padx=(2, 10))
        tk.Label(entry_frame, text="Name:").pack(side=tk.LEFT)
        self.name_var = tk.StringVar()
        tk.Entry(entry_frame, textvariable=self.name_var, width=28).pack(side=tk.LEFT, padx=6)

        # ── Scale settings ────────────────────────────────────────────────────
        scale_frame = tk.Frame(self.win, padx=16)
        scale_frame.grid(row=3, column=0, sticky="ew", pady=(0, 4))
        tk.Label(scale_frame, text="Scale:").pack(side=tk.LEFT)
        self.tl_scale_var = tk.StringVar()
        tk.Entry(scale_frame, textvariable=self.tl_scale_var, width=10).pack(side=tk.LEFT, padx=(4, 2))
        self.tl_scale_unit_var = tk.StringVar(value="px/yr")
        tk.Label(scale_frame, textvariable=self.tl_scale_unit_var).pack(side=tk.LEFT, padx=(0, 16))
        self.tl_freeze_var = tk.BooleanVar(value=False)
        tk.Checkbutton(scale_frame, text="Freeze Scale", variable=self.tl_freeze_var).pack(side=tk.LEFT)

        # ── Ruler date range ──────────────────────────────────────────────────
        ruler_frame = tk.Frame(self.win, padx=16)
        ruler_frame.grid(row=4, column=0, sticky="ew", pady=(0, 4))

        self.tl_ruler_min_num_var   = tk.StringVar()
        self.tl_ruler_min_unit_var  = tk.StringVar(value="CE")
        self.tl_ruler_min_month_var = tk.StringVar()
        self.tl_ruler_min_day_var   = tk.StringVar()
        self.tl_ruler_max_num_var   = tk.StringVar()
        self.tl_ruler_max_unit_var  = tk.StringVar(value="CE")
        self.tl_ruler_max_month_var = tk.StringVar()
        self.tl_ruler_max_day_var   = tk.StringVar()

        def _make_ruler_date_row(grid_row, label, num_var, unit_var, month_var, day_var,
                                 allow_present=False):
            tk.Label(ruler_frame, text=label).grid(row=grid_row, column=0, sticky="w", padx=(0, 4), pady=1)
            row_f = tk.Frame(ruler_frame)
            row_f.grid(row=grid_row, column=1, sticky="w", pady=1)
            units = ["CE", "BCE", "MYA", "BYA"] + (["Present"] if allow_present else [])
            num_entry = tk.Entry(row_f, textvariable=num_var, width=10)
            num_entry.pack(side=tk.LEFT, padx=(0, 2))
            unit_cb = ttk.Combobox(row_f, textvariable=unit_var,
                                   values=units, width=7, state="readonly")
            unit_cb.pack(side=tk.LEFT, padx=(0, 4))
            month_cb = ttk.Combobox(row_f, textvariable=month_var,
                                    values=MONTHS, width=10, state="readonly")
            month_cb.pack(side=tk.LEFT, padx=(0, 2))
            day_cb = ttk.Combobox(row_f, textvariable=day_var,
                                  values=[""] + [str(i) for i in range(1, 32)],
                                  width=4, state="readonly")
            day_cb.pack(side=tk.LEFT)

            def _on_unit(*_):
                u = unit_var.get()
                if u == "Present":
                    num_entry.pack_forget()
                    month_cb.pack_forget()
                    day_cb.pack_forget()
                    num_var.set("")
                    month_var.set("")
                    day_var.set("")
                elif u in ("MYA", "BYA"):
                    num_entry.pack(side=tk.LEFT, padx=(0, 2), before=unit_cb)
                    month_cb.pack_forget()
                    day_cb.pack_forget()
                    month_var.set("")
                    day_var.set("")
                else:  # CE, BCE
                    num_entry.pack(side=tk.LEFT, padx=(0, 2), before=unit_cb)
                    month_cb.pack(side=tk.LEFT, padx=(0, 2), after=unit_cb)
                    day_cb.pack(side=tk.LEFT, after=month_cb)
            unit_var.trace_add("write", _on_unit)
            _on_unit()

        _make_ruler_date_row(0, "Ruler Start:",
                             self.tl_ruler_min_num_var, self.tl_ruler_min_unit_var,
                             self.tl_ruler_min_month_var, self.tl_ruler_min_day_var,
                             allow_present=False)
        _make_ruler_date_row(1, "Ruler End:",
                             self.tl_ruler_max_num_var, self.tl_ruler_max_unit_var,
                             self.tl_ruler_max_month_var, self.tl_ruler_max_day_var,
                             allow_present=True)
        tk.Label(ruler_frame, text="(blank = auto)", font=("Arial", 8), fg="#888888").grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(2, 0))

        # ── Category heading style ────────────────────────────────────────────────
        cat_hdr_frame = tk.Frame(self.win, padx=16)
        cat_hdr_frame.grid(row=5, column=0, sticky="ew", pady=(0, 4))
        tk.Label(cat_hdr_frame, text="Category heading:").pack(side=tk.LEFT)
        self.tl_cat_header_var = tk.StringVar(value="Left")
        ttk.Combobox(cat_hdr_frame, textvariable=self.tl_cat_header_var,
                     values=["Left", "Top", "Both", "None"],
                     width=6, state="readonly").pack(side=tk.LEFT, padx=(4, 16))
        tk.Label(cat_hdr_frame, text="Title position:").pack(side=tk.LEFT)
        self.tl_cat_header_title_pos_var = tk.StringVar(value="Center (View)")
        ttk.Combobox(cat_hdr_frame, textvariable=self.tl_cat_header_title_pos_var,
                     values=["Left", "Left (View)", "Center", "Center (View)"],
                     width=14, state="readonly").pack(side=tk.LEFT, padx=(4, 0))

        # ── Default icon styles — commented out; icons hardcoded to Line ────────
        # icon_frame = tk.Frame(self.win, padx=16)
        # icon_frame.grid(row=5, column=0, sticky="ew", pady=(0, 4))
        # tk.Label(icon_frame, text="Point event icon:").pack(side=tk.LEFT)
        # self.tl_icon_short_var = tk.StringVar(value="Diamond")
        # ttk.Combobox(icon_frame, textvariable=self.tl_icon_short_var,
        #              values=["Diamond", "Circle", "Line"],
        #              width=8, state="readonly").pack(side=tk.LEFT, padx=(4, 16))
        # tk.Label(icon_frame, text="Duration event icon:").pack(side=tk.LEFT)
        # self.tl_icon_long_var = tk.StringVar(value="Box")
        # ttk.Combobox(icon_frame, textvariable=self.tl_icon_long_var,
        #              values=["Box", "Line"],
        #              width=6, state="readonly").pack(side=tk.LEFT, padx=(4, 0))

        # ── Canvas background color ───────────────────────────────────────────
        canvas_bg_frame = tk.Frame(self.win, padx=16)
        canvas_bg_frame.grid(row=6, column=0, sticky="ew", pady=(0, 4))
        tk.Label(canvas_bg_frame, text="Canvas background:").pack(side=tk.LEFT)
        self.tl_canvas_bg_var = tk.StringVar(value="")
        self.canvas_bg_swatch = tk.Label(canvas_bg_frame, width=4, relief="sunken", bd=1)
        self.canvas_bg_swatch.pack(side=tk.LEFT, padx=(6, 4))

        def _apply_canvas_bg(color):
            self.tl_canvas_bg_var.set(color)
            if color:
                self.canvas_bg_swatch.config(bg=color)
            else:
                self.canvas_bg_swatch.config(bg=self.win.cget("bg"))

        def _pick_canvas_bg():
            from tkinter import colorchooser as _cc
            current = self.tl_canvas_bg_var.get() or "#f0ede8"
            result = _cc.askcolor(color=current, parent=self.win,
                                  title="Canvas Background Color")
            if result and result[1]:
                _apply_canvas_bg(result[1])

        self._apply_canvas_bg = _apply_canvas_bg
        tk.Button(canvas_bg_frame, text="Choose...",
                  command=_pick_canvas_bg).pack(side=tk.LEFT)
        tk.Button(canvas_bg_frame, text="Clear",
                  command=lambda: _apply_canvas_bg("")).pack(side=tk.LEFT, padx=(4, 0))

        # ── Background image panel ────────────────────────────────────────────
        _BG_THUMB = 96   # fallback size before buttons are rendered

        bg_img_panel = tk.LabelFrame(self.win, text="Background Image", padx=8, pady=6)
        bg_img_panel.grid(row=7, column=0, sticky="ew", padx=16, pady=(0, 4))

        self._tl_bg_img_bytes  = [None]
        self._tl_bg_img_name   = [None]
        self._tl_bg_thumb_ref  = [None]
        self._tl_bg_tmp_path   = [None]
        self._tl_bg_thumb_size = [_BG_THUMB]
        self._tl_bg_size_var   = tk.StringVar(value="")
        self._tl_bg_pos_var    = tk.StringVar(value="Top")

        # Single row: thumbnail col (left) + button col (right)
        bg_thumb_row = tk.Frame(bg_img_panel)
        bg_thumb_row.pack(fill=tk.X, pady=(0, 6))

        bg_thumb_col = tk.Frame(bg_thumb_row)
        bg_thumb_col.pack(side=tk.LEFT, anchor=tk.N)
        self._bg_thumb_container = tk.Frame(bg_thumb_col, width=_BG_THUMB, height=_BG_THUMB,
                                            relief="sunken", bd=1)
        self._bg_thumb_container.pack()
        self._bg_thumb_container.pack_propagate(False)
        self._bg_thumb_lbl = tk.Label(self._bg_thumb_container, text="No image",
                                      fg="gray", compound=tk.CENTER)
        self._bg_thumb_lbl.pack(fill=tk.BOTH, expand=True)
        tk.Label(bg_thumb_col, textvariable=self._tl_bg_size_var,
                 fg="gray", font=("Arial", 8)).pack(pady=(3, 0))

        # Canvas height hint
        self._tl_canvas_h_var = tk.StringVar(value="")
        tk.Label(bg_thumb_col, textvariable=self._tl_canvas_h_var,
                 fg="gray", font=("Arial", 8), anchor=tk.W).pack(pady=(2, 0))

        bg_btn_frame = tk.Frame(bg_thumb_row)
        bg_btn_frame.pack(side=tk.LEFT, padx=(10, 0), anchor=tk.N)
        self._bg_btn_choose  = tk.Button(bg_btn_frame, text="Choose...",          width=16)
        self._bg_btn_clear   = tk.Button(bg_btn_frame, text="Clear",              width=16)
        self._bg_btn_preview = tk.Button(bg_btn_frame, text="Preview",            width=16)
        self._bg_btn_editor  = tk.Button(bg_btn_frame, text="Open in Editor",     width=16)
        self._bg_btn_reload  = tk.Button(bg_btn_frame, text="Reload from Editor", width=16,
                                         state=tk.DISABLED)
        self._bg_btn_choose.pack(pady=(0, 4))
        self._bg_btn_clear.pack(pady=(0, 4))
        self._bg_btn_preview.pack(pady=(0, 4))
        self._bg_btn_editor.pack(pady=(0, 4))
        self._bg_btn_reload.pack()

        def _bg_sync_thumb(event=None):
            h = bg_btn_frame.winfo_reqheight()
            if h > 10:
                self._tl_bg_thumb_size[0] = h
                self._bg_thumb_container.config(width=h, height=h)
                _bg_render_thumb()
        bg_btn_frame.bind("<Configure>", _bg_sync_thumb)

        def _bg_render_thumb():
            if not self._tl_bg_img_bytes[0]:
                return
            box = self._tl_bg_thumb_size[0]
            try:
                img   = Image.open(io.BytesIO(self._tl_bg_img_bytes[0]))
                scale = box / max(img.width, img.height)
                new_w = max(1, int(img.width  * scale))
                new_h = max(1, int(img.height * scale))
                thumb = img.resize((new_w, new_h), Image.LANCZOS)
                photo = ImageTk.PhotoImage(thumb)
                self._tl_bg_thumb_ref[0] = photo
                self._bg_thumb_lbl.config(image=photo, text="")
            except Exception:
                pass

        def _bg_set_image(data, name):
            self._tl_bg_img_bytes[0] = data
            self._tl_bg_img_name[0]  = name
            if data:
                try:
                    img = Image.open(io.BytesIO(data))
                    self._tl_bg_size_var.set(f"{img.width} \u00d7 {img.height} px")
                    _bg_render_thumb()
                except Exception:
                    self._tl_bg_size_var.set("")
                    self._tl_bg_thumb_ref[0] = None
                    self._bg_thumb_lbl.config(image="", text="(error)", fg="red")
            else:
                self._tl_bg_size_var.set("")
                self._tl_bg_thumb_ref[0] = None
                self._bg_thumb_lbl.config(image="", text="No image", fg="gray")

        self._bg_set_image = _bg_set_image

        def _bg_preview():
            if not self._tl_bg_img_bytes[0]:
                return
            try:
                img = Image.open(io.BytesIO(self._tl_bg_img_bytes[0]))
            except Exception:
                return
            pw, ph = img.width, img.height
            pwin = tk.Toplevel(self.win)
            pwin.title(self._tl_bg_img_name[0] or "Background Image Preview")
            pwin.resizable(True, True)
            sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
            win_w  = min(pw + 20, int(sw * 0.8))
            win_h  = min(ph + 20, int(sh * 0.8))
            pwin.geometry(f"{win_w}x{win_h}+{(sw-win_w)//2}+{(sh-win_h)//2}")
            pwin.columnconfigure(0, weight=1)
            pwin.rowconfigure(0, weight=1)
            hbar = tk.Scrollbar(pwin, orient=tk.HORIZONTAL)
            vbar = tk.Scrollbar(pwin, orient=tk.VERTICAL)
            pcanvas = tk.Canvas(pwin, xscrollcommand=hbar.set,
                                yscrollcommand=vbar.set, bg="#888888")
            hbar.config(command=pcanvas.xview)
            vbar.config(command=pcanvas.yview)
            hbar.grid(row=1, column=0, sticky="ew")
            vbar.grid(row=0, column=1, sticky="ns")
            pcanvas.grid(row=0, column=0, sticky="nsew")
            photo = ImageTk.PhotoImage(img)
            pwin._preview_photo = photo
            pcanvas.create_image(0, 0, anchor=tk.NW, image=photo)
            pcanvas.config(scrollregion=(0, 0, pw, ph))

        def _bg_open_editor():
            import os, tempfile
            if not self._tl_bg_img_bytes[0]:
                messagebox.showwarning("No Image", "Load an image first.", parent=self.win)
                return
            try:
                img = Image.open(io.BytesIO(self._tl_bg_img_bytes[0]))
                tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                img.save(tmp.name, "PNG")
                tmp.close()
                self._tl_bg_tmp_path[0] = tmp.name
                os.startfile(tmp.name)
                self._bg_btn_reload.config(state=tk.NORMAL)
            except Exception as ex:
                messagebox.showerror("Error", str(ex), parent=self.win)

        def _bg_reload_editor():
            import os
            path = self._tl_bg_tmp_path[0]
            if not path or not os.path.exists(path):
                messagebox.showwarning("Reload",
                    "No temp file found. Use 'Open in Editor' first.", parent=self.win)
                return
            try:
                with open(path, "rb") as fh:
                    data = fh.read()
                _bg_set_image(data, self._tl_bg_img_name[0] or os.path.basename(path))
            except Exception as ex:
                messagebox.showerror("Error", str(ex), parent=self.win)

        def _bg_choose():
            from tkinter import filedialog
            import os
            path = filedialog.askopenfilename(
                title="Choose Background Image",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.tiff"),
                           ("All files", "*.*")],
                parent=self.win
            )
            if path:
                with open(path, "rb") as fh:
                    _bg_set_image(fh.read(), os.path.basename(path))

        self._bg_btn_choose.config(command=_bg_choose)
        self._bg_btn_clear.config(command=lambda: _bg_set_image(None, None))
        self._bg_btn_preview.config(command=_bg_preview)
        self._bg_btn_editor.config(command=_bg_open_editor)
        self._bg_btn_reload.config(command=_bg_reload_editor)

        # Bottom: position + canvas height
        bg_pos_frame = tk.Frame(bg_img_panel)
        bg_pos_frame.pack(fill=tk.X)
        tk.Label(bg_pos_frame, text="Position:").pack(side=tk.LEFT)
        ttk.Combobox(bg_pos_frame, textvariable=self._tl_bg_pos_var,
                     values=["Top", "Bottom", "Tile"], width=8,
                     state="readonly").pack(side=tk.LEFT, padx=(6, 0))

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_frame = tk.Frame(self.win)
        btn_frame.grid(row=8, column=0, pady=6)

        # Row 1: New / Save / Delete
        btn_row = tk.Frame(btn_frame)
        btn_row.pack(pady=2)
        tk.Button(btn_row, text="New",     command=self._new,    width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_row, text="Save",    command=self._save,   width=10).pack(side=tk.LEFT, padx=4)
        self.btn_delete = tk.Button(btn_row, text="Delete", command=self._delete, width=10)
        self.btn_delete.pack(side=tk.LEFT, padx=4)

        # Row 2: Save As / Import / Export
        io_row = tk.Frame(btn_frame)
        io_row.pack(pady=2)
        tk.Button(io_row, text="Save As", command=self._save_as,            width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(io_row, text="Import",  command=self._import_from_excel,  width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(io_row, text="Export",  command=self._export_to_excel,    width=10).pack(side=tk.LEFT, padx=4)

        # Row 3: Close — full width
        close_row = tk.Frame(btn_frame)
        close_row.pack(fill=tk.X, pady=(2, 0), padx=4)
        tk.Button(close_row, text="Close", command=self.win.destroy).pack(fill=tk.X)

        self.listbox.bind("<<ListboxSelect>>", self._on_select)
        self._refresh_list()

        self.win.update_idletasks()
        sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
        w, h = self.win.winfo_reqwidth(), self.win.winfo_reqheight()
        self.win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _refresh_list(self, select_id=None):
        timelines = self.db.load_timelines()
        self._timeline_ids = [t[0] for t in timelines]
        self.listbox.delete(0, tk.END)
        for _, title in timelines:
            self.listbox.insert(tk.END, title)
        target = select_id if select_id is not None else self.db.active_timeline_id
        if target in self._timeline_ids:
            idx = self._timeline_ids.index(target)
            self.listbox.selection_set(idx)
            self.listbox.see(idx)
            self._on_select()
        self._update_btn_states()

    def _on_select(self, *_):
        sel = self.listbox.curselection()
        if sel:
            idx = sel[0]
            self._selected_id = self._timeline_ids[idx]
            self.id_var.set(str(self._selected_id))
            self.name_var.set(self.listbox.get(idx))
            state = self.db.load_timeline_view_state(self._selected_id)
            if state:
                disp_val, disp_unit = _px_per_year_to_display(state["px_per_year"])
                self.tl_scale_var.set(f"{disp_val:.3g}")
                self.tl_scale_unit_var.set(disp_unit)
                self.tl_freeze_var.set(state["freeze_scale"])
            else:
                self.tl_scale_var.set("")
                self.tl_scale_unit_var.set("px/yr")
                self.tl_freeze_var.set(False)
            ruler_min, ruler_max, ruler_max_present = self.db.load_timeline_ruler(self._selected_id)
            rmin_num, rmin_unit, rmin_month, rmin_day = _date_val_to_components(ruler_min)
            self.tl_ruler_min_num_var.set(rmin_num)
            self.tl_ruler_min_unit_var.set(rmin_unit)
            self.tl_ruler_min_month_var.set(rmin_month)
            self.tl_ruler_min_day_var.set(rmin_day)
            if ruler_max_present:
                self.tl_ruler_max_unit_var.set("Present")  # trace hides num/month/day
            else:
                rmax_num, rmax_unit, rmax_month, rmax_day = _date_val_to_components(ruler_max)
                self.tl_ruler_max_num_var.set(rmax_num)
                self.tl_ruler_max_unit_var.set(rmax_unit)
                self.tl_ruler_max_month_var.set(rmax_month)
                self.tl_ruler_max_day_var.set(rmax_day)
            # icon_short, icon_long = self.db.load_timeline_icons(self._selected_id)
            # self.tl_icon_short_var.set(icon_short)
            # self.tl_icon_long_var.set(icon_long)
            self.tl_cat_header_var.set(
                self.db.load_timeline_cat_header_style(self._selected_id))
            self.tl_cat_header_title_pos_var.set(
                self.db.load_timeline_cat_header_title_pos(self._selected_id))
            saved_bg = self.db.load_timeline_canvas_bg(self._selected_id) or ""
            self._apply_canvas_bg(saved_bg)
            bg_info = self.db.load_timeline_bg_image(self._selected_id)
            self._bg_set_image(bg_info["image"], bg_info["name"])
            self._tl_bg_pos_var.set(bg_info["pos"])
            # Show current canvas content height
            if self.view and getattr(self.view, "_row_heights", None):
                ch = self.view._total_height()
                self._tl_canvas_h_var.set(f"Canvas height: {ch} px")
            else:
                self._tl_canvas_h_var.set("")
        else:
            self._selected_id = None
            self.id_var.set("")
            self.name_var.set("")
            self.tl_scale_var.set("")
            self.tl_scale_unit_var.set("px/yr")
            self.tl_freeze_var.set(False)
            self.tl_ruler_min_num_var.set("")
            self.tl_ruler_min_unit_var.set("CE")
            self.tl_ruler_min_month_var.set("")
            self.tl_ruler_min_day_var.set("")
            self.tl_ruler_max_num_var.set("")
            self.tl_ruler_max_unit_var.set("CE")
            self.tl_ruler_max_month_var.set("")
            self.tl_ruler_max_day_var.set("")
            # self.tl_icon_short_var.set("Diamond")
            # self.tl_icon_long_var.set("Box")
            self.tl_cat_header_var.set("Left")
            self.tl_cat_header_title_pos_var.set("Center (View)")
            self._apply_canvas_bg("")
            self._bg_set_image(None, None)
            self._tl_bg_pos_var.set("Top")
            self._tl_canvas_h_var.set("")
        self._update_btn_states()

    def _update_btn_states(self):
        can_delete = (self._selected_id is not None
                      and len(self._timeline_ids) > 1)
        self.btn_delete.config(state=tk.NORMAL if can_delete else tk.DISABLED)

    def _new(self):
        self.listbox.selection_clear(0, tk.END)
        self._selected_id = None
        self.id_var.set("0")
        self.name_var.set("")
        self.tl_scale_var.set("")
        self.tl_scale_unit_var.set("px/yr")
        self.tl_freeze_var.set(False)
        self.tl_ruler_min_num_var.set("")
        self.tl_ruler_min_unit_var.set("CE")
        self.tl_ruler_min_month_var.set("")
        self.tl_ruler_min_day_var.set("")
        self.tl_ruler_max_num_var.set("")
        self.tl_ruler_max_unit_var.set("CE")
        self.tl_ruler_max_month_var.set("")
        self.tl_ruler_max_day_var.set("")
        # self.tl_icon_short_var.set("Diamond")
        # self.tl_icon_long_var.set("Box")
        self.tl_cat_header_var.set("Left")
        self.tl_cat_header_title_pos_var.set("Left")
        self._apply_canvas_bg("")
        self._bg_set_image(None, None)
        self._tl_bg_pos_var.set("Top")
        self._tl_canvas_h_var.set("")
        self._update_btn_states()

    def _save(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Save", "Name cannot be empty.", parent=self.win)
            return
        # Parse scale value — blank means clear saved state
        scale_str = self.tl_scale_var.get().strip()
        try:
            scale_val = None
            if scale_str:
                disp_val = float(scale_str)
                if disp_val <= 0:
                    raise ValueError
                scale_val = _display_to_px_per_year(disp_val, self.tl_scale_unit_var.get())
        except ValueError:
            messagebox.showwarning("Save", "Scale must be a positive number (or blank to clear).",
                                   parent=self.win)
            return
        # Parse ruler start/end — blank num means clear (auto-calculate)
        ruler_min = ruler_max = None
        ruler_max_is_present = (self.tl_ruler_max_unit_var.get() == "Present")
        rmin_num = self.tl_ruler_min_num_var.get().strip()
        rmax_num = self.tl_ruler_max_num_var.get().strip()
        if rmin_num:
            try:
                rmin_month_name = self.tl_ruler_min_month_var.get()
                rmin_month = MONTHS.index(rmin_month_name) if rmin_month_name in MONTHS else 0
                rmin_day_str = self.tl_ruler_min_day_var.get()
                rmin_day = int(rmin_day_str) if rmin_day_str else 0
                ruler_min = _date_value(rmin_num, self.tl_ruler_min_unit_var.get(), rmin_month, rmin_day)
                if ruler_min is None:
                    raise ValueError
            except (ValueError, ZeroDivisionError):
                messagebox.showwarning("Save", "Ruler Start must be a positive number.", parent=self.win)
                return
        if not ruler_max_is_present and rmax_num:
            try:
                rmax_month_name = self.tl_ruler_max_month_var.get()
                rmax_month = MONTHS.index(rmax_month_name) if rmax_month_name in MONTHS else 0
                rmax_day_str = self.tl_ruler_max_day_var.get()
                rmax_day = int(rmax_day_str) if rmax_day_str else 0
                ruler_max = _date_value(rmax_num, self.tl_ruler_max_unit_var.get(), rmax_month, rmax_day)
                if ruler_max is None:
                    raise ValueError
            except (ValueError, ZeroDivisionError):
                messagebox.showwarning("Save", "Ruler End must be a positive number.", parent=self.win)
                return
        # Effective end for comparisons (Present = today)
        eff_ruler_max = _today_value() if ruler_max_is_present else ruler_max
        if ruler_min is not None and eff_ruler_max is not None and ruler_min >= eff_ruler_max:
            messagebox.showwarning("Save",
                "Ruler Start must be earlier than Ruler End.", parent=self.win)
            return

        # Warn if events fall outside the ruler range (existing timelines only)
        if self._selected_id and (ruler_min is not None or eff_ruler_max is not None):
            with sqlite3.connect(self.db.db_file) as _conn:
                outside_conditions = []
                outside_params = [self._selected_id]
                if ruler_min is not None:
                    outside_conditions.append("(start_value IS NOT NULL AND start_value < ?)")
                    outside_params.append(ruler_min)
                if eff_ruler_max is not None:
                    outside_conditions.append("(start_value IS NOT NULL AND start_value > ?)")
                    outside_params.append(eff_ruler_max)
                outside_count = _conn.execute(
                    "SELECT COUNT(*) FROM events WHERE timelineid=? AND hidden=0 AND ("
                    + " OR ".join(outside_conditions) + ")",
                    outside_params
                ).fetchone()[0]
            if outside_count:
                plural = "s" if outside_count != 1 else ""
                if not messagebox.askyesno(
                    "Events Outside Ruler Range",
                    f"{outside_count} visible event{plural} start outside the ruler range "
                    f"and will not appear on the timeline.\n\nSave anyway?",
                    parent=self.win
                ):
                    return

        if self.id_var.get() in ("0", ""):
            new_id = self.db.add_timeline(name)
            self.db.active_timeline_id = new_id
            self.db.load()
            self.db.load_categories()
            if scale_val is not None:
                self.db.save_timeline_view_state(new_id, scale_val, self.tl_freeze_var.get())
            self.db.save_timeline_ruler(new_id, ruler_min, ruler_max, ruler_max_is_present)
            self.db.save_timeline_cat_header_style(new_id, self.tl_cat_header_var.get())
            self.db.save_timeline_cat_header_title_pos(new_id, self.tl_cat_header_title_pos_var.get())
            self.db.save_timeline_canvas_bg(new_id, self.tl_canvas_bg_var.get())
            self.db.save_timeline_bg_image(new_id, self._tl_bg_img_bytes[0],
                                           self._tl_bg_img_name[0],
                                           self._tl_bg_pos_var.get())
            # self.db.save_timeline_icons(new_id,
            #                             self.tl_icon_short_var.get(),
            #                             self.tl_icon_long_var.get())
            if self.on_change:
                self.on_change(new_id)
            self._refresh_list(select_id=new_id)
        else:
            self.db.rename_timeline(self._selected_id, name)
            if scale_val is not None:
                self.db.save_timeline_view_state(self._selected_id, scale_val, self.tl_freeze_var.get())
            else:
                self.db.save_timeline_view_state(self._selected_id, None, self.tl_freeze_var.get())
            self.db.save_timeline_ruler(self._selected_id, ruler_min, ruler_max, ruler_max_is_present)
            self.db.save_timeline_cat_header_style(self._selected_id, self.tl_cat_header_var.get())
            self.db.save_timeline_cat_header_title_pos(self._selected_id, self.tl_cat_header_title_pos_var.get())
            self.db.save_timeline_canvas_bg(self._selected_id, self.tl_canvas_bg_var.get())
            self.db.save_timeline_bg_image(self._selected_id, self._tl_bg_img_bytes[0],
                                           self._tl_bg_img_name[0],
                                           self._tl_bg_pos_var.get())
            # self.db.save_timeline_icons(self._selected_id,
            #                             self.tl_icon_short_var.get(),
            #                             self.tl_icon_long_var.get())
            if self.on_change:
                self.on_change(self.db.active_timeline_id)
            self._refresh_list(select_id=self._selected_id)

    def _delete(self):
        if self._selected_id is None:
            return
        if len(self._timeline_ids) <= 1:
            messagebox.showwarning("Delete", "Cannot delete the only timeline.", parent=self.win)
            return
        count = self.db.count_events(self._selected_id)
        event_str = f"{count} event{'s' if count != 1 else ''}"
        if not messagebox.askyesno(
            "Delete Timeline",
            f"Delete \"{self.name_var.get()}\"?\n\n"
            f"This will permanently delete {event_str} and all categories. "
            f"This cannot be undone.",
            parent=self.win,
        ):
            return
        self.db.delete_timeline(self._selected_id)
        remaining = self.db.load_timelines()
        new_id = remaining[0][0]
        self.db.active_timeline_id = new_id
        self.db.load()
        self.db.load_categories()
        if self.on_change:
            self.on_change(new_id)
        self._refresh_list(select_id=new_id)

    def _save_as(self):
        """Copy the selected timeline to a new timeline with a user-supplied name."""
        if self._selected_id is None:
            messagebox.showwarning("Save As", "Please select a timeline to copy.", parent=self.win)
            return

        # Capture src_id NOW — before askstring opens its own event loop,
        # which causes the listbox to lose focus and fire <<ListboxSelect>>
        # with an empty selection, resetting self._selected_id to None.
        src_id = self._selected_id

        new_name = simpledialog.askstring(
            "Save As",
            "Name for the new timeline copy:",
            initialvalue=f"{self.name_var.get()} (copy)",
            parent=self.win,
        )
        if not new_name or not new_name.strip():
            return
        new_name = new_name.strip()

        # ── Step 1: read all source data before writing anything ──────────────
        with sqlite3.connect(self.db.db_file) as rconn:
            src_cats = rconn.execute(
                "SELECT CategoryID, Title, parent_id, sort_order, hidden "
                "FROM Category WHERE timelineid=? ORDER BY CategoryID",
                (src_id,)
            ).fetchall()

            src_events = rconn.execute(
                "SELECT title, year, desc, categoryid, url, "
                "image, image_name, image_type, "
                "start_value, start_display, start_unit, start_month, start_day, "
                "end_value,   end_display,   end_unit,   end_month,   end_day,   "
                "standalone, sort_order, hidden, picture_position "
                "FROM events WHERE timelineid=?",
                (src_id,)
            ).fetchall()

        # ── Step 2: write new timeline, categories and events ─────────────────
        # Use isolation_level=None so Python's implicit transaction management
        # is disabled; we issue BEGIN/COMMIT ourselves as plain SQL to guarantee
        # reliable commit behaviour across all Python 3.x versions.
        conn = sqlite3.connect(self.db.db_file)
        conn.isolation_level = None
        try:
            conn.execute("BEGIN")

            cur = conn.execute("INSERT INTO Timeline (Title) VALUES (?)", (new_name,))
            new_id = cur.lastrowid

            # Copy categories, preserving hierarchy.
            # Parent categories always have a lower CategoryID than their children
            # (they were inserted first), so ORDER BY CategoryID guarantees parents
            # are in cat_id_map before their children are processed.
            cat_id_map = {}
            for old_cid, title, old_parent, sort_ord, hidden in src_cats:
                new_parent = cat_id_map.get(old_parent)   # None for root cats
                cur2 = conn.execute(
                    "INSERT INTO Category (Title, timelineid, parent_id, sort_order, hidden) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (title, new_id, new_parent, sort_ord, hidden)
                )
                cat_id_map[old_cid] = cur2.lastrowid

            for row in src_events:
                (ev_title, year, desc, old_catid, url,
                 image, image_name, image_type,
                 sv, sd, su, sm, sday,
                 ev, ed, eu, em, eday,
                 standalone, sort_ord, ev_hidden, pic_pos) = row
                new_catid = cat_id_map.get(old_catid)
                conn.execute(
                    "INSERT INTO events "
                    "(title, year, desc, categoryid, timelineid, url, "
                    "image, image_name, image_type, "
                    "start_value, start_display, start_unit, start_month, start_day, "
                    "end_value,   end_display,   end_unit,   end_month,   end_day,   "
                    "standalone, sort_order, hidden, picture_position) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (ev_title, year, desc, new_catid, new_id, url,
                     image, image_name, image_type,
                     sv, sd, su, sm, sday,
                     ev, ed, eu, em, eday,
                     standalone, sort_ord, ev_hidden, pic_pos)
                )

            conn.execute("COMMIT")
        except Exception as exc:
            try:
                conn.execute("ROLLBACK")
            except Exception:
                pass
            messagebox.showerror("Save As Failed", str(exc), parent=self.win)
            return
        finally:
            conn.close()

        # ── Step 3: reload in-memory state and refresh UI ─────────────────────
        self.db.active_timeline_id = new_id
        self.db.load()
        self.db.load_categories()
        if self.on_change:
            self.on_change(new_id)
        self._refresh_list(select_id=new_id)

    def _ensure_category_path(self, conn, path_str, timeline_id=None):
        """Resolve or create a category hierarchy from a ' > ' separated path string."""
        if timeline_id is None:
            timeline_id = self.db.active_timeline_id
        parts = [p.strip() for p in path_str.split(" > ") if p.strip()]
        if not parts:
            return self.db._ensure_category(conn, "General")
        parent_id = None
        cat_id = None
        for part in parts:
            if parent_id is None:
                row = conn.execute(
                    "SELECT CategoryID FROM Category WHERE LOWER(Title)=LOWER(?) "
                    "AND timelineid=? AND parent_id IS NULL",
                    (part, timeline_id)
                ).fetchone()
            else:
                row = conn.execute(
                    "SELECT CategoryID FROM Category WHERE LOWER(Title)=LOWER(?) "
                    "AND timelineid=? AND parent_id=?",
                    (part, timeline_id, parent_id)
                ).fetchone()
            if row:
                cat_id = row[0]
            else:
                cur = conn.execute(
                    "INSERT INTO Category (Title, timelineid, parent_id) VALUES (?, ?, ?)",
                    (part, timeline_id, parent_id)
                )
                cat_id = cur.lastrowid
            parent_id = cat_id
        return cat_id

    def _import_from_excel(self):
        # Capture the selected timeline ID NOW, before any dialog opens.
        # Dialogs run their own event loop, which can cause the listbox to lose
        # focus and fire <<ListboxSelect>> with an empty selection, resetting
        # self._selected_id to None (same issue as in _save_as).
        if self._selected_id is None:
            messagebox.showwarning("Import", "Please select a timeline to import into.",
                                   parent=self.win)
            return
        tid = self._selected_id

        # ── Ask Replace vs Append ─────────────────────────────────────────────
        # Build a small modal dialog with three explicit buttons.
        mode = tk.StringVar(value="")
        dlg = tk.Toplevel(self.win)
        dlg.title("Import — Replace or Append?")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.protocol("WM_DELETE_WINDOW", lambda: mode.set("cancel") or dlg.destroy())

        tk.Label(dlg, text="How should the imported data be added to\n"
                           f"\"{self.name_var.get()}\"?",
                 font=("Arial", 10), pady=8, padx=16, justify=tk.CENTER).pack()

        desc_frame = tk.Frame(dlg, padx=16)
        desc_frame.pack(fill=tk.X, pady=(0, 6))
        tk.Label(desc_frame, text="Replace:  delete all existing events and categories first.",
                 font=("Arial", 9), fg="#555555", anchor=tk.W).pack(fill=tk.X)
        tk.Label(desc_frame, text="Append:   add imported events alongside existing ones.",
                 font=("Arial", 9), fg="#555555", anchor=tk.W).pack(fill=tk.X)

        btn_frame = tk.Frame(dlg, pady=8)
        btn_frame.pack()
        tk.Button(btn_frame, text="Replace", width=10,
                  command=lambda: mode.set("replace") or dlg.destroy()).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="Append",  width=10,
                  command=lambda: mode.set("append")  or dlg.destroy()).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="Cancel",  width=10,
                  command=lambda: mode.set("cancel")  or dlg.destroy()).pack(side=tk.LEFT, padx=6)

        dlg.update_idletasks()
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        w, h = dlg.winfo_reqwidth(), dlg.winfo_reqheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.win.wait_window(dlg)

        if mode.get() in ("", "cancel"):
            return
        replace_mode = (mode.get() == "replace")

        # If replacing, confirm when there is existing data
        if replace_mode:
            with sqlite3.connect(self.db.db_file) as _chk:
                event_count = _chk.execute(
                    "SELECT COUNT(*) FROM events WHERE timelineid=?", (tid,)
                ).fetchone()[0]
                cat_count = _chk.execute(
                    "SELECT COUNT(*) FROM Category WHERE timelineid=?", (tid,)
                ).fetchone()[0]
            if event_count or cat_count:
                parts = []
                if event_count:
                    parts.append(f"{event_count} event{'s' if event_count != 1 else ''}")
                if cat_count:
                    parts.append(f"{cat_count} categor{'ies' if cat_count != 1 else 'y'}")
                if not messagebox.askyesno(
                    "Import — Replace Existing Data",
                    f"This will permanently delete {' and '.join(parts)} from "
                    f"\"{self.name_var.get()}\" before importing.\n\nContinue?",
                    parent=self.win,
                ):
                    return

        path = filedialog.askopenfilename(
            filetypes=[("Excel workbook", "*.xlsx")],
            title="Import from Excel",
            parent=self.win,
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            messagebox.showerror("Import Failed", f"Could not open file:\n{e}", parent=self.win)
            return

        if "Events" not in wb.sheetnames:
            messagebox.showwarning("Import", "No 'Events' sheet found in this file.", parent=self.win)
            return

        try:
            with sqlite3.connect(self.db.db_file) as conn:
                if replace_mode:
                    conn.execute("DELETE FROM events    WHERE timelineid=?", (tid,))
                    conn.execute("DELETE FROM Category  WHERE timelineid=?", (tid,))

                if "Categories" in wb.sheetnames:
                    for row in wb["Categories"].iter_rows(min_row=2, values_only=True):
                        if row and row[1]:
                            cid = self._ensure_category_path(conn, str(row[1]), timeline_id=tid)
                            hidden_cat = 1 if len(row) > 2 and row[2] and str(row[2]).strip() not in ("", "0") else 0
                            if hidden_cat:
                                conn.execute("UPDATE Category SET hidden=1 WHERE CategoryID=?", (cid,))

                all_tl_map = {r[1]: r[0] for r in conn.execute("SELECT TimelineID, Title FROM Timeline").fetchall()}

                count = 0
                for row in wb["Events"].iter_rows(min_row=2, values_only=True):
                    if not row or not row[1]:
                        continue
                    cat_str, title, start_disp, end_disp, desc, url, own_row, hidden_val, pic_pos, linked_cat, linked_tl = (list(row) + [None]*11)[:11]
                    title = str(title).strip() if title else ""
                    if not title:
                        continue

                    cat_id     = self._ensure_category_path(conn, str(cat_str) if cat_str else "General", timeline_id=tid)
                    standalone = 1 if own_row and str(own_row).strip() not in ("", "0") else 0
                    hidden_ev  = 1 if hidden_val and str(hidden_val).strip() not in ("", "0") else 0
                    pic_pos_val = str(pic_pos).strip() if pic_pos else None

                    linked_tl_id = None
                    if linked_tl and str(linked_tl).strip():
                        linked_tl_id = all_tl_map.get(str(linked_tl).strip())

                    linked_cat_id = None
                    if linked_cat and str(linked_cat).strip():
                        tl_for_cat = linked_tl_id or tid
                        parts = [p.strip() for p in str(linked_cat).split(" > ") if p.strip()]
                        par = None
                        for part in parts:
                            if par is None:
                                r2 = conn.execute(
                                    "SELECT CategoryID FROM Category WHERE LOWER(Title)=LOWER(?) "
                                    "AND timelineid=? AND parent_id IS NULL", (part, tl_for_cat)
                                ).fetchone()
                            else:
                                r2 = conn.execute(
                                    "SELECT CategoryID FROM Category WHERE LOWER(Title)=LOWER(?) "
                                    "AND timelineid=? AND parent_id=?", (part, tl_for_cat, par)
                                ).fetchone()
                            if r2:
                                linked_cat_id = r2[0]
                                par = linked_cat_id
                            else:
                                linked_cat_id = None
                                break

                    s_num, s_unit, s_month, s_day = _parse_date_display(start_disp)
                    e_num, e_unit, e_month, e_day = _parse_date_display(end_disp)

                    s_value   = _date_value(s_num, s_unit, s_month or 0, s_day or 0) if s_num else None
                    e_value   = _date_value(e_num, e_unit, e_month or 0, e_day or 0) if e_num else None
                    s_display = _date_display(s_num, s_unit, s_month or 0, s_day or 0) if s_num else ""
                    e_display = _date_display(e_num, e_unit, e_month or 0, e_day or 0) if e_num else ""
                    year      = int(s_value) if s_value is not None else 0

                    conn.execute(
                        "INSERT INTO events "
                        "(title, year, desc, categoryid, timelineid, url, "
                        "start_value, start_display, start_unit, start_month, start_day, "
                        "end_value,   end_display,   end_unit,   end_month,   end_day,   standalone, "
                        "hidden, picture_position, linked_categoryid, linked_timelineid) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (title, year, str(desc or ""), cat_id,
                         tid, str(url or ""),
                         s_value, s_display, s_unit, s_month, s_day,
                         e_value, e_display, e_unit, e_month, e_day, standalone,
                         hidden_ev, pic_pos_val, linked_cat_id, linked_tl_id)
                    )
                    count += 1

                # Fix sort_order for categories — ensure_category_path inserts without sort_order
                from itertools import groupby
                all_cats = conn.execute(
                    "SELECT CategoryID, parent_id FROM Category "
                    "WHERE timelineid=? ORDER BY parent_id, CategoryID",
                    (tid,)
                ).fetchall()
                for _, grp in groupby(all_cats, key=lambda r: r[1]):
                    for i, (cid, _) in enumerate(grp):
                        conn.execute("UPDATE Category SET sort_order=? WHERE CategoryID=?", (i, cid))

                # Fix sort_order for events.
                # In append mode start numbering new events after the existing max
                # so they don't collide with pre-existing sort orders.
                if replace_mode:
                    all_events = conn.execute(
                        "SELECT id, categoryid FROM events "
                        "WHERE timelineid=? ORDER BY categoryid, id",
                        (tid,)
                    ).fetchall()
                    for _, grp in groupby(all_events, key=lambda r: r[1]):
                        for i, (eid, _) in enumerate(grp):
                            conn.execute("UPDATE events SET sort_order=? WHERE id=?", (i, eid))
                else:
                    # Append: only fix the rows that have no sort_order yet (NULL)
                    null_events = conn.execute(
                        "SELECT id, categoryid FROM events "
                        "WHERE timelineid=? AND sort_order IS NULL ORDER BY categoryid, id",
                        (tid,)
                    ).fetchall()
                    for cat_id_grp, grp in groupby(null_events, key=lambda r: r[1]):
                        max_order = conn.execute(
                            "SELECT COALESCE(MAX(sort_order), -1) FROM events "
                            "WHERE timelineid=? AND categoryid=? AND sort_order IS NOT NULL",
                            (tid, cat_id_grp)
                        ).fetchone()[0]
                        for j, (eid, _) in enumerate(grp):
                            conn.execute("UPDATE events SET sort_order=? WHERE id=?",
                                         (max_order + 1 + j, eid))

        except Exception as e:
            messagebox.showerror("Import Failed", str(e), parent=self.win)
            return

        self.db.active_timeline_id = tid
        self.db.load()
        self.db.load_categories()
        if self.on_import:
            self.on_import()
        self._refresh_list(select_id=tid)
        action = "Replaced with" if replace_mode else "Appended"
        messagebox.showinfo("Import",
                            f"{action} {count} event{'s' if count != 1 else ''}.",
                            parent=self.win)

    def _export_to_excel(self):
        title = self.name_var.get() or "timeline"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=f"{title}.xlsx",
            title="Export to Excel",
            parent=self.win,
        )
        if not path:
            return

        # Build cross-timeline lookups for linked fields
        with sqlite3.connect(self.db.db_file) as _conn:
            _cat_rows = _conn.execute("SELECT CategoryID, Title, parent_id FROM Category").fetchall()
            _cat_info = {r[0]: (r[1], r[2]) for r in _cat_rows}
            def _build_cat_path(cid):
                if cid not in _cat_info:
                    return ""
                t, p = _cat_info[cid]
                return (_build_cat_path(p) + " > " + t) if p else t
            all_linked_cat_paths = {cid: _build_cat_path(cid) for cid in _cat_info}
            all_tl_titles = {r[0]: r[1] for r in _conn.execute("SELECT TimelineID, Title FROM Timeline").fetchall()}

        wb = openpyxl.Workbook()

        ws_events = wb.active
        ws_events.title = "Events"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="4A6FA5")
        center = Alignment(horizontal="center")

        event_headers = ["Category", "Title", "Start Date", "End Date", "Description", "URL",
                         "Own Row", "Hidden", "Picture Position", "Linked Category", "Linked Timeline"]
        for col, text in enumerate(event_headers, 1):
            cell = ws_events.cell(row=1, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        cat_path = {n["title"]: n["path"] for n in self.db.cat_nodes}

        for row_idx, e in enumerate(self.db.events, 2):
            cat = e.get("category") or "general"
            linked_cid = e.get("linked_categoryid")
            linked_tid = e.get("linked_timelineid")
            ws_events.cell(row=row_idx, column=1,  value=cat_path.get(cat, cat))
            ws_events.cell(row=row_idx, column=2,  value=e.get("title", ""))
            ws_events.cell(row=row_idx, column=3,  value=e.get("start_display") or "")
            ws_events.cell(row=row_idx, column=4,  value=e.get("end_display") or "")
            ws_events.cell(row=row_idx, column=5,  value=e.get("desc") or "")
            ws_events.cell(row=row_idx, column=6,  value=e.get("url") or "")
            ws_events.cell(row=row_idx, column=7,  value=1 if e.get("standalone") else 0)
            ws_events.cell(row=row_idx, column=8,  value=1 if e.get("hidden") else 0)
            ws_events.cell(row=row_idx, column=9,  value=e.get("picture_position") or "")
            ws_events.cell(row=row_idx, column=10, value=all_linked_cat_paths.get(linked_cid, "") if linked_cid else "")
            ws_events.cell(row=row_idx, column=11, value=all_tl_titles.get(linked_tid, "") if linked_tid else "")

        ws_events.column_dimensions["A"].width = 25
        ws_events.column_dimensions["B"].width = 35
        ws_events.column_dimensions["C"].width = 18
        ws_events.column_dimensions["D"].width = 18
        ws_events.column_dimensions["E"].width = 50
        ws_events.column_dimensions["F"].width = 40
        ws_events.column_dimensions["G"].width = 10
        ws_events.column_dimensions["H"].width = 8
        ws_events.column_dimensions["I"].width = 18
        ws_events.column_dimensions["J"].width = 30
        ws_events.column_dimensions["K"].width = 25

        ws_cats = wb.create_sheet("Categories")

        cat_headers = ["Category", "Full Path", "Hidden"]
        for col, text in enumerate(cat_headers, 1):
            cell = ws_cats.cell(row=1, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, node in enumerate(self.db.cat_nodes, 2):
            ws_cats.cell(row=row_idx, column=1, value=node["title"])
            ws_cats.cell(row=row_idx, column=2, value=node["path"])
            ws_cats.cell(row=row_idx, column=3, value=1 if node.get("hidden") else 0)

        ws_cats.column_dimensions["A"].width = 25
        ws_cats.column_dimensions["B"].width = 50
        ws_cats.column_dimensions["C"].width = 8

        try:
            wb.save(path)
            messagebox.showinfo("Export", f"Exported to {os.path.basename(path)}", parent=self.win)
        except Exception as e:
            messagebox.showerror("Export Failed", str(e), parent=self.win)



if __name__ == "__main__":
    root = tk.Tk()
    db = TimelineDB()
    timelines = db.load_timelines()
    if not timelines:
        db.add_timeline("My Timeline")
        timelines = db.load_timelines()
    last_id = db.last_timeline_id()
    match = next((t for t in timelines if t[0] == last_id), None) or timelines[0]
    db.active_timeline_id = match[0]
    db.load()
    TimelineView(root, db)
    root.mainloop()
